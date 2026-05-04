"""Build the Executive Deliverable sheet from a Pattern competitive landscape workbook.

Usage:
    py build_deliverable.py --source <input.xlsx> [--output <out.xlsx>] [--sort-by priority|fit|moat|source]

Reads the source `Competitive Landscape` sheet, synthesizes ~108 executive rows per company,
applies Pattern brand styling, writes an `Executive Deliverable` sheet alongside the original.

Requires: openpyxl
"""
from __future__ import annotations
import argparse
import re
import sys
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================
# Pattern brand colors
# ============================================================
NAVY = "0F4761"
BRIGHT_BLUE = "4280F4"
PURPLE_BLUE = "3A00FD"
WHITE = "FFFFFF"
BLACK = "000000"
ALT_FILL = "F5F8FF"
BORDER_GREY = "DDDDDD"
SECTION_FILL = "F5F8FF"

VERDICT_FILL = {"ACQUIRE": ("E2EFDA", "1A5C1A"), "MONITOR": ("FFF3CD", "7A5C00"), "PASS": ("FFE0E0", "8B1A1A")}

FONT_BODY = "Wix Madefor Display"
FONT_BOLD = "Wix Madefor Display SemiBold"

BANNED_FILLER = {"leading", "robust", "innovative", "comprehensive", "cutting-edge", "best-in-class"}

# ============================================================
# Output row spec — (section, label, [source_field, ...])
# ============================================================
ROWS: list[tuple[str, str, list[str]]] = [
    # Market taxonomy
    ("Market taxonomy", "Value chain stage", ["Commerce Value Chain Stage"]),
    ("Market taxonomy", "Category and subcategory", ["Primary Category", "Sub-Category"]),
    ("Market taxonomy", "Workflow layer", ["Workflow Layer", "Primary Use Case"]),
    ("Market taxonomy", "Buyer and ICP focus", ["Buyer Segment", "Primary ICP"]),
    ("Market taxonomy", "Secondary buyer segment", ["Secondary ICP", "Buyer Segment"]),
    ("Market taxonomy", "Customer scale focus", ["Customer Scale Focus", "Enterprise Customer Count"]),
    ("Market taxonomy", "Geography served", ["Geography Served", "HQ Country"]),
    ("Market taxonomy", "Platform ecosystem", ["Platform Ecosystem", "Native Platform Integrations"]),
    ("Market taxonomy", "Channel focus", ["Channel Focus", "Sales Motion", "Partner Motion"]),
    ("Market taxonomy", "Category maturity", ["Category Maturity", "Category Leadership Position"]),
    ("Market taxonomy", "Market tailwinds", ["Market Tailwind Tags", "White-Space Relevance"]),
    # Product capability
    ("Product capability", "Core product wedge", ["Core Product Description", "Primary Use Case"]),
    ("Product capability", "Primary use case", ["Primary Use Case", "Workflow Criticality"]),
    ("Product capability", "Key feature set", ["Key Features", "Core Product Description"]),
    ("Product capability", "Differentiated capability", ["Differentiated Features", "Competitive Differentiation"]),
    ("Product capability", "AI and agent posture", ["AI Capability", "Agentic Capability"]),
    ("Product capability", "API and protocol maturity", ["Agent-Callable API Availability", "Protocol Support", "API Documentation Quality"]),
    ("Product capability", "Integration breadth", ["Integration Breadth", "Native Platform Integrations"]),
    ("Product capability", "Commerce platform depth", ["Native Platform Integrations", "Platform Ecosystem"]),
    ("Product capability", "Enterprise systems depth", ["ERP / OMS / WMS Integrations", "Implementation Model"]),
    ("Product capability", "Payments and carrier depth", ["Payment / PSP Integrations", "Carrier Integrations"]),
    ("Product capability", "Implementation model", ["Implementation Model", "Estimated Time to Value"]),
    ("Product capability", "Time to value", ["Estimated Time to Value", "Implementation Model"]),
    ("Product capability", "Workflow criticality", ["Workflow Criticality", "Workflow Criticality if Removed"]),
    ("Product capability", "System-of-record role", ["System of Record Status", "Owns Operational Records"]),
    ("Product capability", "Operational record ownership", ["Owns Operational Records", "Data Captured"]),
    ("Product capability", "Switching cost driver", ["Switching Cost Driver", "Switching Cost Score"]),
    ("Product capability", "Removal impact", ["Workflow Criticality if Removed", "Replacement Risk"]),
    # Data and compliance
    ("Data and compliance", "Data captured", ["Data Captured", "Owns Operational Records"]),
    ("Data and compliance", "PII and sensitivity exposure", ["PII Exposure", "Data Sensitivity Tier"]),
    ("Data and compliance", "PCI and payment exposure", ["PCI Exposure", "Payment / PSP Integrations"]),
    ("Data and compliance", "Security posture", ["SOC 2 Status", "Security Risk Level"]),
    ("Data and compliance", "Privacy readiness", ["GDPR / CCPA Readiness", "Data Retention Policy"]),
    ("Data and compliance", "AI and product compliance", ["EU AI Act Exposure", "ESPR Exposure"]),
    ("Data and compliance", "Cross-retailer data use", ["Cross-Retailer Data Usage", "Data Network Effect"]),
    ("Data and compliance", "Data network effect", ["Data Network Effect", "Data Network Score"]),
    ("Data and compliance", "Compliance risk notes", ["Compliance Risk Notes", "Security Risk Level"]),
    ("Data and compliance", "Data retention posture", ["Data Retention Policy", "Compliance Risk Notes"]),
    # Commercial traction
    ("Commercial traction", "Revenue scale", ["Estimated ARR / Revenue", "Estimated Enterprise Value"]),
    ("Commercial traction", "Growth trajectory", ["Revenue Growth Rate", "Hiring Velocity"]),
    ("Commercial traction", "Customer proof", ["Customer Count", "Notable Customers"]),
    ("Commercial traction", "Enterprise penetration", ["Enterprise Customer Count", "Primary ICP"]),
    ("Commercial traction", "Notable customer signal", ["Notable Customers", "Customer Scale Focus"]),
    ("Commercial traction", "Transaction volume signal", ["GMV / Order / Shipment Volume", "Customer Count"]),
    ("Commercial traction", "Retention quality", ["Retention / NRR", "Gross Retention"]),
    ("Commercial traction", "Churn risk", ["Churn Risk", "Gross Retention"]),
    ("Commercial traction", "ACV range", ["ACV Range", "Pricing Model"]),
    ("Commercial traction", "Pricing model", ["Pricing Model", "Usage-Based Pricing Exposure"]),
    ("Commercial traction", "Usage pricing exposure", ["Usage-Based Pricing Exposure", "GMV / Order / Shipment Volume"]),
    ("Commercial traction", "Margin profile", ["Gross Margin Estimate", "Implementation Model"]),
    ("Commercial traction", "Sales motion", ["Sales Motion", "Sales Cycle Length"]),
    ("Commercial traction", "Sales cycle", ["Sales Cycle Length", "ACV Range"]),
    ("Commercial traction", "Partner motion", ["Partner Motion", "Channel Focus"]),
    ("Commercial traction", "Marketplace presence", ["App Store / Marketplace Presence", "Platform Ecosystem"]),
    # Competitive position
    ("Competitive position", "Direct competitor set", ["Direct Competitors", "Primary Category"]),
    ("Competitive position", "Substitute pressure", ["Substitute Competitors", "Replacement Risk"]),
    ("Competitive position", "Platform-native threat", ["Platform-Native Threats", "Platform Lock-In Score"]),
    ("Competitive position", "Differentiation angle", ["Competitive Differentiation", "Differentiated Features"]),
    ("Competitive position", "Competitive weakness", ["Competitive Weakness", "Red Flags"]),
    ("Competitive position", "Moat type", ["Moat Type", "Total Moat Score"]),
    ("Competitive position", "Data network strength", ["Data Network Score", "Data Network Effect"]),
    ("Competitive position", "Switching cost strength", ["Switching Cost Score", "Switching Cost Driver"]),
    ("Competitive position", "Platform lock-in strength", ["Platform Lock-In Score", "Native Platform Integrations"]),
    ("Competitive position", "Regulatory or liability moat", ["Regulatory / Liability Score", "Compliance Risk Notes"]),
    ("Competitive position", "Physical network moat", ["Physical Network Score", "Fit with Pattern Fulfillment Network"]),
    ("Competitive position", "Overall moat score", ["Total Moat Score", "Moat Durability"]),
    ("Competitive position", "Moat durability", ["Moat Durability", "Replacement Risk"]),
    ("Competitive position", "Replacement risk", ["Replacement Risk", "Substitute Competitors", "Platform-Native Threats"]),
    ("Competitive position", "Category leadership", ["Category Leadership Position", "Category Maturity"]),
    ("Competitive position", "White-space relevance", ["White-Space Relevance", "Market Tailwind Tags"]),
    # Pattern fit
    ("Pattern fit", "Strategic relevance", ["Pattern Strategic Relevance", "Strategic Fit Score"]),
    ("Pattern fit", "Capability fit", ["Fit with Pattern Capabilities", "New Capability Unlocked"]),
    ("Pattern fit", "Customer base fit", ["Fit with Pattern Customer Base", "Cross-Sell Potential"]),
    ("Pattern fit", "Data asset fit", ["Fit with Pattern Data Assets", "Data Captured"]),
    ("Pattern fit", "Fulfillment network fit", ["Fit with Pattern Fulfillment Network", "Carrier Integrations"]),
    ("Pattern fit", "AI strategy fit", ["Fit with Pattern AI / Agent Strategy", "AI Capability"]),
    ("Pattern fit", "Cross-sell potential", ["Cross-Sell Potential", "Revenue Synergy Potential"]),
    ("Pattern fit", "Product integration potential", ["Product Integration Potential", "Integration Breadth"]),
    ("Pattern fit", "Revenue synergy potential", ["Revenue Synergy Potential", "Fit with Pattern Customer Base"]),
    ("Pattern fit", "Cost synergy potential", ["Cost Synergy Potential", "Implementation Model"]),
    ("Pattern fit", "Defensive value", ["Defensive Value", "Platform-Native Threats"]),
    ("Pattern fit", "New capability unlocked", ["New Capability Unlocked", "Fit with Pattern Capabilities"]),
    # M&A lens
    ("M&A lens", "Acquisition thesis", ["Acquisition Thesis", "Pattern Strategic Relevance"]),
    ("M&A lens", "Build-buy-partner stance", ["Build / Buy / Partner Recommendation", "Deal Priority"]),
    ("M&A lens", "Deal priority", ["Deal Priority", "Strategic Fit Score"]),
    ("M&A lens", "Valuation signal", ["Estimated Enterprise Value", "Estimated Revenue Multiple"]),
    ("M&A lens", "Comparable transaction signal", ["Comparable Transactions", "Estimated Revenue Multiple"]),
    ("M&A lens", "Funding and investor context", ["Funding Raised", "Last Funding Round", "Key Investors"]),
    ("M&A lens", "Exit pressure", ["Investor Pressure / Exit Likelihood", "Seller Motivation"]),
    ("M&A lens", "Seller motivation", ["Seller Motivation", "Likely Process Timing"]),
    ("M&A lens", "Cap table complexity", ["Cap Table Complexity", "Key Investors"]),
    ("M&A lens", "Process timing", ["Likely Process Timing", "Pipeline Status"]),
    ("M&A lens", "Acquisition difficulty", ["Acquisition Difficulty", "Integration Complexity"]),
    ("M&A lens", "Integration complexity", ["Integration Complexity", "Product Integration Potential"]),
    ("M&A lens", "Diligence risk", ["Diligence Risk", "Compliance Risk Notes", "Data Sensitivity Tier"]),
    ("M&A lens", "Red flags", ["Red Flags", "Competitive Weakness", "Churn Risk"]),
    ("M&A lens", "Open diligence question", ["Key Open Questions", "Diligence Risk"]),
    # Pipeline ops
    ("Pipeline operations", "Pipeline status", ["Pipeline Status", "Deal Owner"]),
    ("Pipeline operations", "Deal ownership", ["Deal Owner", "Relationship Strength"]),
    ("Pipeline operations", "Research ownership", ["Research Owner", "Human Review Status"]),
    ("Pipeline operations", "Review freshness", ["Last Reviewed Date", "Last Updated Date"]),
    ("Pipeline operations", "Outreach recency", ["Last Outreach Date", "Relationship Strength"]),
    ("Pipeline operations", "Recommended next action", ["Next Action", "Build / Buy / Partner Recommendation"]),
    ("Pipeline operations", "Next action owner", ["Next Action Owner", "Deal Owner"]),
    ("Pipeline operations", "Target contact", ["Target Contact", "Key Executives"]),
    ("Pipeline operations", "Relationship strength", ["Relationship Strength", "Intro Path"]),
    ("Pipeline operations", "Intro path", ["Intro Path", "Relationship Strength"]),
    ("Pipeline operations", "Evidence quality", ["Evidence Links", "Source Confidence Score"]),
    ("Pipeline operations", "Human review status", ["Human Review Status", "Agent-Generated Summary"]),
]


# ============================================================
# Helpers
# ============================================================
def split_rating_rationale(text: Optional[str]) -> tuple[str, str]:
    """Split a 'Rating — Rationale' cell. Returns (rating, rationale). Empty parts are ''."""
    if text is None:
        return "", ""
    s = str(text).strip()
    if not s:
        return "", ""
    # Split on em-dash with surrounding spaces (most common from pipeline output)
    for sep in [" — ", " — ", " - "]:
        if sep in s:
            parts = s.split(sep, 1)
            return parts[0].strip(), parts[1].strip()
    return s, ""


def truncate_words(text: str, max_words: int = 11) -> str:
    """Trim a phrase to max_words. Strip trailing periods."""
    if not text:
        return "—"
    words = re.findall(r"\S+", text)
    if len(words) <= max_words:
        out = " ".join(words)
    else:
        out = " ".join(words[:max_words])
    return out.rstrip(".,;:")


def synthesize_phrase(label: str, source_values: list[tuple[str, str]]) -> str:
    """Build the executive phrase from a list of (rating, rationale) tuples for the source fields.

    Strategy: prefer the rationale of the first non-empty source. If only ratings exist (no rationale),
    concatenate ratings. If both are empty, return '—'.
    """
    # Find first source with a rationale (the rich content)
    for rating, rationale in source_values:
        if rationale:
            return truncate_words(rationale, max_words=11)

    # No rationale anywhere — fall back to concatenated ratings
    ratings = [r for r, _ in source_values if r and r != "—"]
    if not ratings:
        return "—"

    # Numeric score handling — pass through as-is
    if len(ratings) == 1 and re.fullmatch(r"-?\d+(\.\d+)?", ratings[0]):
        return ratings[0]

    return truncate_words(", ".join(ratings[:2]), max_words=10)


def find_field_index(ws: Worksheet, field_col: int, field_name: str) -> Optional[int]:
    """Find row number where col `field_col` matches field_name (case/space tolerant)."""
    target = re.sub(r"\s+", " ", field_name.strip().lower())
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=field_col).value
        if cell is None:
            continue
        if re.sub(r"\s+", " ", str(cell).strip().lower()) == target:
            return row
    return None


def detect_layout(ws: Worksheet) -> tuple[int, int, int, list[tuple[int, str]]]:
    """Detect the source sheet layout.

    Returns (header_row, section_col, field_col, [(company_col, company_name), ...]).
    Heuristic: scan for a row containing 'Company Name' in any cell, then row above is header.
    """
    field_col = None
    section_col = None
    cn_row = None
    for row in range(1, min(ws.max_row, 30) + 1):
        for col in range(1, min(ws.max_column, 10) + 1):
            v = ws.cell(row=row, column=col).value
            if v and re.sub(r"\s+", " ", str(v).strip().lower()) == "company name":
                cn_row = row
                field_col = col
                section_col = col - 1 if col > 1 else col
                break
        if cn_row:
            break

    if cn_row is None:
        raise RuntimeError("Could not find 'Company Name' field in source sheet")

    # Header row is usually 1 above the first data row, and companies are in cells of cn_row from field_col+1 onward
    header_row = cn_row
    companies: list[tuple[int, str]] = []
    for col in range(field_col + 1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v and str(v).strip():
            companies.append((col, str(v).strip()))

    return header_row, section_col, field_col, companies


def parse_score(text: str) -> Optional[float]:
    """Extract a leading number from a string like '16 — strong moat from ...' or '8.2'."""
    if not text:
        return None
    m = re.match(r"\s*(-?\d+(?:\.\d+)?)", str(text))
    return float(m.group(1)) if m else None


def derive_verdict(b_b_p: str) -> str:
    """Map Build/Buy/Partner Recommendation → ACQUIRE/MONITOR/PASS."""
    s = (b_b_p or "").lower()
    if "buy" in s or "acquire" in s or "active diligence" in s:
        return "ACQUIRE"
    if "monitor" in s or "watch" in s or "track" in s or "partner" in s:
        return "MONITOR"
    if "pass" in s or "no " in s or s.startswith("decline"):
        return "PASS"
    return "—"


def derive_posture(verdict: str, fit: Optional[float], moat: Optional[float], complexity: str) -> str:
    """Generate a 7–10-word posture phrase based on score pattern."""
    cx = (complexity or "").lower()
    if verdict == "ACQUIRE" and fit and fit >= 7.5 and "high" not in cx:
        return "Pursue active diligence this quarter"
    if verdict == "ACQUIRE" and "high" in cx:
        return "Partner first, revisit acquisition next year"
    if verdict == "MONITOR" and moat and moat >= 14:
        return "Strong moat, revisit on inflection signals"
    if verdict == "MONITOR":
        return "Track for entry window over next year"
    if verdict == "PASS" and moat and moat >= 14:
        return "Useful benchmark, weak Pattern fit"
    if verdict == "PASS":
        return "Off the table, no near-term action"
    return "Validate thesis before leadership outreach"


# ============================================================
# Main builder
# ============================================================
def build(source_path: Path, output_path: Optional[Path], sort_by: str) -> Path:
    wb = load_workbook(source_path)
    src_name = "Competitive Landscape" if "Competitive Landscape" in wb.sheetnames else wb.sheetnames[0]
    src = wb[src_name]
    print(f"source sheet: {src_name}")

    header_row, section_col, field_col, companies = detect_layout(src)
    print(f"layout: header_row={header_row} field_col={field_col} companies={len(companies)}")

    # Build field-row map
    field_row: dict[str, int] = {}
    for r in range(1, src.max_row + 1):
        v = src.cell(row=r, column=field_col).value
        if v:
            key = re.sub(r"\s+", " ", str(v).strip())
            field_row[key.lower()] = r

    def get_cell(field_name: str, company_col: int) -> tuple[str, str]:
        """Return (rating, rationale) for the given field/company."""
        key = re.sub(r"\s+", " ", field_name.strip().lower())
        # Try exact, then with various separator normalizations
        candidates = [key, key.replace(" / ", "/"), key.replace("/", " / ")]
        for c in candidates:
            r = field_row.get(c)
            if r:
                v = src.cell(row=r, column=company_col).value
                return split_rating_rationale(v)
        return "", ""

    # Compute verdicts and scores per company
    company_data: list[dict] = []
    for col, name in companies:
        bbp = get_cell("Build / Buy / Partner Recommendation", col)[0] or get_cell("Build/Buy/Partner Recommendation", col)[0]
        verdict = derive_verdict(bbp)
        moat_r, _ = get_cell("Total Moat Score", col)
        fit_r, _ = get_cell("Strategic Fit Score", col)
        # M&A attractiveness — derive from deal priority numeric if available, else None
        ma_r, _ = get_cell("Deal Priority", col)
        complexity_r, _ = get_cell("Integration Complexity", col)
        thesis_r, thesis_rat = get_cell("Acquisition Thesis", col)
        company_data.append({
            "col": col,
            "name": name,
            "verdict": verdict,
            "moat": parse_score(moat_r),
            "fit": parse_score(fit_r),
            "ma": parse_score(ma_r),
            "complexity": complexity_r,
            "thesis": thesis_rat or thesis_r,
        })

    # Sort
    if sort_by == "priority":
        order = {"ACQUIRE": 0, "MONITOR": 1, "PASS": 2, "—": 3}
        company_data.sort(key=lambda c: (order.get(c["verdict"], 9), -(c["fit"] or 0), -(c["moat"] or 0)))
    elif sort_by == "fit":
        company_data.sort(key=lambda c: -(c["fit"] or 0))
    elif sort_by == "moat":
        company_data.sort(key=lambda c: -(c["moat"] or 0))
    # source order already preserved if 'source'

    # Create or replace output sheet
    out_name = "Executive Deliverable"
    if out_name in wb.sheetnames:
        del wb[out_name]
    out = wb.create_sheet(out_name, index=1)

    # Build the sheet
    write_deliverable(out, company_data, get_cell)

    # Save
    target = output_path or source_path
    wb.save(target)
    print(f"saved: {target}")
    return target


def write_deliverable(out: Worksheet, company_data: list[dict], get_cell):
    n_companies = len(company_data)

    thin = Side(border_style="thin", color=BORDER_GREY)
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Column widths
    out.column_dimensions["A"].width = 18
    out.column_dimensions["B"].width = 28
    for i, _ in enumerate(company_data):
        out.column_dimensions[get_column_letter(3 + i)].width = 32

    # Row 1: title band
    out.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + n_companies)
    c = out.cell(row=1, column=1, value="Competitive Landscape — Executive View")
    c.font = Font(name=FONT_BOLD, size=20, color=WHITE, bold=True)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    out.row_dimensions[1].height = 36

    # Row 2: subtitle
    out.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + n_companies)
    c = out.cell(row=2, column=1, value="Concise strategic readout from detailed research")
    c.font = Font(name=FONT_BODY, size=10, color=BRIGHT_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    out.row_dimensions[2].height = 18

    # Row 3: company headers
    hdr_a = out.cell(row=3, column=1, value="Verdict")
    hdr_a.font = Font(name=FONT_BOLD, size=10, color=NAVY, bold=True)
    hdr_a.fill = PatternFill("solid", fgColor=SECTION_FILL)
    hdr_a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    hdr_b = out.cell(row=3, column=2, value="Recommendation")
    hdr_b.font = Font(name=FONT_BOLD, size=9, color=BLACK, bold=True)
    hdr_b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for i, cd in enumerate(company_data):
        c = out.cell(row=3, column=3 + i, value=cd["name"])
        c.font = Font(name=FONT_BOLD, size=11, color=WHITE, bold=True)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    out.row_dimensions[3].height = 30

    # Row 4-6: verdict / posture / score / thesis
    verdict_row = 4
    out.cell(row=verdict_row, column=2, value="Verdict").font = Font(name=FONT_BOLD, size=9, bold=True)
    out.cell(row=verdict_row, column=2).alignment = Alignment(vertical="center", indent=1)
    for i, cd in enumerate(company_data):
        v = cd["verdict"]
        c = out.cell(row=verdict_row, column=3 + i, value=v)
        if v in VERDICT_FILL:
            fill_color, text_color = VERDICT_FILL[v]
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.font = Font(name=FONT_BOLD, size=11, color=text_color, bold=True)
        else:
            c.font = Font(name=FONT_BODY, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    out.row_dimensions[verdict_row].height = 28

    posture_row = 5
    out.cell(row=posture_row, column=2, value="Recommended posture").font = Font(name=FONT_BOLD, size=9, bold=True)
    out.cell(row=posture_row, column=2).alignment = Alignment(vertical="center", indent=1)
    for i, cd in enumerate(company_data):
        phrase = derive_posture(cd["verdict"], cd["fit"], cd["moat"], cd["complexity"])
        c = out.cell(row=posture_row, column=3 + i, value=phrase)
        c.font = Font(name=FONT_BODY, size=9)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.border = border
    out.row_dimensions[posture_row].height = 22

    score_row = 6
    out.cell(row=score_row, column=2, value="Score (moat / fit / ma)").font = Font(name=FONT_BOLD, size=9, bold=True)
    out.cell(row=score_row, column=2).alignment = Alignment(vertical="center", indent=1)
    for i, cd in enumerate(company_data):
        m = f"{cd['moat']:g}" if cd['moat'] is not None else "—"
        f = f"{cd['fit']:g}" if cd['fit'] is not None else "—"
        a = f"{cd['ma']:g}" if cd['ma'] is not None else "—"
        c = out.cell(row=score_row, column=3 + i, value=f"{m} / {f} / {a}")
        c.font = Font(name=FONT_BOLD, size=10, color=NAVY, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    out.row_dimensions[score_row].height = 22

    thesis_row = 7
    out.cell(row=thesis_row, column=2, value="Acquisition thesis").font = Font(name=FONT_BOLD, size=9, bold=True)
    out.cell(row=thesis_row, column=2).alignment = Alignment(vertical="top", indent=1)
    for i, cd in enumerate(company_data):
        c = out.cell(row=thesis_row, column=3 + i, value=truncate_words(cd["thesis"] or "", max_words=14))
        c.font = Font(name=FONT_BODY, size=9, italic=True)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        c.border = border
    out.row_dimensions[thesis_row].height = 36

    # Freeze panes at C8 (below verdict block, after both label cols)
    out.freeze_panes = "C8"

    # Section + body rows
    cur_row = 8
    cur_section = None
    for section, label, source_fields in ROWS:
        if section != cur_section:
            # Section header row
            out.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=2 + n_companies)
            sh = out.cell(row=cur_row, column=1, value=section)
            sh.font = Font(name=FONT_BOLD, size=11, color=NAVY, bold=True)
            sh.fill = PatternFill("solid", fgColor=SECTION_FILL)
            sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            out.row_dimensions[cur_row].height = 20
            cur_row += 1
            cur_section = section

        # Field row
        # Column A blank, column B = label
        a = out.cell(row=cur_row, column=1, value="")
        a.fill = PatternFill("solid", fgColor=ALT_FILL) if (cur_row % 2 == 0) else PatternFill(fill_type=None)
        b = out.cell(row=cur_row, column=2, value=label)
        b.font = Font(name=FONT_BOLD, size=9, color=BLACK, bold=True)
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        b.fill = PatternFill("solid", fgColor=ALT_FILL) if (cur_row % 2 == 0) else PatternFill(fill_type=None)
        b.border = border

        for i, cd in enumerate(company_data):
            # Pull (rating, rationale) for each source field, in order
            source_values = [get_cell(f, cd["col"]) for f in source_fields]
            phrase = synthesize_phrase(label, source_values)
            cc = out.cell(row=cur_row, column=3 + i, value=phrase)
            cc.font = Font(name=FONT_BODY, size=9, color=BLACK)
            cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            cc.border = border
            if cur_row % 2 == 0:
                cc.fill = PatternFill("solid", fgColor=ALT_FILL)
        out.row_dimensions[cur_row].height = 22
        cur_row += 1

    # Quality checks
    qc_warnings = quality_checks(out, company_data)
    if qc_warnings:
        print("\n=== QC warnings ===")
        for w in qc_warnings:
            print(f"  - {w}")
    else:
        print("\nQC: all checks passed")


def quality_checks(out: Worksheet, company_data: list[dict]) -> list[str]:
    warnings: list[str] = []
    n = len(company_data)
    if out.max_row < 80:
        warnings.append(f"deliverable has only {out.max_row} rows (expected ≥80)")
    if out.max_column < n + 2:
        warnings.append(f"deliverable has {out.max_column} cols, expected {n + 2}")
    # Sample 12 random body cells
    import random
    rows_to_check = random.sample(range(8, out.max_row + 1), min(12, out.max_row - 7))
    for r in rows_to_check:
        for col in range(3, 3 + n):
            v = out.cell(row=r, column=col).value
            if v and isinstance(v, str):
                wc = len(re.findall(r"\S+", v))
                if wc > 14:
                    warnings.append(f"row {r} col {get_column_letter(col)}: {wc} words ('{v[:60]}...')")
                low = v.lower()
                for filler in BANNED_FILLER:
                    if filler in low:
                        warnings.append(f"row {r} col {get_column_letter(col)}: banned filler '{filler}'")
                        break
    # Verdict cells
    for i in range(n):
        v = out.cell(row=4, column=3 + i).value
        if v not in ("ACQUIRE", "MONITOR", "PASS", "—"):
            warnings.append(f"verdict col {3+i}: unexpected value '{v}'")
    return warnings


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--source", required=True, help="Path to source .xlsx")
    p.add_argument("--output", default=None, help="Path to output .xlsx (default: in-place)")
    p.add_argument("--sort-by", default="priority", choices=["priority", "fit", "moat", "source"])
    args = p.parse_args()
    src = Path(args.source).expanduser().resolve()
    if not src.exists():
        print(f"source not found: {src}", file=sys.stderr)
        sys.exit(2)
    out = Path(args.output).expanduser().resolve() if args.output else None
    build(src, out, args.sort_by)


if __name__ == "__main__":
    main()
