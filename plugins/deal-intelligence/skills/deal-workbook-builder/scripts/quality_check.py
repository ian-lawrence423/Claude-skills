"""
Deal Workbook Quality Check
============================
Runs after every build or edit of a deal workbook (Driver Tree, KPI Tree,
NTB Registry, MOIC Bridge tabs). Checks and auto-fixes:

  1. Column widths  — numeric/label columns too narrow to display content
  2. Row heights    — wrapped-text rows truncating commentary
  3. Number formats — missing or wrong formats on value cells
  4. Formula errors — #REF!, #DIV/0!, #VALUE!, #N/A, #NAME? anywhere
  5. Text fit audit — cells where content visually overflows based on
                      character count vs column width estimate

Usage:
    python quality_check.py <workbook.xlsx> [--fix] [--verbose]

Flags:
    --fix      Apply auto-fixes (widths, heights, formats). Default: report only.
    --verbose  Print every cell checked, not just issues.

Exit codes:
    0  No issues (or all fixed with --fix)
    1  Issues found and not fixed
"""
import sys
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

# ---------------------------------------------------------------------------
# Per-tab column width specs (col letter -> (min_width, max_width))
# These override the auto-fit calculation for columns whose content width
# is dominated by wrap-text commentary (would auto-fit to 200+ chars).
# ---------------------------------------------------------------------------
TAB_COL_SPECS = {
    "DRIVER TREE": {
        "A": (1.5,   1.5),    # indent spacer
        "B": (34,    36),     # driver label
        "C": (9,     10),     # source tag
        "D": (22,    28),     # metric description (wrap)
        "E": (7,     8),      # unit
        "F": (10,    11),     # FY2025A
        "G": (10,    11),     # FY2026E (Entry)
        "H": (10,    11),     # FY2028E (Exit)
        "I": (10,    11),     # Delta
        "J": (12,    13),     # Revenue impact $M
        "K": (9,     10),     # Incr EBITDA %
        "L": (12,    13),     # EBITDA impact $M
        "M": (13,    14),     # EV impact $M
        "N": (10,    11),     # MOIC delta
    },
    "INPUTS": {
        "A": (1.5,   1.5),
        "B": (40,    45),     # label (long)
        "C": (13,    14),     # value
        "D": (3,     4),      # spacer
        "E": (42,    50),     # notes (wrap)
    },
    "KPI TREE": {
        "A": (1.5,   1.5),
        "B": (4,     5),      # level
        "C": (32,    36),     # KPI name
        "D": (14,    16),     # category
        "E": (22,    26),     # model line
        "F": (12,    13),     # frequency
        "G": (14,    15),     # actual (FY2025A)
        "H": (14,    15),     # budget (FY2026E)
        "I": (14,    15),     # exit year (FY2028E)
        "J": (12,    13),     # $ variance
        "K": (11,    12),     # % variance
        "L": (14,    16),     # status
        "M": (36,    45),     # notes (wrap)
    },
    "NTB REGISTRY": {
        "A": (1.5,   1.5),
        "B": (8,     9),      # NTB label
        "C": (22,    24),     # theme
        "D": (50,    55),     # thesis summary (wrap)
        "E": (14,    15),     # base EV
        "F": (14,    15),     # base MOIC delta
        "G": (14,    15),     # upside EV
        "H": (14,    15),     # upside MOIC
    },
    "MOIC BRIDGE": {
        "A": (1.5,   1.5),
        "B": (48,    52),     # component label
        "C": (16,    18),     # equity value $M
        "D": (16,    18),     # MOIC contribution
        "E": (16,    18),     # cumulative MOIC
    },
    "FINANCIALS": {
        "B": (34,    38),
        "C": (12,    13),
        "D": (12,    13),
        "E": (12,    13),
        "F": (12,    13),
        "G": (42,    50),     # notes (wrap)
    },
}

# Minimum row height (pts) for rows containing wrapped commentary text
MIN_WRAP_HEIGHT = 28
DEFAULT_ROW_HEIGHT = 14

# Number format patterns: (substring_match_in_format, replacement_format)
# Applied when a cell has no format or a generic format
EXPECTED_FORMATS = {
    # Key: column letter in given tab -> desired format
    "DRIVER TREE": {
        "J": '#,##0;(#,##0);"-"',
        "L": '#,##0;(#,##0);"-"',
        "M": '#,##0;(#,##0);"-"',
        "N": '+0.000x;-0.000x;"-"',
        "K": "0.0%",
    },
    "NTB REGISTRY": {
        "E": "#,##0.000",
        "F": '+0.000x;-0.000x;"-"',
        "G": "#,##0.000",
        "H": '+0.000x;-0.000x;"-"',
    },
    "MOIC BRIDGE": {
        "D": '+0.000x;-0.000x;"-"',
        "E": "0.00x",
    },
    "INPUTS": {
        "C": "General",  # mixed: checked contextually
    },
}

FORMULA_ERROR_TOKENS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def col_letter(ws_col_int):
    return get_column_letter(ws_col_int)

def estimate_text_width(text):
    """Rough character-count width estimate (Excel units ≈ chars * 1.1 for Arial 9pt)."""
    if not text:
        return 0
    s = str(text)
    # For wrap-text cells, use longest single line
    lines = s.split("\n")
    return max(len(line) for line in lines) * 1.1

def estimate_wrapped_height(text, col_width_chars, font_size_pt=9):
    """Estimate row height (pts) needed to show wrapped text.

    Caps at 150pt (about 10-11 lines of 9pt text) to avoid runaway
    estimates when a narrow helper column contains long text that is
    really meant to span across merged/adjacent cells.
    """
    MAX_HEIGHT = 150
    if not text:
        return DEFAULT_ROW_HEIGHT
    s = str(text)
    if s.startswith("="):
        return DEFAULT_ROW_HEIGHT
    # Use effective column width: at least 20 chars so narrow label cols
    # don't produce 500pt estimates for spanned header text.
    effective_width = max(col_width_chars, 20)
    chars_per_line = max(1, int(effective_width / 1.1))
    raw_chars = len(s)
    approx_lines = max(1, (raw_chars + chars_per_line - 1) // chars_per_line)
    line_h = font_size_pt * 1.4
    return min(MAX_HEIGHT, max(DEFAULT_ROW_HEIGHT, int(approx_lines * line_h) + 4))


# ---------------------------------------------------------------------------
# Check / fix functions
# ---------------------------------------------------------------------------

def check_column_widths(ws, tab_name, fix=False, issues=None, verbose=False):
    if issues is None:
        issues = []
    specs = TAB_COL_SPECS.get(tab_name, {})

    for col_letter_str, (min_w, max_w) in specs.items():
        current = ws.column_dimensions[col_letter_str].width or 0
        if current < min_w * 0.85:  # 15% tolerance
            msg = f"  [{tab_name}] col {col_letter_str}: width {current:.1f} < min {min_w} — {'FIXED' if fix else 'NEEDS FIX'}"
            issues.append(("width", tab_name, col_letter_str, current, min_w))
            if verbose or not fix:
                print(msg)
            if fix:
                ws.column_dimensions[col_letter_str].width = min_w
                print(msg)
        elif verbose:
            print(f"  [{tab_name}] col {col_letter_str}: width {current:.1f} OK (min={min_w})")

    return issues


def check_row_heights(ws, tab_name, fix=False, issues=None, verbose=False):
    if issues is None:
        issues = []
    specs = TAB_COL_SPECS.get(tab_name, {})

    for row in ws.iter_rows():
        row_num = row[0].row
        current_h = ws.row_dimensions[row_num].height or DEFAULT_ROW_HEIGHT

        for cell in row:
            if not cell.value:
                continue
            is_wrap = cell.alignment and cell.alignment.wrap_text
            if not is_wrap:
                continue

            col_l = get_column_letter(cell.column)
            col_w = specs.get(col_l, (DEFAULT_ROW_HEIGHT, 60))[0]
            needed = estimate_wrapped_height(cell.value, col_w)

            if needed > current_h + 4:  # 4pt tolerance
                msg = (f"  [{tab_name}] row {row_num} col {col_l}: "
                       f"wrap height {current_h:.0f}pt < needed ~{needed:.0f}pt "
                       f"— {'FIXED' if fix else 'NEEDS FIX'}")
                issues.append(("height", tab_name, row_num, col_l, current_h, needed))
                if verbose or not fix:
                    print(msg)
                if fix:
                    ws.row_dimensions[row_num].height = max(needed, MIN_WRAP_HEIGHT)
                    print(msg)
                break  # one fix per row

    return issues


def check_formula_errors(ws, tab_name, issues=None, verbose=False):
    if issues is None:
        issues = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v and isinstance(v, str) and any(e in v for e in FORMULA_ERROR_TOKENS):
                msg = f"  [{tab_name}] {cell.coordinate}: formula error value = {v!r}"
                print(msg)
                issues.append(("formula_error", tab_name, cell.coordinate, v))
    return issues


def check_number_formats(ws, tab_name, fix=False, issues=None, verbose=False):
    if issues is None:
        issues = []
    fmt_specs = EXPECTED_FORMATS.get(tab_name, {})
    if not fmt_specs:
        return issues

    for col_l, expected_fmt in fmt_specs.items():
        col_idx = column_index_from_string(col_l)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value is None:
                continue
            if isinstance(cell.value, str) and not cell.value.startswith("="):
                continue  # skip text labels
            current_fmt = cell.number_format or "General"
            if expected_fmt == "General":
                continue  # mixed column, skip
            if current_fmt == "General" or current_fmt == "@":
                msg = (f"  [{tab_name}] {cell.coordinate}: "
                       f"format is '{current_fmt}', expected '{expected_fmt}' "
                       f"— {'FIXED' if fix else 'NEEDS FIX'}")
                issues.append(("format", tab_name, cell.coordinate, current_fmt, expected_fmt))
                if verbose or not fix:
                    print(msg)
                if fix:
                    cell.number_format = expected_fmt
                    print(msg)

    return issues


def check_text_overflow(ws, tab_name, issues=None, verbose=False):
    """Detect cells where text content clearly exceeds column width (no wrap)."""
    if issues is None:
        issues = []
    specs = TAB_COL_SPECS.get(tab_name, {})

    for row in ws.iter_rows():
        for cell in row:
            if not cell.value or isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            is_wrap = cell.alignment and cell.alignment.wrap_text
            if is_wrap:
                continue  # wrap-text is fine — covered by row height check
            col_l = get_column_letter(cell.column)
            col_w = specs.get(col_l, (None,))[0]
            if col_w is None:
                continue
            text_w = estimate_text_width(cell.value)
            if text_w > col_w * 1.3:  # 30% overflow threshold
                msg = (f"  [{tab_name}] {cell.coordinate}: "
                       f"text width ~{text_w:.0f} overflows col width {col_w} "
                       f"(value: {str(cell.value)[:40]!r})")
                if verbose:
                    print(msg)
                issues.append(("overflow", tab_name, cell.coordinate, text_w, col_w))

    return issues


# ---------------------------------------------------------------------------
# Main runner
# ---------------------------------------------------------------------------

def run_quality_check(workbook_path: str, fix: bool = False, verbose: bool = False):
    path = Path(workbook_path)
    if not path.exists():
        print(f"ERROR: File not found: {path}")
        sys.exit(2)

    print(f"\n{'='*60}")
    print(f"DEAL WORKBOOK QUALITY CHECK")
    print(f"File: {path.name}")
    print(f"Mode: {'FIX' if fix else 'REPORT ONLY'}")
    print(f"{'='*60}\n")

    wb = load_workbook(path)
    all_issues = []

    TABS_TO_CHECK = [
        "DRIVER TREE", "INPUTS", "KPI TREE",
        "NTB REGISTRY", "MOIC BRIDGE", "FINANCIALS",
    ]

    for tab_name in TABS_TO_CHECK:
        if tab_name not in wb.sheetnames:
            if verbose:
                print(f"[SKIP] Tab '{tab_name}' not present in workbook")
            continue

        ws = wb[tab_name]
        print(f"--- {tab_name} ---")

        check_column_widths(ws, tab_name, fix=fix, issues=all_issues, verbose=verbose)
        check_row_heights(ws, tab_name, fix=fix, issues=all_issues, verbose=verbose)
        check_formula_errors(ws, tab_name, issues=all_issues, verbose=verbose)
        check_number_formats(ws, tab_name, fix=fix, issues=all_issues, verbose=verbose)
        check_text_overflow(ws, tab_name, issues=all_issues, verbose=verbose)
        print()

    # Summary
    by_type = {}
    for issue in all_issues:
        t = issue[0]
        by_type[t] = by_type.get(t, 0) + 1

    print(f"{'='*60}")
    print(f"SUMMARY: {len(all_issues)} issue(s) found")
    # overflow is always report-only; formula_error is never auto-fixed
    REPORT_ONLY = {"overflow", "formula_error"}
    for t, count in sorted(by_type.items()):
        if t in REPORT_ONLY:
            status = "REVIEW MANUALLY"
        elif fix:
            status = "FIXED"
        else:
            status = "OPEN"
        print(f"  {t}: {count} ({status})")
    print(f"{'='*60}\n")

    if fix and all_issues:
        # Save via temp file to avoid PermissionError when file is open in Excel
        import tempfile, shutil
        fixable = [i for i in all_issues if i[0] != "overflow"]
        if fixable:
            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            import os
            os.close(tmp_fd)
            try:
                wb.save(tmp_path)
                shutil.copy2(tmp_path, path)
                print(f"Saved fixes to: {path}\n")
            except PermissionError:
                shutil.copy2(tmp_path, str(path).replace(".xlsx", "_qc_fixed.xlsx"))
                fixed_path = str(path).replace(".xlsx", "_qc_fixed.xlsx")
                print(f"WARNING: Original file locked — fixes saved to: {fixed_path}\n")
                print("Close the file in Excel and rename/replace manually.\n")
            finally:
                os.unlink(tmp_path)

    return all_issues


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Deal workbook quality check")
    parser.add_argument("workbook", help="Path to .xlsx workbook")
    parser.add_argument("--fix", action="store_true", help="Apply auto-fixes")
    parser.add_argument("--verbose", action="store_true", help="Show all cells checked")
    args = parser.parse_args()

    issues = run_quality_check(args.workbook, fix=args.fix, verbose=args.verbose)
    sys.exit(0 if not issues else 1)
