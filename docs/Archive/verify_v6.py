import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\IanLawrence\github\Claude-skills\docs\Archive\Shopify_Deal_Workbook_v6.xlsx")

print("=== KPI TREE header + sample rows ===")
ws = wb["KPI TREE"]
for r in [6, 7, 8, 9, 15, 18]:
    for c in range(6, 14):
        v = ws.cell(r, c).value
        if v is not None:
            print(f"  ({r},{c})={repr(str(v)[:70])}")

print("\n=== DRIVER TREE key label rows ===")
ws_dt = wb["DRIVER TREE"]
for r in [1, 2, 3, 5, 8, 13, 14, 21, 28, 30, 31, 32, 34, 35]:
    v2 = ws_dt.cell(r, 2).value
    v4 = ws_dt.cell(r, 4).value
    if v2:
        print(f"  B{r}={repr(str(v2)[:80])}")
    if v4:
        print(f"  D{r}={repr(str(v4)[:80])}")
