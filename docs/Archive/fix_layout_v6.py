"""
fix_layout_v6.py
=================
Two changes on top of v5:

1. KPI TREE
   - Swap G (Budget) ↔ H (Actual) so Actual is left of Budget
   - Insert new column I = FY2028E Exit year model/analyst references
   - Shift old I→J, J→K, K→L, L→M
   - Rewrite variance ($), variance %, status formulas for new column positions
   - Update header row 6 labels + column widths

2. DRIVER TREE readability
   - Remove overlong commentary rows B2/B3
   - Shorten column D descriptions to ≤80 chars
   - Clean up column B sub-header text (excessive indentation / verbose labels)
   - Trim value bridge header (row 28) and value bridge row labels (30-34)
   - Keep ALL formulas in F-N intact (no formula changes)
"""

import sys
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook
from openpyxl.styles import numbers

SRC = r"C:\Users\IanLawrence\github\Claude-skills\docs\Archive\Shopify_Deal_Workbook_v5.xlsx"
DST = r"C:\Users\IanLawrence\github\Claude-skills\docs\Archive\Shopify_Deal_Workbook_v6.xlsx"

wb = load_workbook(SRC)

# ============================================================
# 1. KPI TREE — reorder columns + add exit year
# ============================================================
ws = wb["KPI TREE"]

# Exit year (FY2028E) references / values by data row.
# None → leave blank (header rows, qualitative KPIs without model exit refs).
EXIT_REFS = {
    7:  "='FINANCIAL MODEL'!I41",                                          # GMV ($B)
    8:  "='FINANCIAL MODEL'!I17",                                          # Total Revenue
    9:  "=INPUTS!$C$61",                                                   # Sub Solutions
    10: "=INPUTS!$C$57",                                                   # Merch Solutions
    11: "='FINANCIAL MODEL'!I44",                                          # MRR
    12: "=IFERROR('FINANCIAL MODEL'!I19/'FINANCIAL MODEL'!I17,\"-\")",     # Blended GM
    13: 0.82,                                                              # Sub GM (analyst)
    14: 0.40,                                                              # Merch GM (analyst)
    15: "='FINANCIAL MODEL'!I26",                                          # EBITDA Margin
    16: "='FINANCIAL MODEL'!I9",                                           # S&M %
    17: "='FINANCIAL MODEL'!I10",                                          # R&D %
    18: "=IFERROR('FINANCIAL MODEL'!I24/'FINANCIAL MODEL'!I17,\"-\")",     # OpEx %
    19: None,                                                              # GMV KPIs header
    20: "='FINANCIAL MODEL'!I42",                                          # GPV%
    21: "='FINANCIAL MODEL'!I43",                                          # Payments Take Rate
    22: 0.85,                                                              # B2B GMV% (analyst)
    23: 0.50,                                                              # Intl GMV% (analyst)
    24: "='FINANCIAL MODEL'!I38",                                          # FCF
    25: "=IFERROR('FINANCIAL MODEL'!I38/'FINANCIAL MODEL'!I17,\"-\")",     # FCF Margin
    26: "='FINANCIAL MODEL'!I14",                                          # Capex %
}

# Rows where LOWER is better (cost/efficiency metrics):
# status = "On Track" if actual is not too much HIGHER than budget
COST_ROWS = {16, 17, 18, 26}

# Snapshot current values for cols G–L (cols 7–12), rows 6–26
snap = {}
for r in range(6, 27):
    snap[r] = {
        "G": ws.cell(r, 7).value,
        "H": ws.cell(r, 8).value,
        "I": ws.cell(r, 9).value,
        "J": ws.cell(r, 10).value,
        "K": ws.cell(r, 11).value,
        "L": ws.cell(r, 12).value,
    }

# Also snapshot number formats from H column (will apply to new G and exit I)
fmt_snap = {}
for r in range(7, 27):
    fmt_snap[r] = {
        "H_fmt": ws.cell(r, 8).number_format,   # old H = Actual
        "G_fmt": ws.cell(r, 7).number_format,   # old G = Budget
    }

# Clear cols G–M for rows 6–26
for r in range(6, 27):
    for c in range(7, 14):
        ws.cell(r, c).value = None

# ── Header row 6 ────────────────────────────────────────────────────────────
ws.cell(6, 7).value = "FY2025A\nActual"
ws.cell(6, 8).value = "FY2026E\nBudget"
ws.cell(6, 9).value = "FY2028E\nExit"
ws.cell(6, 10).value = "Variance ($)"
ws.cell(6, 11).value = "Variance %"
ws.cell(6, 12).value = "Status"
ws.cell(6, 13).value = "Notes"

# ── Data rows 7–26 ──────────────────────────────────────────────────────────
for r in range(7, 27):
    old_g = snap[r]["G"]   # was Budget
    old_h = snap[r]["H"]   # was Actual
    old_k = snap[r]["K"]   # was Status

    # Col G = Actual (was H)
    ws.cell(r, 7).value = old_h
    ws.cell(r, 7).number_format = fmt_snap[r]["H_fmt"]

    # Col H = Budget (was G)
    ws.cell(r, 8).value = old_g
    ws.cell(r, 8).number_format = fmt_snap[r]["G_fmt"]

    # Col I = Exit year
    exit_val = EXIT_REFS.get(r)
    ws.cell(r, 9).value = exit_val
    if exit_val is not None:
        ws.cell(r, 9).number_format = fmt_snap[r]["H_fmt"]  # same format as actual

    # Col J = Variance ($) — only if both actual and budget are present
    has_data = old_h is not None and old_g is not None
    if has_data:
        ws.cell(r, 10).value = f"=IFERROR(G{r}-H{r},\"-\")"

    # Col K = Variance % — only if variance $ is present
    if has_data:
        ws.cell(r, 11).value = f"=IFERROR((G{r}-H{r})/ABS(H{r}),\"-\")"

    # Col L = Status (updated column references)
    if old_k is None:
        pass
    elif old_k == "N/A":
        ws.cell(r, 12).value = "N/A"
    elif has_data:
        if r in COST_ROWS:
            # Lower is better: on track if actual not too much higher than budget
            ws.cell(r, 12).value = (
                f'=IFERROR(IF((G{r}-H{r})/ABS(H{r})<0.05,"On Track",'
                f'IF((G{r}-H{r})/ABS(H{r})<0.15,"Watch","Off Plan")),"-")'
            )
        else:
            # Higher is better: on track if actual not too far below budget
            ws.cell(r, 12).value = (
                f'=IFERROR(IF((G{r}-H{r})/ABS(H{r})>-0.05,"On Track",'
                f'IF((G{r}-H{r})/ABS(H{r})>-0.15,"Watch","Behind Plan")),"-")'
            )

    # Col M = Notes (old L)
    ws.cell(r, 13).value = snap[r]["L"]

# ── Column widths ────────────────────────────────────────────────────────────
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 14
ws.column_dimensions["I"].width = 14
ws.column_dimensions["J"].width = 12
ws.column_dimensions["K"].width = 11
ws.column_dimensions["L"].width = 14
ws.column_dimensions["M"].width = 42

print("KPI TREE: columns reordered, exit year added.")

# ============================================================
# 2. DRIVER TREE — readability cleanup
# ============================================================
ws_dt = wb["DRIVER TREE"]

# ── Title / commentary rows ──────────────────────────────────────────────────
ws_dt["B1"].value = "DRIVER TREE  |  SHOPIFY"
ws_dt["B2"].value = (
    "Paasche revenue decomposition: "
    "GMV Volume Effect + Attach Rate Improvement → EBITDA → EV → MOIC"
)
ws_dt["B3"].value = None   # Remove long methodology text — formula footnote row 35 covers it

# ── Section header row 5 ────────────────────────────────────────────────────
ws_dt["B5"].value = "REVENUE DRIVER DECOMPOSITION  —  EBITDA  —  EV  —  MOIC"

# ── Column D: shorten descriptions ─────────────────────────────────────────
SHORT_D = {
    8:  "ΔGMV × entry attach rate (240.9 bps) — sub-driver splits per INPUTS rows 68–71",
    9:  "Same-cohort GMV (40% of ΔGMV)",
    10: "New merchant GMV (30% of ΔGMV)",
    11: "International GMV (20% of ΔGMV) — +45% YoY",
    12: "B2B / POS / Channel GMV (10% of ΔGMV) — +80% YoY",
    13: "Total ΔGMV × entry attach rate 240.9 bps",
    14: "Δattach rate × exit GMV — GPV + non-payments components",
    15: "GPV% of GMV — IC Memo: $67M per 100bps at base GMV",
    16: "Residual attach rate (MCA, Balance, Markets)",
    17: "Total merch rev delta − GMV volume effect",
    20: "Revenue ≈ MRR × Revenue/MRR multiplier (INPUTS!C86)",
    22: "Plus MRR ($M) — $25M+ GMV cohort, 28%→39% of MRR",
    23: "Core MRR ($M) — Basic/Standard/Advanced plans",
}
for row, text in SHORT_D.items():
    ws_dt.cell(row, 4).value = text

# ── Column B: clean up sub-header labels ────────────────────────────────────
ws_dt["B8"].value  = "  GMV Volume Effect"
ws_dt["B14"].value = "  Attach Rate Improvement"
ws_dt["B21"].value = "    MRR Build"

# ── Value creation bridge section ────────────────────────────────────────────
ws_dt["B28"].value = "VALUE CREATION BRIDGE"
ws_dt["B30"].value = "Revenue-Driven EBITDA Growth  (EBITDA delta × entry multiple 48.5×)"
ws_dt["B31"].value = "Multiple Compression  (48.5× entry → 45.0× exit × exit EBITDA $4,687M)"
ws_dt["B32"].value = "Net Cash Build  (net cash $6.3B entry → $12.4B exit)"
ws_dt["B33"].value = "TOTAL VALUE CREATION"
ws_dt["B34"].value = "MOIC CHECK  =  1.00× + attribution  (should ≈ INPUTS!C34)"
ws_dt["B35"].value = (
    "Paasche: EV = EBITDA delta × entry multiple 48.5×. "
    "Multiple compression uses exit EBITDA. Net cash sign: negative = cash-rich."
)

print("DRIVER TREE: labels and commentary cleaned up.")

wb.save(DST)
print(f"\nSaved: {DST}")
