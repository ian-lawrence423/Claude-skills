"""
fix_format_v7.py
================
Fixes on top of v6:

1. KPI TREE — variance column correction
   - Replace broken J (absolute delta w/ % format) + K (variance %, General format) with:
       new J: single smart "Δ vs Budget" column
              margins → (G-H)*100 formatted +0.0"ppt";-0.0"ppt";"-"
              dollars → (G-H)/ABS(H) formatted +0.0%;-0.0%;"-"
   - Shift: K = Status (was L), L = Notes (was M), clear M

2. DRIVER TREE — fix margin delta rows
   - I15 (GPV%): change formula to (H-G)*100, format +0.0"ppt";-0.0"ppt";"-"
   - (Other delta cells already show correct units: bps or $M or $B)

3. McKinsey-style formatting — both tabs
   KPI TREE:
     - Remove ALL row-level fills (E8EDF2, F7F8FA, EBF4EA, D9E2F3)
     - Bold L1 rows; normal L2; bold NS row with thin bottom border
     - Thin separator borders between major groups
     - Add conditional formatting for Status column K
   DRIVER TREE:
     - Section headers (rows 7, 20, 28, 29): keep dark navy 0F4761
     - Sub-header labels (rows 8, 14): bold black, no fill
     - Subtotal rows (13, 17, 24): F0F0F0 light grey, bold
     - Total/bridge total rows (18, 26, 33): EBEBEB, bold, thin top border
     - MOIC check row 34: keep 0F4761 navy with white bold
     - Driver data rows: white, no fill
     - Remove D9E2F3 blue fills from totals
"""

import sys
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

SRC = r"C:\Users\IanLawrence\github\Claude-skills\docs\Archive\Shopify_Deal_Workbook_v6.xlsx"
DST = r"C:\Users\IanLawrence\github\Claude-skills\docs\Archive\Shopify_Deal_Workbook_v7.xlsx"

wb = load_workbook(SRC)

# ─── Style helpers ───────────────────────────────────────────────────────────
def fill(hex6):
    return PatternFill("solid", fgColor=hex6)

NO_FILL = PatternFill(fill_type=None)

def fnt(bold=False, color="000000", sz=9, italic=False, name="Calibri"):
    return Font(name=name, size=sz, bold=bold, color=color, italic=italic)

def bdr(top=None, bot=None, color="C0C0C0"):
    s = lambda c: Side(style="thin", color=c) if c else Side(style=None)
    return Border(top=s(color) if top else Side(style=None),
                  bottom=s(color) if bot else Side(style=None))

NAVY    = "0F4761"
LT_GREY = "F0F0F0"
MED_GREY= "EBEBEB"
WHITE   = "FFFFFF"
BLACK   = "000000"
GRN_TXT = "375623"
AMB_TXT = "9C5700"
RED_TXT = "C00000"
GRY_TXT = "666666"

# Row classifications for KPI TREE
L1_ROWS = {7, 8, 12, 15, 19, 24}        # bold, section-opener
L2_ROWS = {9, 10, 11, 13, 14, 16, 17,   # normal, indented
           18, 20, 21, 22, 23, 25, 26}
MARGIN_ROWS = {12,13,14,15,16,17,18,20,21,22,23,25,26}
DOLLAR_ROWS = {7,8,9,10,11,24}

# ============================================================
# 1. KPI TREE — variance column fix + formatting
# ============================================================
ws = wb["KPI TREE"]

# ── 1a. Fix variance columns ─────────────────────────────────────────────────
# Save current L (Status) and M (Notes) for rows 6-26
snap_status = {r: ws.cell(r, 12).value for r in range(6, 27)}
snap_notes  = {r: ws.cell(r, 13).value for r in range(6, 27)}

# Clear J, K, L, M for rows 6-26
for r in range(6, 27):
    for c in (10, 11, 12, 13):
        ws.cell(r, c).value = None
        ws.cell(r, c).number_format = "General"

# Header row 6
ws.cell(6, 10).value = "Δ vs Budget"
ws.cell(6, 11).value = "Status"
ws.cell(6, 12).value = "Notes"

PPT_FMT = '+0.0"ppt";-0.0"ppt";"-"'
PCT_FMT = "+0.0%;-0.0%;\"-\""

for r in range(7, 27):
    has_actual = ws.cell(r, 7).value is not None
    has_budget = ws.cell(r, 8).value is not None

    # Col J: Δ vs Budget
    if has_actual and has_budget:
        if r in MARGIN_ROWS:
            ws.cell(r, 10).value = f"=IFERROR((G{r}-H{r})*100,\"-\")"
            ws.cell(r, 10).number_format = PPT_FMT
        else:
            ws.cell(r, 10).value = f"=IFERROR((G{r}-H{r})/ABS(H{r}),\"-\")"
            ws.cell(r, 10).number_format = PCT_FMT

    # Col K: Status (was L)
    ws.cell(r, 11).value = snap_status[r]

    # Col L: Notes (was M)
    ws.cell(r, 12).value = snap_notes[r]

# ── 1b. Update column widths ─────────────────────────────────────────────────
ws.column_dimensions["J"].width = 13   # Δ vs Budget
ws.column_dimensions["K"].width = 13   # Status
ws.column_dimensions["L"].width = 44   # Notes
for c in ["M"]:
    ws.column_dimensions[c].width = 3

# ── 1c. Remove all row-level fills; apply McKinsey structure ─────────────────
for r in range(7, 27):
    is_l1 = r in L1_ROWS
    row_bold = is_l1

    for c in range(2, 13):   # B through L
        cell = ws.cell(r, c)
        cell.fill = NO_FILL

        # Font: bold for L1, grey italic for Notes (col L)
        existing = cell.font
        is_notes = (c == 12)
        is_val_col = c in (7, 8, 9)    # Actual / Budget / Exit
        is_hdr_col = c in (2, 3, 4, 5, 6)

        if is_notes:
            cell.font = fnt(bold=False, color=GRY_TXT, italic=True, sz=9)
        elif is_hdr_col:
            cell.font = fnt(bold=row_bold, color=BLACK, sz=9)
        else:
            cell.font = fnt(bold=False, color=BLACK, sz=9)

    # Thin top separator on first row of each major group
    # Groups: 7 (NS), 8 (Revenue block), 12 (Margins), 15 (EBITDA), 19 (GMV), 24 (Cash)
    if r in {7, 12, 15, 19, 24}:
        for c in range(2, 13):
            cell = ws.cell(r, c)
            cell.border = bdr(top=True, bot=False, color="C8C8C8")
    # NS row 7: also add bottom border (special row)
    if r == 7:
        for c in range(2, 13):
            cell = ws.cell(r, c)
            cell.border = bdr(top=True, bot=True, color="C8C8C8")

    # Bold the B/C label cells for L1
    ws.cell(r, 2).font = fnt(bold=row_bold, sz=9)
    ws.cell(r, 3).font = fnt(bold=row_bold, sz=9)

# ── 1d. Conditional formatting for Status column K ───────────────────────────
grn_dxf = DifferentialStyle(font=Font(bold=True, color=GRN_TXT))
amb_dxf = DifferentialStyle(font=Font(bold=True, color=AMB_TXT))
red_dxf = DifferentialStyle(font=Font(bold=True, color=RED_TXT))

r_green = Rule(type="containsText", operator="containsText", text="On Track", dxf=grn_dxf)
r_green.formula = ['NOT(ISERROR(SEARCH("On Track",K7)))']
r_amber = Rule(type="containsText", operator="containsText", text="Watch", dxf=amb_dxf)
r_amber.formula = ['NOT(ISERROR(SEARCH("Watch",K7)))']
r_red   = Rule(type="containsText", operator="containsText", text="Plan", dxf=red_dxf)
r_red.formula   = ['NOT(ISERROR(SEARCH("Plan",K7)))']

ws.conditional_formatting.add("K7:K26", r_green)
ws.conditional_formatting.add("K7:K26", r_amber)
ws.conditional_formatting.add("K7:K26", r_red)

print("KPI TREE: variance fixed, formatting applied.")

# ============================================================
# 2. DRIVER TREE — fix margin delta + formatting
# ============================================================
ws_dt = wb["DRIVER TREE"]

# ── 2a. Fix I15 (GPV%): delta should be ppt not raw decimal ──────────────────
# I15 currently = =H15-G15 with format 0.0% (shows e.g. 8.0% for 0.08 — ambiguous)
# Change to (H15-G15)*100 with +0.0"ppt" format so it reads "+8.0ppt"
ws_dt.cell(15, 9).value = "=(H15-G15)*100"
ws_dt.cell(15, 9).number_format = '+0.0"ppt";-0.0"ppt";"-"'

# ── 2b. Remove D9E2F3 blue fills from total rows; replace with clean grey ────
BLUE_FILL = "00D9E2F3"   # openpyxl stores fgColor as ARGB with alpha prefix
NAVY_FILL = "000F4761"

TOTAL_ROWS    = {18, 26}   # section totals → clean white + bold + border
SUBTOTAL_ROWS = {13, 17, 24}   # subtotals → light grey
BRIDGE_ROWS   = {30, 31, 32}   # value bridge driver rows → white

for r in range(1, 36):
    row_fill = ws_dt.row_dimensions[r].fill  # not useful but read for reference
    for c in range(2, 15):
        cell = ws_dt.cell(r, c)
        rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else "00000000"

        if rgb == BLUE_FILL:
            # Was a blue-fill total row — change to white + bold + top border
            cell.fill = NO_FILL
            if c in (2, 3, 4):
                cell.font = fnt(bold=True, sz=9)
            else:
                cell.font = fnt(bold=True, sz=9)
            cell.border = bdr(top=True, bot=True, color="AAAAAA")

        elif r in SUBTOTAL_ROWS:
            cell.fill = fill(LT_GREY)

        elif r in BRIDGE_ROWS:
            cell.fill = NO_FILL

# ── 2c. Sub-section label rows (8, 14): remove black fill, set to bold black ──
for r_label in (8, 14):
    for c in range(2, 15):
        cell = ws_dt.cell(r_label, c)
        rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else "00000000"
        if rgb == "00000000":   # black fill — remove it
            cell.fill = NO_FILL
        cell.font = fnt(bold=True, sz=9)

# ── 2d. Row 35 footnote: grey italic ────────────────────────────────────────
for c in range(2, 15):
    ws_dt.cell(35, c).fill = NO_FILL
    ws_dt.cell(35, c).font = fnt(bold=False, color=GRY_TXT, italic=True, sz=8)

print("DRIVER TREE: margin delta fixed, formatting applied.")

wb.save(DST)
print(f"\nSaved: {DST}")
