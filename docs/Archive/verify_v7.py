import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\IanLawrence\github\Claude-skills\docs\Archive\Shopify_Deal_Workbook_v7.xlsx")

print("=== KPI TREE cols G-L rows 6-15 (check variance + status) ===")
ws = wb["KPI TREE"]
for r in range(6, 16):
    line = f"Row {r:2d}:"
    for c in range(7, 13):
        cell = ws.cell(r, c)
        v = cell.value
        fmt = cell.number_format
        fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else "none"
        if v is not None:
            line += f"  [{chr(64+c)}:{repr(str(v)[:30])} fmt={fmt!r} fill={fill}]"
    print(line)

print("\n=== DRIVER TREE delta col I for margin rows ===")
ws_dt = wb["DRIVER TREE"]
for r in [15]:  # GPV%
    cell = ws_dt.cell(r, 9)
    print(f"  I{r}: val={repr(cell.value)} fmt={cell.number_format!r}")

print("\n=== DRIVER TREE fill check for key rows ===")
for r in [7, 8, 9, 13, 18, 26, 33, 34]:
    cell_b = ws_dt.cell(r, 2)
    fill = cell_b.fill.fgColor.rgb if cell_b.fill and cell_b.fill.fgColor else "none"
    bold = cell_b.font.bold if cell_b.font else False
    print(f"  Row {r:2d} B: {repr(str(cell_b.value)[:45])} fill={fill} bold={bold}")
