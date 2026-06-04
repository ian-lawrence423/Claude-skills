import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

PATH = r"C:\Users\IanLawrence\github\Claude-skills\docs\Archive\Shopify_Deal_Workbook_v7.xlsx"
wb = load_workbook(PATH)
ws = wb["KPI TREE"]

# QC widened M to 36 (old notes spec) — M is now empty, shrink it
ws.column_dimensions["M"].width = 3

# Also clear any M column values that QC may have touched (shouldn't have, but be safe)
for r in range(6, 27):
    ws.cell(r, 13).value = None

wb.save(PATH)
print("M column reset to narrow.")
