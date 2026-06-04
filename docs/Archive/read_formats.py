import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(r"C:\Users\IanLawrence\github\Claude-skills\docs\Archive\Shopify_Deal_Workbook_v6.xlsx")

print("=== KPI TREE column layout (rows 6-26, cols B-M) ===")
ws = wb["KPI TREE"]
for r in range(6, 27):
    row_data = []
    for c in range(2, 14):
        cell = ws.cell(r, c)
        v = cell.value
        if v is not None:
            fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else "none"
            fmt = cell.number_format
            bold = cell.font.bold if cell.font else False
            row_data.append(f"  {get_column_letter(c)}{r}: val={repr(str(v)[:45])} fill={fill} fmt={fmt!r} bold={bold}")
    if row_data:
        print(f"Row {r}:")
        print("\n".join(row_data))

print("\n=== DRIVER TREE key row formatting (cols B,F,G,H,I,J) ===")
ws_dt = wb["DRIVER TREE"]
for r in [5,6,7,8,9,13,14,15,17,18,19,20,24,25,26,27,28,29,30,33,34,35]:
    row_data = []
    for c in [2, 6, 7, 8, 9, 10]:
        cell = ws_dt.cell(r, c)
        v = cell.value
        fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else "none"
        bold = cell.font.bold if cell.font else False
        fmt = cell.number_format
        if v is not None:
            row_data.append(f"  {get_column_letter(c)}{r}: {repr(str(v)[:35])} fill={fill} fmt={fmt!r} bold={bold}")
    if row_data:
        print(f"Row {r}:")
        print("\n".join(row_data))
