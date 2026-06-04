import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\IanLawrence\github\Claude-skills\docs\Archive\Shopify_Deal_Workbook_v5.xlsx")

print("=== DRIVER TREE ===")
ws = wb["DRIVER TREE"]
for row in ws.iter_rows(min_row=1, max_row=40):
    vals = []
    for cell in row:
        v = cell.value
        if v is not None:
            vals.append(f"  {cell.coordinate}={repr(v)[:90]}")
    if vals:
        print("\n".join(vals))
