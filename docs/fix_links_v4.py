"""
Fix v4 workbook linkage issues:
1. NTB REGISTRY E7-E12, F7-F13 -- replace hardcodes with DRIVER TREE formula links
2. MOIC BRIDGE -- fix 1-row offset, wrong column refs, add D14 (compression), add row 15 (net cash)

NTB-to-DRIVER TREE mapping (Paasche decomposition):
  NTB 1 GMV Flywheel = DRIVER TREE M13 (GMV vol) + M16 (non-payments) + M23 (core MRR)
                     = M18 - M15 + M23  [all merch ex-GPV, plus core sub]
  NTB 2 Payments     = DRIVER TREE M15 (GPV penetration EV)
  NTB 3 AI           = 0 (no model basis)
  NTB 4 Enterprise   = DRIVER TREE M22 (Plus/Enterprise MRR EV)
  NTB 5 Compression  = DRIVER TREE L31 (value bridge multiple compression EV)
  Net Cash           = DRIVER TREE L32 (value bridge net cash EV)
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WB = r"C:\Users\IanLawrence\github\Claude-skills\docs\Shopify_Deal_Workbook_v4.xlsx"

FONT    = "Arial"
SZ      = 9
C_NAVY  = "0F4761"
C_TEAL  = "D9E2F3"
C_GREY  = "F2F2F2"
C_BLACK = "000000"
C_WHITE = "FFFFFF"
C_SLATE = "444444"
C_BDR   = "DDDDDD"
C_RED   = "C00000"
C_GREEN = "375623"

def fn(bold=False, color=C_BLACK, sz=SZ, italic=False):
    return Font(name=FONT, bold=bold, color=color, size=sz, italic=italic)

def fl(c):
    return PatternFill("solid", fgColor=c)

def bd():
    t = Side(style="thin", color=C_BDR)
    return Border(top=t, bottom=t, left=t, right=t)

def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def sc(ws, r, c, val, bold=False, color=C_BLACK, bg=None, h="right",
       fmt=None, sz=SZ, italic=False, wrap=False):
    cell = ws.cell(r, c, val)
    cell.font = Font(name=FONT, bold=bold, color=color, size=sz, italic=italic)
    if bg:
        cell.fill = fl(bg)
    cell.alignment = Alignment(horizontal=h, vertical="center", wrap_text=wrap)
    cell.border = bd()
    if fmt:
        cell.number_format = fmt
    return cell

wb = load_workbook(WB)

# ==============================================================================
# 1. FIX NTB REGISTRY (cols E=5, F=6)
# ==============================================================================
ntb = wb["NTB REGISTRY"]

FMT_EV   = "#,##0.000"
FMT_MOIC = '+0.000x;-0.000x;"-"'

# NTB 1 (row 7): GMV Flywheel
ntb.cell(7, 5, "='DRIVER TREE'!M13+'DRIVER TREE'!M16+'DRIVER TREE'!M23").number_format = FMT_EV
ntb.cell(7, 6, "=IFERROR(E7/INPUTS!$C$21,\"-\")").number_format = FMT_MOIC

# NTB 2 (row 8): Payments/GPV
ntb.cell(8, 5, "='DRIVER TREE'!M15").number_format = FMT_EV
ntb.cell(8, 6, "=IFERROR(E8/INPUTS!$C$21,\"-\")").number_format = FMT_MOIC

# NTB 3 (row 9): AI Commerce -- no model basis
ntb.cell(9, 5, 0).number_format = FMT_EV
ntb.cell(9, 6, 0).number_format = FMT_MOIC

# NTB 4 (row 10): Enterprise Plus
ntb.cell(10, 5, "='DRIVER TREE'!M22").number_format = FMT_EV
ntb.cell(10, 6, "=IFERROR(E10/INPUTS!$C$21,\"-\")").number_format = FMT_MOIC

# NTB 5 (row 11): Multiple Compression Risk
ntb.cell(11, 5, "='DRIVER TREE'!L31").number_format = FMT_EV
ntb.cell(11, 6, "=IFERROR(E11/INPUTS!$C$21,\"-\")").number_format = FMT_MOIC

# TOTAL (row 12): NTB 1-5 + Net Cash from value bridge
ntb.cell(12, 5, "=SUM(E7:E11)+'DRIVER TREE'!L32").number_format = FMT_EV
ntb.cell(12, 5).font = fn(bold=True, color=C_NAVY)
ntb.cell(12, 6, "=IFERROR(E12/INPUTS!$C$21,\"-\")").number_format = FMT_MOIC
ntb.cell(12, 6).font = fn(bold=True, color=C_NAVY)

# MOIC CHECK (row 13)
ntb.cell(13, 6, "=IFERROR(1+F12,\"-\")").number_format = '0.000"x"'
ntb.cell(13, 6).font = fn(bold=True, color=C_NAVY, sz=10)

print("NTB REGISTRY: E7:F13 linked to DRIVER TREE formulas")

# ==============================================================================
# 2. FIX MOIC BRIDGE
#    Bugs: 1-row offset (reads registry row 8 for NTB-1, should be row 7),
#          wrong columns (reads F/G instead of E/F),
#          C14/D14 empty (multiple compression missing),
#          no net cash row (row 15 blank)
# ==============================================================================
mb = wb["MOIC BRIDGE"]

FMT_EV_B = "#,##0;(#,##0);\"-\""
FMT_MC_B = '+0.000x;-0.000x;"-"'
FMT_CUM  = "0.00x"

# Row 8: NTB-1 GMV -> registry row 7
mb.cell(8, 3, "=IFERROR('NTB REGISTRY'!E7,\"-\")").number_format = FMT_EV_B
mb.cell(8, 4, "=IFERROR('NTB REGISTRY'!F7,\"-\")").number_format = FMT_MC_B

# Row 9: NTB-2 Payments -> registry row 8
mb.cell(9, 3, "=IFERROR('NTB REGISTRY'!E8,\"-\")").number_format = FMT_EV_B
mb.cell(9, 4, "=IFERROR('NTB REGISTRY'!F8,\"-\")").number_format = FMT_MC_B

# Row 10: NTB-3 AI -> registry row 9 (= 0)
mb.cell(10, 3, "=IFERROR('NTB REGISTRY'!E9,\"-\")").number_format = FMT_EV_B
mb.cell(10, 4, "=IFERROR('NTB REGISTRY'!F9,\"-\")").number_format = FMT_MC_B

# Row 11: NTB-4 Enterprise -> registry row 10
mb.cell(11, 3, "=IFERROR('NTB REGISTRY'!E10,\"-\")").number_format = FMT_EV_B
mb.cell(11, 4, "=IFERROR('NTB REGISTRY'!F10,\"-\")").number_format = FMT_MC_B

# Row 12: NTB-5 FCF -- no registry entry, not modeled
mb.cell(12, 2, "NTB-5: FCF Margin Expands to 20%+  [not modeled in base case]")
mb.cell(12, 3, 0).number_format = FMT_EV_B
mb.cell(12, 4, 0).number_format = FMT_MC_B

# Row 13: NTB-6 Macro -- no registry entry, not modeled
mb.cell(13, 2, "NTB-6: Macro Does Not Materially Deteriorate  [not modeled in base case]")
mb.cell(13, 3, 0).number_format = FMT_EV_B
mb.cell(13, 4, 0).number_format = FMT_MC_B

# Row 14: Multiple Compression -- was missing C14 / D14
bg14 = C_GREY
mb.cell(14, 2, "Multiple Re-rating / Compression  (48.5x entry to 45.0x exit x $4.7B exit EBITDA)")
mb.cell(14, 2).font = fn(bold=True, color=C_RED)
mb.cell(14, 2).fill = fl(bg14)
mb.cell(14, 2).alignment = al("left")
mb.cell(14, 2).border = bd()
sc(mb, 14, 3, "='DRIVER TREE'!L31", bold=True, color=C_RED, bg=bg14, fmt=FMT_EV_B)
sc(mb, 14, 4, "='DRIVER TREE'!N31", bold=True, color=C_RED, bg=bg14, fmt=FMT_MC_B)
sc(mb, 14, 5, "=IFERROR(E13+D14,\"-\")", bold=True, color=C_RED, bg=bg14, fmt=FMT_CUM)

# Row 15: Net Cash Build (was blank)
mb.row_dimensions[15].height = 14
bg15 = "FFFFFF"
sc(mb, 15, 2, "Net Cash Build  (Net cash $6.3B to $12.4B; INPUTS rows 20 / 29)",
   bold=True, color=C_GREEN, bg=bg15, h="left")
sc(mb, 15, 3, "='DRIVER TREE'!L32", bold=True, color=C_GREEN, bg=bg15, fmt=FMT_EV_B)
sc(mb, 15, 4, "='DRIVER TREE'!N32", bold=True, color=C_GREEN, bg=bg15, fmt=FMT_MC_B)
sc(mb, 15, 5, "=IFERROR(E14+D15,\"-\")", bold=True, color=C_GREEN, bg=bg15, fmt=FMT_CUM)
for col in range(6, 9):
    c = mb.cell(15, col)
    c.fill = fl(bg15)
    c.border = bd()

# Row 16 note: E15 cumulative should now match E16 (direct MOIC)
# E16 = INPUTS!C30/C21 is independent sanity check -- leave unchanged

print("MOIC BRIDGE: offsets fixed, D14/C14 added, row 15 net cash added")

# ==============================================================================
# Save
# ==============================================================================
wb.save(WB)
print(f"\nSaved: {WB}")

# Verification
print("\n=== NTB REGISTRY E7:F13 ===")
wb2 = load_workbook(WB)
ntb2 = wb2["NTB REGISTRY"]
for r in range(7, 14):
    e = ntb2.cell(r, 5).value
    f = ntb2.cell(r, 6).value
    print(f"  Row {r}: E={e} | F={f}")

print("\n=== MOIC BRIDGE rows 8-16 ===")
mb2 = wb2["MOIC BRIDGE"]
for r in range(8, 17):
    b  = mb2.cell(r, 2).value
    c3 = mb2.cell(r, 3).value
    c4 = mb2.cell(r, 4).value
    c5 = mb2.cell(r, 5).value
    if any([c3, c4, c5, b]):
        print(f"  Row {r} [{str(b or '')[:40]}]: C={c3} | D={c4} | E={c5}")
