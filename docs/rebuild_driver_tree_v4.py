"""
Shopify Deal Workbook v4 — Formula-based Driver Tree + KPI Tree fix
Replaces all hardcoded values with Excel formula chains:
  FINANCIAL MODEL (source data) → INPUTS (assumptions) → DRIVER TREE (impact calcs) → NTB REGISTRY

INPUTS new rows 46-86: Driver Tree Assumptions
  - Rows 48-62: auto-linked to FINANCIAL MODEL (black formula text)
  - Rows 65, 68-71, 75-76, 80-85: editable analyst inputs (blue text)
  - Row 86: derived multiplier (black formula)

DRIVER TREE:
  - Metric columns (F/G/H) = formulas referencing INPUTS
  - Delta col (I) = H-G
  - Rev impact (J) = formula from metric delta × rate assumptions
  - EBITDA impact (L) = J × K (incremental margin from INPUTS)
  - EV impact (M) = L × INPUTS!C18 (entry multiple)
  - MOIC delta (N) = M / INPUTS!C21 (entry equity)

KPI TREE:
  - Rows 9-10 (Sub/Merch Solutions Revenue): formula-linked to INPUTS
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WB_IN  = r"C:\Users\IanLawrence\github\Claude-skills\docs\Shopify_Deal_Workbook_v3.xlsx"
WB_OUT = r"C:\Users\IanLawrence\github\Claude-skills\docs\Shopify_Deal_Workbook_v4.xlsx"

# ── Style constants ────────────────────────────────────────────────────────────
C_NAVY    = "0F4761"
C_BLUE    = "4280F4"
C_INDIGO  = "3A00FD"
C_WHITE   = "FFFFFF"
C_BLACK   = "000000"
C_SLATE   = "444444"
C_GREY    = "F2F2F2"
C_BDR     = "DDDDDD"
C_TEAL    = "D9E2F3"
C_GREEN   = "375623"
C_RED     = "C00000"
C_ORANGE  = "C55A11"
C_INPUT   = "0000FF"   # blue = user-editable hardcoded input
FONT      = "Arial"
SZ        = 9

def fn(bold=False, color=C_BLACK, sz=SZ, italic=False):
    return Font(name=FONT, bold=bold, color=color, size=sz, italic=italic)

def fl(c): return PatternFill("solid", fgColor=c)

def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bd():
    t = Side(style="thin", color=C_BDR)
    return Border(top=t, bottom=t, left=t, right=t)

def bd_bottom():
    return Border(bottom=Side(style="thin", color=C_BDR))

def set_cell(ws, row, col, value=None, formula=None, bold=False, color=C_BLACK,
             bg=None, h="left", fmt=None, sz=SZ, italic=False, border=True, wrap=False):
    c = ws.cell(row=row, column=col, value=formula if formula else value)
    c.font = Font(name=FONT, bold=bold, color=color, size=sz, italic=italic)
    if bg: c.fill = fl(bg)
    c.alignment = Alignment(horizontal=h, vertical="center", wrap_text=wrap)
    if border: c.border = bd()
    if fmt: c.number_format = fmt
    return c

# ── Load workbook ──────────────────────────────────────────────────────────────
wb = load_workbook(WB_IN)

# ══════════════════════════════════════════════════════════════════════════════
# 1. UPDATE INPUTS TAB — add Driver Tree Assumptions section (rows 46-86)
#    Replace old EBITDA pool section with new formula-linked assumptions
# ══════════════════════════════════════════════════════════════════════════════
inp = wb["INPUTS"]

# Clear old content in rows 46-95
for r in range(46, 96):
    for c in range(1, 20):
        inp.cell(r, c).value = None
        inp.cell(r, c).font = Font(name=FONT, size=SZ)
        inp.cell(r, c).fill = PatternFill(fill_type=None)
        inp.cell(r, c).border = Border()

def inp_section(row, label):
    c = inp.cell(row, 2, value=label)
    c.font = fn(bold=True, color=C_WHITE, sz=SZ)
    c.fill = fl(C_NAVY)
    c.alignment = al("left")
    for col in range(3, 8):
        inp.cell(row, col).fill = fl(C_NAVY)
    inp.row_dimensions[row].height = 14

def inp_subhdr(row, label):
    c = inp.cell(row, 2, value=label)
    c.font = fn(bold=True, color=C_BLUE, sz=SZ)
    c.alignment = al("left")
    inp.cell(row, 2).border = bd_bottom()
    inp.row_dimensions[row].height = 12

def inp_row(row, label, value=None, formula=None, is_input=False, note=None):
    inp.row_dimensions[row].height = 14
    c = inp.cell(row, 2, value=label)
    c.font = fn(color=C_SLATE, sz=SZ - 1)
    c.alignment = al("left")
    if value is not None or formula is not None:
        vc = inp.cell(row, 3, value=formula if formula else value)
        vc.font = fn(color=C_INPUT if is_input else C_BLACK, bold=is_input)
        vc.alignment = al("right")
        vc.border = bd()
    if note:
        nc = inp.cell(row, 5, value=note)
        nc.font = fn(italic=True, color=C_SLATE, sz=7)
        nc.alignment = al("left", wrap=True)
    inp.cell(row, 2).border = bd_bottom()

# Row 45: spacer
inp.row_dimensions[45].height = 8

# Row 46: Section header
inp_section(46, "DRIVER TREE ASSUMPTIONS")

# Row 47: Sub-header
inp.row_dimensions[47].height = 10
inp_subhdr(47, "  Model Links  [black = auto-updated from FINANCIAL MODEL]")

# Rows 48-62: Model-linked inputs (formulas)
inp_row(48, "Entry GMV ($B)  [FY2026E]",
        formula="='FINANCIAL MODEL'!G41",
        note="GMV at entry year; drives volume effect calculations")
inp_row(49, "Exit GMV ($B)  [FY2028E]",
        formula="='FINANCIAL MODEL'!I41")
inp_row(50, "Entry Attach Rate (bps)  [FY2026E]",
        formula="='FINANCIAL MODEL'!G43*10000",
        note="Merchant Solutions Rev / GMV, in bps; used as baseline for volume effect")
inp_row(51, "Exit Attach Rate (bps)  [FY2028E]",
        formula="='FINANCIAL MODEL'!I43*10000")
inp_row(52, "Entry GPV Penetration (%)  [FY2026E]",
        formula="='FINANCIAL MODEL'!G42",
        note="% of GMV processed through Shopify Payments")
inp_row(53, "Exit GPV Penetration (%)  [FY2028E]",
        formula="='FINANCIAL MODEL'!I42")
inp_row(54, "Entry MRR ($M)  [FY2026E]",
        formula="='FINANCIAL MODEL'!G44")
inp_row(55, "Exit MRR ($M)  [FY2028E]",
        formula="='FINANCIAL MODEL'!I44")
inp_row(56, "Entry Merchant Solutions Revenue ($M)",
        formula="='FINANCIAL MODEL'!G41*'FINANCIAL MODEL'!G43*1000",
        note="GMV ($B) × Attach Rate (decimal) × 1000 → $M")
inp_row(57, "Exit Merchant Solutions Revenue ($M)",
        formula="='FINANCIAL MODEL'!I41*'FINANCIAL MODEL'!I43*1000")
inp_row(58, "Entry Total Revenue ($M)  [FY2026E]",
        formula="='FINANCIAL MODEL'!G17")
inp_row(59, "Exit Total Revenue ($M)  [FY2028E]",
        formula="='FINANCIAL MODEL'!I17")
inp_row(60, "Entry Subscription Solutions Revenue ($M)",
        formula="=C58-C56",
        note="Total Rev − Merchant Solutions Rev")
inp_row(61, "Exit Subscription Solutions Revenue ($M)",
        formula="=C59-C57")
inp_row(62, "Base GMV ($B)  [FY2025A — for GPV revenue scaling]",
        formula="='FINANCIAL MODEL'!F41",
        note="IC Memo $67M/100bps disclosure is at FY2025A GMV; scale by exit/base ratio")

# Row 63: spacer
inp.row_dimensions[63].height = 6

# Row 64: IC Memo inputs sub-header
inp_subhdr(64, "  IC Memo / Analyst Inputs  [blue = editable]")

inp_row(65, "GPV Revenue per 100bps penetration at Base GMV ($M)",
        value=67, is_input=True,
        note="IC Memo v2 disclosure: each 100bps Shopify Payments penetration = $67M revenue at FY2025A GMV ($378.4B)")

# Row 66: spacer
inp.row_dimensions[66].height = 6

# Row 67: GMV allocations sub-header
inp_subhdr(67, "  GMV Sub-Driver Allocations  [% of incremental GMV — analyst judgment]")
for row, label, val, note in [
    (68, "Same-Cohort Productivity (%)",   0.40, "Existing merchant same-store GMV growth"),
    (69, "New Merchant Acquisition (%)",   0.30, "Net new merchant cohort GMV adds"),
    (70, "Geographic Expansion (Intl) (%)", 0.20, "IC Memo: International GMV +45% YoY Q1'26"),
    (71, "B2B / POS / Channel (%)",        0.10, "IC Memo: B2B +80%, POS +33%, Shop App"),
]:
    inp_row(row, label, value=val, is_input=True, note=note)
    inp.cell(row, 3).number_format = "0%"

inp_row(72, "Allocation Sum Check (must = 100%)",
        formula="=C68+C69+C70+C71",
        note="⚠ Must equal 100% — adjust allocations above if not")
inp.cell(72, 3).number_format = "0%"

# Row 73: spacer
inp.row_dimensions[73].height = 6

# Row 74: EBITDA margins sub-header
inp_subhdr(74, "  Incremental EBITDA Margins  [calibrated to model EBITDA delta]")
for row, label, val, note in [
    (75, "Merchant Solutions Incr. EBITDA Margin (%)", 0.214,
     "37.7% seg. gross margin − 16.3% incremental opex rate; calibrated to model EBITDA delta"),
    (76, "Subscription Solutions Incr. EBITDA Margin (%)", 0.676,
     "82.0% seg. gross margin − 14.4% incremental opex rate"),
]:
    inp_row(row, label, value=val, is_input=True, note=note)
    inp.cell(row, 3).number_format = "0.0%"

# Row 77: note
inp.row_dimensions[77].height = 11
inp.cell(77, 2, value="  Note: Incr. opex rate = ΔOpEx / ΔRevenue = ($4,841M−$3,787M) / ($22,138M−$14,818M) = 14.4%").font = fn(italic=True, color=C_SLATE, sz=7)

# Row 78: spacer
inp.row_dimensions[78].height = 6

# Row 79: Subscription MRR sub-header
inp_subhdr(79, "  Subscription MRR Sub-Drivers  [from model.xlsx Drivers sheet]")
for row, label, val, note in [
    (80, "FY2025A Plus / Enterprise MRR ($M)", 69.7,  "model.xlsx Drivers row 65 col BO"),
    (81, "Entry Plus / Enterprise MRR ($M)  [FY2026E]", 83.6, "model.xlsx Drivers row 65 col BT"),
    (82, "Exit Plus / Enterprise MRR ($M)  [FY2028E]",  114.5, "model.xlsx Drivers row 65 col CD"),
    (83, "FY2025A Core MRR ($M)", 135.3, "model.xlsx Drivers row 70 col BO"),
    (84, "Entry Core MRR ($M)  [FY2026E]", 155.6, "model.xlsx Drivers row 70 col BT"),
    (85, "Exit Core MRR ($M)  [FY2028E]",  176.4, "model.xlsx Drivers row 70 col CD"),
]:
    inp_row(row, label, value=val, is_input=True, note=note)

inp_row(86, "Revenue / MRR Multiplier (implied)",
        formula="=C60/C54",
        note="Entry Sub Revenue ÷ Entry MRR; used to convert incremental MRR → incremental revenue")
inp.cell(86, 3).number_format = '0.0"×"'

# ══════════════════════════════════════════════════════════════════════════════
# 2. REBUILD DRIVER TREE TAB — all formula-based
# ══════════════════════════════════════════════════════════════════════════════
if "DRIVER TREE" in wb.sheetnames:
    del wb["DRIVER TREE"]
dt = wb.create_sheet("DRIVER TREE")

# Position sheet after FINANCIALS
sheets = wb.sheetnames
# Move to after MOIC BRIDGE (before NTB REGISTRY)
idx = sheets.index("MOIC BRIDGE") if "MOIC BRIDGE" in sheets else len(sheets) - 2
wb.move_sheet("DRIVER TREE", offset=idx - len(sheets) + 1)

# Column widths
for col, width in zip("ABCDEFGHIJKLMN",
                      [1.5, 34, 9, 22, 7, 10, 10, 10, 10, 12, 9, 12, 13, 10]):
    dt.column_dimensions[col].width = width
dt.freeze_panes = "F7"

# ── Row 1: Title ───────────────────────────────────────────────────────────────
dt.row_dimensions[1].height = 22
dt.cell(1, 2, "DRIVER TREE  ·  SHOPIFY DEAL WORKBOOK").font = fn(bold=True, color=C_NAVY, sz=11)

dt.row_dimensions[2].height = 13
dt.cell(2, 2,
    "Revenue driver decomposition (FY2026E entry → FY2028E exit) → EBITDA cascade → EV/MOIC attribution"
    "  |  All metric and impact cells are live Excel formulas — change INPUTS or FINANCIAL MODEL and this tab reprices automatically."
).font = fn(italic=True, color=C_SLATE, sz=7)

dt.row_dimensions[3].height = 13
dt.cell(3, 2,
    "Decomposition: GMV volume effect (ΔGMV × entry attach rate 240.9bps) + attach rate improvement (exit GMV × Δbps)."
    "  EBITDA margins: Merchant 21.4%, Sub 67.6% (INPUTS rows 75-76).  EV = EBITDA × INPUTS!C18 (entry multiple, Paasche)."
    "  MOIC = EV / INPUTS!C21."
).font = fn(italic=True, color=C_SLATE, sz=7)
dt.cell(3, 2).alignment = al("left", wrap=True)
dt.row_dimensions[3].height = 22

dt.row_dimensions[4].height = 6

# ── Row 5: Section header ──────────────────────────────────────────────────────
dt.row_dimensions[5].height = 16
for c in range(2, 15):
    dt.cell(5, c).fill = fl(C_NAVY)
dt.cell(5, 2, "REVENUE DRIVER DECOMPOSITION  →  EBITDA  →  EV  →  MOIC"
).font = fn(bold=True, color=C_WHITE, sz=SZ)
dt.cell(5, 2).alignment = al("left")

# ── Row 6: Column headers ──────────────────────────────────────────────────────
dt.row_dimensions[6].height = 34
hdrs = ["Driver", "Source", "KPI Metric", "Unit",
        "FY2025A", "FY2026E\n(Entry)", "FY2028E\n(Exit)", "Δ (28E–26E)",
        "Revenue\nImpact ($M)", "Incr.\nEBITDA %", "EBITDA\nImpact ($M)",
        "EV Impact\n($M, ×entry)", "MOIC\nDelta (×)"]
for i, h in enumerate(hdrs):
    c = dt.cell(6, 2 + i, h)
    c.font = fn(bold=True, color=C_WHITE, sz=7)
    c.fill = fl(C_NAVY)
    c.alignment = Alignment(horizontal="center" if i > 0 else "left", vertical="center", wrap_text=True)
    c.border = bd()

# ── Helper: write a driver row ─────────────────────────────────────────────────
# Columns: B(2)=label, C(3)=source, D(4)=metric, E(5)=unit,
#          F(6)=FY25A, G(7)=FY26E, H(8)=FY28E, I(9)=delta, J(10)=rev,
#          K(11)=EBITDA%, L(12)=EBITDA, M(13)=EV, N(14)=MOIC
def dt_row(ws, r, label, source="", metric="", unit="",
           f_25a=None, f_26e=None, f_28e=None,
           f_delta=None, f_rev=None, f_ebitda_mgn=None, f_ebitda=None, f_ev=None, f_moic=None,
           fmt_metric="#,##0.0",
           indent=0, is_section=False, is_subtotal=False, is_total=False, is_negative=False):
    ws.row_dimensions[r].height = 14 if not is_section else 15

    bg = (C_NAVY if is_section else
          C_TEAL if is_total else
          C_GREY if (is_subtotal or r % 2 == 0) else "FFFFFF")

    txt = (C_WHITE if is_section else
           C_RED if is_negative else
           C_NAVY if is_total else
           C_BLACK)
    bold = is_section or is_total or is_subtotal

    prefix = "    " * indent
    ws.cell(r, 2, prefix + label).font = fn(bold=bold, color=txt if not is_section else C_WHITE)
    ws.cell(r, 2).fill = fl(bg)
    ws.cell(r, 2).alignment = al("left")
    ws.cell(r, 2).border = bd()

    for col, val, clr in [(3, source, C_ORANGE if source == "Analyst" else
                                       C_INDIGO if source == "IC Memo" else C_SLATE),
                           (4, metric, C_SLATE),
                           (5, unit, C_SLATE)]:
        c = ws.cell(r, col, val)
        c.font = fn(color=clr if not is_section else C_WHITE, sz=7)
        c.fill = fl(bg)
        c.alignment = al("center" if col in (3, 5) else "left")
        c.border = bd()

    # Metric value columns F(6), G(7), H(8)
    for col, formula in [(6, f_25a), (7, f_26e), (8, f_28e)]:
        if formula:
            c = ws.cell(r, col, formula)
            c.font = fn(color=C_BLACK if not is_section else C_WHITE)
            c.fill = fl(bg)
            c.alignment = al("right")
            c.border = bd()
            c.number_format = fmt_metric

    # Delta col I(9)
    if f_delta:
        c = ws.cell(r, 9, f_delta)
        c.font = fn(bold=bold, color=txt if not is_section else C_WHITE)
        c.fill = fl(bg)
        c.alignment = al("right")
        c.border = bd()
        c.number_format = fmt_metric

    # Impact columns J(10) through N(14)
    impact_fmts = ['#,##0;(#,##0);"-"', "0.0%", '#,##0;(#,##0);"-"',
                   '#,##0;(#,##0);"-"', '+0.000x;-0.000x;"-"']
    for col, formula, fmt in zip(range(10, 15),
                                  [f_rev, f_ebitda_mgn, f_ebitda, f_ev, f_moic],
                                  impact_fmts):
        if formula:
            c = ws.cell(r, col, formula)
            c.font = fn(bold=bold, color=txt if not is_section else C_WHITE)
            c.fill = fl(bg)
            c.alignment = al("right")
            c.border = bd()
            c.number_format = fmt

# ── ROW LAYOUT ─────────────────────────────────────────────────────────────────

# Row 7: MERCHANT SOLUTIONS header
dt_row(dt, 7, "MERCHANT SOLUTIONS REVENUE", source="Model",
       metric="Revenue = GMV × Attach Rate", unit="$M",
       f_26e="=INPUTS!$C$56", f_28e="=INPUTS!$C$57",
       f_delta="=H7-G7",
       is_section=True)

# Row 8: GMV Volume Effect sub-header
dt.row_dimensions[8].height = 12
dt.cell(8, 2, "  GMV Growth → Volume Effect").font = fn(bold=True, color=C_BLUE, sz=SZ - 1)
dt.cell(8, 4, "Δ GMV ($B) × entry attach rate (bps); sub-driver splits = analyst allocations in INPUTS rows 68-71"
        ).font = fn(italic=True, color=C_SLATE, sz=7)
dt.cell(8, 4).alignment = al("left", wrap=True)
for c in range(2, 15):
    dt.cell(8, c).border = bd_bottom()

# Rows 9-12: GMV sub-drivers (each uses a different allocation % from INPUTS rows 68-71)
gmv_rows = [
    (9,  "Same-Cohort Productivity",               "Analyst",  "GMV attributed $B (40% alloc. of ΔGMV)",           "INPUTS!$C$68"),
    (10, "New Merchant Acquisition",               "Analyst",  "GMV attributed $B (30% alloc. of ΔGMV)",           "INPUTS!$C$69"),
    (11, "Geographic Expansion (International)",   "IC Memo",  "GMV attributed $B (20% alloc.) — Intl +45% YoY",   "INPUTS!$C$70"),
    (12, "B2B / POS / Channel",                    "IC Memo",  "GMV attributed $B (10% alloc.) — B2B +80% YoY",    "INPUTS!$C$71"),
]
for r, label, src, metric, pct_ref in gmv_rows:
    dt_row(dt, r, label, source=src, metric=metric, unit="$B",
           f_25a=f"='FINANCIAL MODEL'!F41*{pct_ref}",
           f_26e=f"=INPUTS!$C$48*{pct_ref}",
           f_28e=f"=INPUTS!$C$49*{pct_ref}",
           f_delta=f"=H{r}-G{r}",
           f_rev=f"=I{r}*(INPUTS!$C$50/10000)*1000",
           f_ebitda_mgn="=INPUTS!$C$75",
           f_ebitda=f"=J{r}*K{r}",
           f_ev=f"=L{r}*INPUTS!$C$18",
           f_moic=f"=M{r}/INPUTS!$C$21",
           fmt_metric="#,##0.0",
           indent=2)

# Row 13: GMV Volume subtotal
dt_row(dt, 13, "  Subtotal: GMV Volume Effect", source="",
       metric="Total ΔGMV × entry attach rate", unit="$B",
       f_26e="=INPUTS!$C$48", f_28e="=INPUTS!$C$49",
       f_delta="=H13-G13",
       f_rev="=SUM(J9:J12)",
       f_ebitda_mgn="=INPUTS!$C$75",
       f_ebitda="=J13*K13",
       f_ev="=L13*INPUTS!$C$18",
       f_moic="=M13/INPUTS!$C$21",
       fmt_metric="#,##0.0",
       indent=1, is_subtotal=True)

# Row 14: Attach Rate sub-header
dt.row_dimensions[14].height = 12
dt.cell(14, 2, "  Attach Rate Improvement").font = fn(bold=True, color=C_BLUE, sz=SZ - 1)
dt.cell(14, 4, "Δ attach rate (bps) × FY2028E exit GMV ($B) → revenue in $M"
        ).font = fn(italic=True, color=C_SLATE, sz=7)
for c in range(2, 15):
    dt.cell(14, c).border = bd_bottom()

# Row 15: GPV Penetration
dt_row(dt, 15, "GPV Penetration (71% → 79%)", source="Model + IC Memo",
       metric="GPV% of GMV — IC Memo: $67M per 100bps at base GMV ($378.4B), scaled to exit GMV",
       unit="%",
       f_25a="='FINANCIAL MODEL'!F42",
       f_26e="=INPUTS!$C$52",
       f_28e="=INPUTS!$C$53",
       f_delta="=H15-G15",
       f_rev="=(H15-G15)*100*INPUTS!$C$65*INPUTS!$C$49/INPUTS!$C$62",
       f_ebitda_mgn="=INPUTS!$C$75",
       f_ebitda="=J15*K15",
       f_ev="=L15*INPUTS!$C$18",
       f_moic="=M15/INPUTS!$C$21",
       fmt_metric="0.0%",
       indent=2)

# Row 16: Non-Payments Attach (residual = total attach effect - GPV portion)
dt_row(dt, 16, "Non-Payments Attach (MCA, Balance, Markets)", source="Model",
       metric="Residual attach rate improvement after GPV component; Shopify Capital, Balance, Markets",
       unit="bps residual",
       f_rev="=J17-J15",
       f_ebitda_mgn="=INPUTS!$C$75",
       f_ebitda="=J16*K16",
       f_ev="=L16*INPUTS!$C$18",
       f_moic="=M16/INPUTS!$C$21",
       indent=2)

# Row 17: Attach Rate subtotal
dt_row(dt, 17, "  Subtotal: Attach Rate Improvement", source="",
       metric="Total merch rev delta − GMV volume effect", unit="bps",
       f_26e="=INPUTS!$C$50", f_28e="=INPUTS!$C$51",
       f_delta="=H17-G17",
       f_rev="=INPUTS!$C$57-INPUTS!$C$56-J13",
       f_ebitda_mgn="=INPUTS!$C$75",
       f_ebitda="=J17*K17",
       f_ev="=L17*INPUTS!$C$18",
       f_moic="=M17/INPUTS!$C$21",
       fmt_metric="#,##0.0",
       indent=1, is_subtotal=True)

# Row 18: Merchant Solutions total
dt_row(dt, 18, "TOTAL: Merchant Solutions Revenue", source="Model",
       unit="$M",
       f_26e="=INPUTS!$C$56", f_28e="=INPUTS!$C$57",
       f_delta="=H18-G18",
       f_rev="=J13+J17",
       f_ebitda_mgn="=INPUTS!$C$75",
       f_ebitda="=L18/J18",        # show implied blended margin
       f_ev="=J18*INPUTS!$C$75*INPUTS!$C$18",
       f_moic="=J18*INPUTS!$C$75*INPUTS!$C$18/INPUTS!$C$21",
       fmt_metric="#,##0",
       is_total=True)
# Fix EBITDA and EV for total row (sum, not multiply)
dt.cell(18, 12, "=L13+L17").number_format = '#,##0;(#,##0);"-"'
dt.cell(18, 12).font = fn(bold=True, color=C_NAVY)
dt.cell(18, 12).fill = fl(C_TEAL)
dt.cell(18, 12).border = bd()
dt.cell(18, 12).alignment = al("right")
dt.cell(18, 13, "=L18*INPUTS!$C$18").number_format = '#,##0;(#,##0);"-"'
dt.cell(18, 13).font = fn(bold=True, color=C_NAVY)
dt.cell(18, 13).fill = fl(C_TEAL)
dt.cell(18, 13).border = bd()
dt.cell(18, 13).alignment = al("right")
dt.cell(18, 14, "=M18/INPUTS!$C$21").number_format = '+0.000x;-0.000x;"-"'
dt.cell(18, 14).font = fn(bold=True, color=C_NAVY)
dt.cell(18, 14).fill = fl(C_TEAL)
dt.cell(18, 14).border = bd()
dt.cell(18, 14).alignment = al("right")
# Override EBITDA margin col to show $, not implied %
dt.cell(18, 11, "=L18").number_format = '#,##0;(#,##0);"-"'
dt.cell(18, 11).font = fn(bold=True, color=C_NAVY)
dt.cell(18, 11).fill = fl(C_TEAL)
dt.cell(18, 11).border = bd()
dt.cell(18, 11).alignment = al("right")

# Row 19: spacer
dt.row_dimensions[19].height = 6

# Row 20: SUBSCRIPTION SOLUTIONS header
dt_row(dt, 20, "SUBSCRIPTION SOLUTIONS REVENUE", source="Model",
       metric="Revenue = MRR × implied multiplier (C86)", unit="$M",
       f_26e="=INPUTS!$C$60", f_28e="=INPUTS!$C$61",
       f_delta="=H20-G20",
       is_section=True)

# Row 21: note
dt.row_dimensions[21].height = 11
dt.cell(21, 2, "    MRR Build: MRR × Revenue/MRR multiplier (INPUTS!C86) = implied annual revenue per $1M MRR"
        ).font = fn(italic=True, color=C_SLATE, sz=7)

# Rows 22-23: MRR sub-drivers
dt_row(dt, 22, "Plus / Enterprise MRR Growth", source="Model",
       metric="Plus MRR ($M) — $25M+ GMV cohort, fastest-growing Q1'26; 28% → 39% of MRR",
       unit="$M MRR",
       f_25a="=INPUTS!$C$80",
       f_26e="=INPUTS!$C$81",
       f_28e="=INPUTS!$C$82",
       f_delta="=H22-G22",
       f_rev="=I22*INPUTS!$C$86",
       f_ebitda_mgn="=INPUTS!$C$76",
       f_ebitda="=J22*K22",
       f_ev="=L22*INPUTS!$C$18",
       f_moic="=M22/INPUTS!$C$21",
       fmt_metric="#,##0.0",
       indent=2)

dt_row(dt, 23, "Core SMB MRR Growth", source="Model",
       metric="Core MRR ($M) — Basic/Standard/Advanced plan merchants",
       unit="$M MRR",
       f_25a="=INPUTS!$C$83",
       f_26e="=INPUTS!$C$84",
       f_28e="=INPUTS!$C$85",
       f_delta="=H23-G23",
       f_rev="=I23*INPUTS!$C$86",
       f_ebitda_mgn="=INPUTS!$C$76",
       f_ebitda="=J23*K23",
       f_ev="=L23*INPUTS!$C$18",
       f_moic="=M23/INPUTS!$C$21",
       fmt_metric="#,##0.0",
       indent=2)

# Row 24: Subscription total
dt_row(dt, 24, "TOTAL: Subscription Solutions Revenue", source="Model",
       unit="$M",
       f_26e="=INPUTS!$C$60", f_28e="=INPUTS!$C$61",
       f_delta="=H24-G24",
       f_rev="=SUM(J22:J23)",
       f_ebitda_mgn="=INPUTS!$C$76",
       f_ebitda="=SUM(L22:L23)",
       f_ev="=L24*INPUTS!$C$18",
       f_moic="=M24/INPUTS!$C$21",
       fmt_metric="#,##0",
       is_total=True)
# Fix EBITDA col to show $ not %, EV/MOIC normal
dt.cell(24, 11, "=SUM(L22:L23)").number_format = '#,##0;(#,##0);"-"'
dt.cell(24, 11).font = fn(bold=True, color=C_NAVY)
dt.cell(24, 11).fill = fl(C_TEAL)
dt.cell(24, 11).border = bd()
dt.cell(24, 11).alignment = al("right")

# Row 25: spacer
dt.row_dimensions[25].height = 6

# Row 26: TOTAL REVENUE
dt_row(dt, 26, "TOTAL REVENUE GROWTH", source="Model",
       unit="$M",
       f_26e="=INPUTS!$C$58", f_28e="=INPUTS!$C$59",
       f_delta="=H26-G26",
       f_rev="=J18+J24",
       f_ebitda_mgn="=(L18+L24)/J26",
       f_ebitda="=L18+L24",
       f_ev="=M18+M24",
       f_moic="=N18+N24",
       fmt_metric="#,##0",
       is_total=True)
# Teal background on entire row
for c in range(2, 15):
    dt.cell(26, c).fill = fl(C_TEAL)
    if not dt.cell(26, c).value:
        dt.cell(26, c).border = bd()

# Row 27: spacer
dt.row_dimensions[27].height = 8

# ── VALUE CREATION BRIDGE ──────────────────────────────────────────────────────
# Row 28: Section header
dt.row_dimensions[28].height = 16
for c in range(2, 15):
    dt.cell(28, c).fill = fl(C_NAVY)
    dt.cell(28, c).border = bd()
dt.cell(28, 2, "VALUE CREATION BRIDGE  (Revenue drivers + Multiple compression + Net cash = Total MOIC)"
        ).font = fn(bold=True, color=C_WHITE, sz=SZ)
dt.cell(28, 2).alignment = al("left")

# Row 29: Bridge column headers
dt.row_dimensions[29].height = 28
bridge_hdrs = {2: "Component", 10: "EBITDA\nEffect ($M)", 11: "×\nMultiple",
               12: "EV\nImpact ($M)", 14: "MOIC\nDelta (×)"}
for col, lbl in bridge_hdrs.items():
    c = dt.cell(29, col, lbl)
    c.font = fn(bold=True, color=C_WHITE, sz=7)
    c.fill = fl(C_NAVY)
    c.alignment = Alignment(horizontal="center" if col > 2 else "left", vertical="center", wrap_text=True)
    c.border = bd()
for c in range(3, 15):
    if c not in bridge_hdrs:
        dt.cell(29, c).fill = fl(C_NAVY)
        dt.cell(29, c).border = bd()

# Row 30: Revenue-driven
dt.row_dimensions[30].height = 14
bg30 = "FFFFFF"
set_cell(dt, 30, 2, "Revenue-Driven EBITDA Growth  (EBITDA delta × entry multiple 48.5×)", bold=True, bg=bg30, h="left")
set_cell(dt, 30, 10, "=L26", bold=True, bg=bg30, h="right", fmt='#,##0;(#,##0);"-"')
set_cell(dt, 30, 11, "=INPUTS!$C$18", bold=True, bg=bg30, h="right", fmt='0.0"×"')
set_cell(dt, 30, 12, "=J30*K30", bold=True, bg=bg30, h="right", fmt='#,##0;(#,##0);"-"')
set_cell(dt, 30, 14, "=L30/INPUTS!$C$21", bold=True, bg=bg30, h="right", fmt='+0.000x;-0.000x;"-"')
for c in range(3, 15):
    if not dt.cell(30, c).value:
        dt.cell(30, c).fill = fl(bg30); dt.cell(30, c).border = bd()

# Row 31: Multiple compression
dt.row_dimensions[31].height = 14
bg31 = C_GREY
set_cell(dt, 31, 2, "Multiple Compression  (48.5× entry → 45.0× exit) × exit EBITDA $4,687M", bold=True, color=C_RED, bg=bg31, h="left")
set_cell(dt, 31, 10, "=(INPUTS!$C$27-INPUTS!$C$18)*INPUTS!$C$26", bold=True, color=C_RED, bg=bg31, h="right", fmt='#,##0;(#,##0)')
set_cell(dt, 31, 11, 1.0, bold=True, bg=bg31, h="right", fmt='0.0"×"')
set_cell(dt, 31, 12, "=J31", bold=True, color=C_RED, bg=bg31, h="right", fmt='#,##0;(#,##0)')
set_cell(dt, 31, 14, "=L31/INPUTS!$C$21", bold=True, color=C_RED, bg=bg31, h="right", fmt='+0.000x;-0.000x')
for c in range(3, 15):
    if not dt.cell(31, c).value:
        dt.cell(31, c).fill = fl(bg31); dt.cell(31, c).border = bd()

# Row 32: Net cash build
dt.row_dimensions[32].height = 14
bg32 = "FFFFFF"
set_cell(dt, 32, 2, "Net Cash Build  (Net cash $6.3B → $12.4B; INPUTS rows 20, 29)", bold=True, color=C_GREEN, bg=bg32, h="left")
set_cell(dt, 32, 10, "=-(INPUTS!$C$29-INPUTS!$C$20)", bold=True, color=C_GREEN, bg=bg32, h="right", fmt='#,##0;(#,##0)')
set_cell(dt, 32, 11, 1.0, bold=True, bg=bg32, h="right", fmt='0.0"×"')
set_cell(dt, 32, 12, "=J32", bold=True, color=C_GREEN, bg=bg32, h="right", fmt='#,##0;(#,##0)')
set_cell(dt, 32, 14, "=L32/INPUTS!$C$21", bold=True, color=C_GREEN, bg=bg32, h="right", fmt='+0.000x;-0.000x')
for c in range(3, 15):
    if not dt.cell(32, c).value:
        dt.cell(32, c).fill = fl(bg32); dt.cell(32, c).border = bd()

# Row 33: TOTAL VALUE CREATION
dt.row_dimensions[33].height = 16
for c in range(2, 15):
    dt.cell(33, c).fill = fl(C_NAVY); dt.cell(33, c).border = bd()
dt.cell(33, 2, "TOTAL VALUE CREATION").font = fn(bold=True, color=C_WHITE, sz=SZ)
dt.cell(33, 2).alignment = al("left")
set_cell(dt, 33, 12, "=L30+L31+L32", bold=True, color=C_WHITE, bg=C_NAVY, h="right",
         fmt='#,##0;(#,##0)')
set_cell(dt, 33, 14, "=L33/INPUTS!$C$21", bold=True, color=C_WHITE, bg=C_NAVY, h="right",
         fmt='+0.000x;-0.000x')

# Row 34: MOIC CHECK
dt.row_dimensions[34].height = 20
for c in range(2, 15):
    dt.cell(34, c).fill = fl(C_TEAL); dt.cell(34, c).border = bd()
dt.cell(34, 2, "MOIC CHECK  =  1.00× + total attribution  (should = INPUTS!C34)"
        ).font = fn(bold=True, color=C_NAVY, sz=SZ)
dt.cell(34, 2).alignment = al("left")
c34 = dt.cell(34, 14, "=1+N33")
c34.font = fn(bold=True, color=C_NAVY, sz=13)
c34.fill = fl(C_TEAL)
c34.alignment = al("right")
c34.number_format = '0.00"×"'
c34.border = bd()
# Also show the INPUTS MOIC for comparison
dt.cell(34, 12, "=INPUTS!$C$34").font = fn(bold=True, color=C_SLATE, sz=9)
dt.cell(34, 12).fill = fl(C_TEAL)
dt.cell(34, 12).alignment = al("right")
dt.cell(34, 12).number_format = '0.00"×"'
dt.cell(34, 12).border = bd()
dt.cell(34, 11, "INPUTS MOIC:").font = fn(italic=True, color=C_SLATE, sz=7)
dt.cell(34, 11).fill = fl(C_TEAL)
dt.cell(34, 11).alignment = al("right")
dt.cell(34, 11).border = bd()

# Row 35: Footnote
dt.row_dimensions[35].height = 22
dt.cell(35, 2,
    "Footnote: MOIC check uses Paasche decomposition (EV = EBITDA delta × entry multiple 48.5×, same as INPUTS!C56 pool). "
    "Multiple compression uses (exit mult − entry mult) × exit EBITDA (Paasche price effect). "
    "GMV sub-driver %s are analyst allocations (INPUTS rows 68-71); change them and all downstream cells reprice. "
    "EBITDA margins (rows 75-76) calibrated to model delta; recalibrate if FINANCIAL MODEL assumptions change."
).font = fn(italic=True, color=C_SLATE, sz=7)
dt.cell(35, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ══════════════════════════════════════════════════════════════════════════════
# 3. FIX KPI TREE — rows 9 and 10 to use INPUTS formulas
# ══════════════════════════════════════════════════════════════════════════════
kpi = wb["KPI TREE"]

# Row 9: Subscription Solutions — Budget (G9) and Actual (H9)
# Budget = INPUTS!C60 (entry sub rev, which links to FINANCIAL MODEL)
kpi.cell(9, 7, "=INPUTS!C60")   # G = Budget FY2026E
kpi.cell(9, 7).number_format = "#,##0"
kpi.cell(9, 7).alignment = al("right")
# Actual FY2025A = total rev - merch solutions rev
kpi.cell(9, 8, "='FINANCIAL MODEL'!F17-'FINANCIAL MODEL'!F41*'FINANCIAL MODEL'!F43*1000")
kpi.cell(9, 8).number_format = "#,##0"
kpi.cell(9, 8).alignment = al("right")

# Row 10: Merchant Solutions — Budget (G10) and Actual (H10)
kpi.cell(10, 7, "=INPUTS!C56")  # G = Budget FY2026E (entry merch rev)
kpi.cell(10, 7).number_format = "#,##0"
kpi.cell(10, 7).alignment = al("right")
kpi.cell(10, 8, "='FINANCIAL MODEL'!F41*'FINANCIAL MODEL'!F43*1000")
kpi.cell(10, 8).number_format = "#,##0"
kpi.cell(10, 8).alignment = al("right")

# Fix variance formulas if they exist (cols I and J = $ and % variance)
for r in [9, 10]:
    for col, formula, fmt in [
        (9,  f"=IFERROR(G{r}-H{r}," + '"-")', '#,##0;(#,##0);"-"'),
        (10, f"=IFERROR((G{r}-H{r})/H{r}," + '"-")', '+0.0%;-0.0%;"-"'),
    ]:
        kpi.cell(r, col, formula).number_format = fmt
        kpi.cell(r, col).alignment = al("right")

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(WB_OUT)
print(f"Saved: {WB_OUT}")
print()
print("Formula architecture summary:")
print("  FINANCIAL MODEL → INPUTS rows 48-62 (auto-linked)")
print("  INPUTS rows 65-85 (blue = editable analyst inputs)")
print("  DRIVER TREE metric cols (F/G/H) = INPUTS formula refs")
print("  DRIVER TREE impact cols (J/L/M/N) = chain formulas from metric cols + INPUTS rates")
print("  MOIC check (N34) = 1 + N33; compare to INPUTS!C34 shown in L34")
print("  KPI TREE rows 9-10: Sub/Merch Solutions Revenue now formula-linked to INPUTS")
