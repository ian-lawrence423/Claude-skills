"""
Shopify Deal Workbook v2 rebuild script.
Goals:
  1. Convert hardcodes to formulas (INPUTS driver attribution, DRIVER TREE, NTB REGISTRY, KPI TREE)
  2. Apply McKinsey/Pattern formatting system
"""
import shutil, os
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\IanLawrence\github\Claude-skills\docs\Shopify_Deal_Workbook.xlsx"
TMP = r"C:\tmp\Shopify_Deal_Workbook_v2.xlsx"
DST = r"C:\Users\IanLawrence\github\Claude-skills\docs\Shopify_Deal_Workbook.xlsx"

# ── colour palette ──────────────────────────────────────────────────────────
NAV_FILL   = "0F4761"
SUB_FILL   = "E8EDF2"
BODY_BG    = "FFFFFF"
ALT_BG     = "F7F8FA"
TOTAL_BG   = "EEF2F7"
GREEN_BG   = "EBF4EA"; GREEN_TXT  = "375623"
ORANGE_BG  = "FEF3E8"; ORANGE_TXT = "C55A11"
RED_BG     = "FDEBEB"; RED_TXT    = "C00000"
HDR_TXT    = "FFFFFF"
BODY_TXT   = "1F2937"
LABEL_TXT  = "374151"
SEC_TXT    = "9CA3AF"
INPUT_TXT  = "1D4ED8"
FORM_TXT   = "000000"
LINK_TXT   = "047857"
T1_BG      = "EBF4EA"
T2_BG      = "EBF4EA"
T3_BG      = "FEF3E8"
T4_BG      = "FDEBEB"

def solid(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def mkfont(bold=False, size=9, color=BODY_TXT, italic=False, name="Arial"):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def thin_bottom(color="E5E7EB"):
    return Border(bottom=Side(style="thin", color=color))

def thick_bottom(color=NAV_FILL):
    return Border(bottom=Side(style="medium", color=color))

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_cell(cell, bold=False, size=9, fg=BODY_TXT, fill=None,
               italic=False, ha="left", border=None, num_fmt=None):
    cell.font = mkfont(bold=bold, size=size, color=fg)
    if fill:
        cell.fill = solid(fill)
    else:
        cell.fill = PatternFill(fill_type=None)
    cell.alignment = align(h=ha)
    if border:
        cell.border = border
    if num_fmt:
        cell.number_format = num_fmt

# ── apply standard row styling ───────────────────────────────────────────────
def style_header_row(ws, row, cols, title=None):
    """Navy header row spanning cols"""
    for c in cols:
        cell = ws.cell(row=row, column=c)
        style_cell(cell, bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
                   ha="left", border=thick_bottom(NAV_FILL))
    ws.row_dimensions[row].height = 18
    if title is not None:
        ws.cell(row=row, column=cols[0]).value = title

def style_subheader_row(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        style_cell(cell, bold=True, size=9, fg=LABEL_TXT, fill=SUB_FILL,
                   border=thin_bottom("9CA3AF"))
    ws.row_dimensions[row].height = 16

def style_body_row(ws, row, cols, alt=False):
    bg = ALT_BG if alt else BODY_BG
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = solid(bg)
        if not cell.font or cell.font.color.rgb in ("00000000", "FF000000",
                                                     "00000000"):
            pass  # preserve existing font color (input blue / formula black)
        cell.border = thin_bottom()
    ws.row_dimensions[row].height = 15

def style_total_row(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = solid(TOTAL_BG)
        cell.border = thick_bottom()
    ws.row_dimensions[row].height = 15

def style_spacer_row(ws, row):
    ws.row_dimensions[row].height = 12
    for c in range(1, ws.max_column + 2):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill(fill_type=None)
        cell.border = Border()

# ══════════════════════════════════════════════════════════════════════════════
# LOAD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = load_workbook(SRC)

# Remove all tab colours
for sn in wb.sheetnames:
    wb[sn].sheet_properties.tabColor = None

# ══════════════════════════════════════════════════════════════════════════════
# INPUTS TAB — add driver attribution section rows 50–88
# ══════════════════════════════════════════════════════════════════════════════
ws_inp = wb["INPUTS"]

# ── Section title row 50 ─────────────────────────────────────────────────────
ws_inp.cell(50, 2).value = "Driver Attribution Assumptions"
style_cell(ws_inp.cell(50, 2), bold=True, size=11, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
ws_inp.cell(50, 3).fill = solid(NAV_FILL)
ws_inp.cell(50, 3).border = thick_bottom()
ws_inp.row_dimensions[50].height = 18

# spacer 51
style_spacer_row(ws_inp, 51)

# ── Sub-section: Model-Derived Parameters ───────────────────────────────────
ws_inp.cell(52, 2).value = "Model-Derived Parameters"
style_cell(ws_inp.cell(52, 2), bold=True, size=9, fg=LABEL_TXT, fill=SUB_FILL,
           border=thin_bottom("9CA3AF"))
ws_inp.cell(52, 3).fill = solid(SUB_FILL)
ws_inp.cell(52, 3).border = thin_bottom("9CA3AF")
ws_inp.row_dimensions[52].height = 16

# col headers row 53
for c, txt in [(2, "Parameter"), (3, "Value ($M)")]:
    style_cell(ws_inp.cell(53, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL), ha="left")
    ws_inp.cell(53, c).value = txt
ws_inp.row_dimensions[53].height = 18

model_params = [
    (54, "EBITDA Growth ($M)",          "=C26-C16"),
    (55, "Entry EV/EBITDA Multiple",    "=C18"),
    (56, "EBITDA-Driven EV Pool ($M)",  "=(C26-C16)*C18"),
    (57, "Multiple Compression Effect ($M)", "=(C27-C18)*C26"),
    (58, "Net Cash Build ($M)",         "=-(C29-C20)"),
    (59, "Total Value Creation ($M)",   "=C30-C21"),
]
for i, (r, lbl, frm) in enumerate(model_params):
    alt = (i % 2 == 1)
    bg = ALT_BG if alt else BODY_BG
    ws_inp.cell(r, 2).value = lbl
    style_cell(ws_inp.cell(r, 2), fg=LABEL_TXT, fill=bg, border=thin_bottom())
    ws_inp.cell(r, 3).value = frm
    style_cell(ws_inp.cell(r, 3), fg=FORM_TXT, fill=bg, border=thin_bottom(),
               ha="right", num_fmt='#,##0')
    ws_inp.row_dimensions[r].height = 15

# spacer 60, 61
for r in (60, 61):
    style_spacer_row(ws_inp, r)

# ── Sub-section: Revenue Driver Attribution ──────────────────────────────────
ws_inp.cell(61, 2).value = "Revenue Driver Attribution"
style_cell(ws_inp.cell(61, 2), bold=True, size=9, fg=LABEL_TXT, fill=SUB_FILL,
           border=thin_bottom("9CA3AF"))
ws_inp.cell(61, 3).fill = solid(SUB_FILL)
ws_inp.cell(61, 3).border = thin_bottom("9CA3AF")
for c in [4,5,6]:
    ws_inp.cell(61, c).fill = solid(SUB_FILL)
    ws_inp.cell(61, c).border = thin_bottom("9CA3AF")
ws_inp.row_dimensions[61].height = 16

# col headers row 62
for c, txt in [(2,"Driver"), (3,"% Alloc"), (4,"Base EV ($M)"), (5,"Upside Factor"), (6,"Upside EV ($M)")]:
    style_cell(ws_inp.cell(62, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL), ha="left" if c==2 else "right")
    ws_inp.cell(62, c).value = txt
ws_inp.row_dimensions[62].height = 18

rev_drivers = [
    (63, "GMV Same-Cohort (B1a)",           0.272, 1.40),
    (64, "New Merchant Acquisition (A1)",   0.163, 1.47),
    (65, "Geographic ex-NA (B1e)",          0.087, 1.75),
    (66, "GPV Penetration (B2a)",           0.131, 1.67),
    (67, "Payments Take Rate (B2b)",        0.065, 1.67),
    (68, "Capital / MCA (B2c)",             0.022, 2.50),
    (69, "Enterprise Plus MRR (A2)",        0.044, 2.00),
    (70, "Enterprise GMV Concentration (B1c)", 0.022, 2.50),
]
for i, (r, lbl, alloc, factor) in enumerate(rev_drivers):
    alt = (i % 2 == 1)
    bg = ALT_BG if alt else BODY_BG
    ws_inp.cell(r, 2).value = lbl
    style_cell(ws_inp.cell(r, 2), fg=LABEL_TXT, fill=bg, border=thin_bottom())
    ws_inp.cell(r, 3).value = alloc
    style_cell(ws_inp.cell(r, 3), fg=INPUT_TXT, fill=bg, border=thin_bottom(),
               ha="right", num_fmt='0.0%')
    ws_inp.cell(r, 4).value = f"=C{r}*$C$56"
    style_cell(ws_inp.cell(r, 4), fg=FORM_TXT, fill=bg, border=thin_bottom(),
               ha="right", num_fmt='#,##0')
    ws_inp.cell(r, 5).value = factor
    style_cell(ws_inp.cell(r, 5), fg=INPUT_TXT, fill=bg, border=thin_bottom(),
               ha="right", num_fmt='0.00x')
    ws_inp.cell(r, 6).value = f"=D{r}*E{r}"
    style_cell(ws_inp.cell(r, 6), fg=FORM_TXT, fill=bg, border=thin_bottom(),
               ha="right", num_fmt='#,##0')
    ws_inp.row_dimensions[r].height = 15

# spacer 71, sub-section header 72
style_spacer_row(ws_inp, 71)

ws_inp.cell(72, 2).value = "Margin Driver Attribution"
style_cell(ws_inp.cell(72, 2), bold=True, size=9, fg=LABEL_TXT, fill=SUB_FILL,
           border=thin_bottom("9CA3AF"))
for c in [3,4,5,6]:
    ws_inp.cell(72, c).fill = solid(SUB_FILL)
    ws_inp.cell(72, c).border = thin_bottom("9CA3AF")
ws_inp.row_dimensions[72].height = 16

# col headers row 73
for c, txt in [(2,"Driver"), (3,"bps"), (4,"Base EV ($M)"), (5,"Factor"), (6,"Upside EV ($M)")]:
    style_cell(ws_inp.cell(73, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL), ha="left" if c==2 else "right")
    ws_inp.cell(73, c).value = txt
ws_inp.row_dimensions[73].height = 18

margin_drivers = [
    (74, "OpEx Leverage (bps improvement)", 240, "=(C74/10000)*$C$25*$C$27", 1.75),
    (75, "Sub Solutions Gross Margin (bps)", 60,  "=(C75/10000)*$C$25*$C$27", 1.50),
]
for i, (r, lbl, bps, base_frm, factor) in enumerate(margin_drivers):
    alt = (i % 2 == 1)
    bg = ALT_BG if alt else BODY_BG
    ws_inp.cell(r, 2).value = lbl
    style_cell(ws_inp.cell(r, 2), fg=LABEL_TXT, fill=bg, border=thin_bottom())
    ws_inp.cell(r, 3).value = bps
    style_cell(ws_inp.cell(r, 3), fg=INPUT_TXT, fill=bg, border=thin_bottom(),
               ha="right", num_fmt='#,##0')
    ws_inp.cell(r, 4).value = base_frm
    style_cell(ws_inp.cell(r, 4), fg=FORM_TXT, fill=bg, border=thin_bottom(),
               ha="right", num_fmt='#,##0')
    ws_inp.cell(r, 5).value = factor
    style_cell(ws_inp.cell(r, 5), fg=INPUT_TXT, fill=bg, border=thin_bottom(),
               ha="right", num_fmt='0.00x')
    ws_inp.cell(r, 6).value = f"=D{r}*E{r}"
    style_cell(ws_inp.cell(r, 6), fg=FORM_TXT, fill=bg, border=thin_bottom(),
               ha="right", num_fmt='#,##0')
    ws_inp.row_dimensions[r].height = 15

# spacer 76, sub-section header 77
style_spacer_row(ws_inp, 76)

ws_inp.cell(77, 2).value = "Formula-Derived & Asymmetric Drivers"
style_cell(ws_inp.cell(77, 2), bold=True, size=9, fg=LABEL_TXT, fill=SUB_FILL,
           border=thin_bottom("9CA3AF"))
for c in [3,4,5,6]:
    ws_inp.cell(77, c).fill = solid(SUB_FILL)
    ws_inp.cell(77, c).border = thin_bottom("9CA3AF")
ws_inp.row_dimensions[77].height = 16

# col headers row 78
for c, txt in [(2,"Driver"), (3,"Input ($M)"), (4,"Base EV ($M)"), (5,""), (6,"Upside EV ($M)")]:
    style_cell(ws_inp.cell(78, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL), ha="left" if c==2 else "right")
    ws_inp.cell(78, c).value = txt
ws_inp.row_dimensions[78].height = 18

asym_drivers = [
    (79, "Multiple Compression / Re-rating", None,  "=(C27-C18)*C26", 5000),
    (80, "Macro / Consumer Risk ($M)",        -5000, "=C80",           2000),
    (81, "AI / Sidekick-C1 (upside only)",    None,  0,                15000),
    (82, "AI / UCP-C3 (upside only)",         None,  0,                25000),
]
for i, (r, lbl, inp_val, base_val, upside_val) in enumerate(asym_drivers):
    alt = (i % 2 == 1)
    bg = ALT_BG if alt else BODY_BG
    ws_inp.cell(r, 2).value = lbl
    style_cell(ws_inp.cell(r, 2), fg=LABEL_TXT, fill=bg, border=thin_bottom())
    if inp_val is not None:
        ws_inp.cell(r, 3).value = inp_val
        style_cell(ws_inp.cell(r, 3), fg=INPUT_TXT, fill=bg, border=thin_bottom(),
                   ha="right", num_fmt='#,##0')
    else:
        ws_inp.cell(r, 3).value = None
        ws_inp.cell(r, 3).fill = solid(bg)
        ws_inp.cell(r, 3).border = thin_bottom()
    if isinstance(base_val, str):
        ws_inp.cell(r, 4).value = base_val
        style_cell(ws_inp.cell(r, 4), fg=FORM_TXT, fill=bg, border=thin_bottom(),
                   ha="right", num_fmt='#,##0')
    else:
        ws_inp.cell(r, 4).value = base_val
        style_cell(ws_inp.cell(r, 4), fg=FORM_TXT, fill=bg, border=thin_bottom(),
                   ha="right", num_fmt='#,##0')
    ws_inp.cell(r, 5).fill = solid(bg)
    ws_inp.cell(r, 5).border = thin_bottom()
    ws_inp.cell(r, 6).value = upside_val
    style_cell(ws_inp.cell(r, 6), fg=INPUT_TXT, fill=bg, border=thin_bottom(),
               ha="right", num_fmt='#,##0')
    ws_inp.row_dimensions[r].height = 15

# spacer 83, 84
style_spacer_row(ws_inp, 83)

# ── Attribution Check ─────────────────────────────────────────────────────────
ws_inp.cell(84, 2).value = "Attribution Check"
style_cell(ws_inp.cell(84, 2), bold=True, size=9, fg=LABEL_TXT, fill=SUB_FILL,
           border=thin_bottom("9CA3AF"))
for c in [3,4,5,6]:
    ws_inp.cell(84, c).fill = solid(SUB_FILL)
    ws_inp.cell(84, c).border = thin_bottom("9CA3AF")
ws_inp.row_dimensions[84].height = 16

# col headers row 85
for c, txt in [(2,"Check"), (3,""), (4,"Value ($M)"), (5,""), (6,"")]:
    style_cell(ws_inp.cell(85, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL))
    ws_inp.cell(85, c).value = txt
ws_inp.row_dimensions[85].height = 18

check_rows = [
    (86, "Total Base EV from Drivers",   "=SUM(D63:D70)+SUM(D74:D75)+D79+D80"),
    (87, "Total Value Creation (INPUTS)", "=C30-C21"),
    (88, "Unattributed / Rounding",      "=D87-D86"),
]
for i, (r, lbl, frm) in enumerate(check_rows):
    alt = (i % 2 == 1)
    bg = TOTAL_BG if r == 88 else (ALT_BG if alt else BODY_BG)
    ws_inp.cell(r, 2).value = lbl
    style_cell(ws_inp.cell(r, 2), bold=(r==88), fg=LABEL_TXT, fill=bg,
               border=thin_bottom() if r < 88 else thick_bottom())
    ws_inp.cell(r, 4).value = frm
    style_cell(ws_inp.cell(r, 4), fg=FORM_TXT, fill=bg,
               border=thin_bottom() if r < 88 else thick_bottom(),
               ha="right", num_fmt='#,##0')
    for c in [3,5,6]:
        ws_inp.cell(r, c).fill = solid(bg)
        ws_inp.cell(r, c).border = thin_bottom() if r < 88 else thick_bottom()
    ws_inp.row_dimensions[r].height = 15

# ── column widths for new cols D,E,F ─────────────────────────────────────────
ws_inp.column_dimensions["D"].width = 18
ws_inp.column_dimensions["E"].width = 14
ws_inp.column_dimensions["F"].width = 18

# ── Reformat existing INPUTS rows 1-47 ───────────────────────────────────────
# Title row 2
style_cell(ws_inp.cell(2, 2), bold=True, size=14, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
ws_inp.cell(2, 3).fill = solid(NAV_FILL)
ws_inp.cell(2, 3).border = thick_bottom()
ws_inp.row_dimensions[2].height = 22

def format_inputs_section(ws, header_row, col_hdr_row, data_rows,
                          input_rows=None, formula_rows=None):
    """Apply consistent formatting to an INPUTS section."""
    # Sub-section header
    style_cell(ws.cell(header_row, 2), bold=True, size=11, fg=HDR_TXT,
               fill=NAV_FILL, border=thick_bottom())
    ws.cell(header_row, 3).fill = solid(NAV_FILL)
    ws.cell(header_row, 3).border = thick_bottom()
    ws.row_dimensions[header_row].height = 18
    # Column header row
    for c in [2, 3]:
        style_cell(ws.cell(col_hdr_row, c), bold=True, size=9, fg=HDR_TXT,
                   fill=NAV_FILL, border=thick_bottom(NAV_FILL))
    ws.row_dimensions[col_hdr_row].height = 18
    # Data rows
    for i, r in enumerate(data_rows):
        alt = (i % 2 == 1)
        bg = ALT_BG if alt else BODY_BG
        ws.cell(r, 2).fill = solid(bg)
        ws.cell(r, 2).border = thin_bottom()
        ws.cell(r, 2).font = mkfont(color=LABEL_TXT)
        is_input = (input_rows and r in input_rows)
        is_formula = (formula_rows and r in formula_rows)
        c3 = ws.cell(r, 3)
        if is_input:
            style_cell(c3, fg=INPUT_TXT, fill=bg, border=thin_bottom(), ha="right")
        elif is_formula or (not is_input):
            style_cell(c3, fg=FORM_TXT, fill=bg, border=thin_bottom(), ha="right")
        ws.row_dimensions[r].height = 15

# Spacer row 3
style_spacer_row(ws_inp, 3)
# Company & Deal section
format_inputs_section(ws_inp, 4, 5, list(range(6, 12)),
                      input_rows=set(range(6, 12)))
# Spacer 12
style_spacer_row(ws_inp, 12)
# Entry Assumptions
format_inputs_section(ws_inp, 13, 14, list(range(15, 22)),
                      input_rows={15,16,17,18,20},
                      formula_rows={19,21})
# Spacer 22
style_spacer_row(ws_inp, 22)
# Exit Assumptions
format_inputs_section(ws_inp, 23, 24, list(range(25, 31)),
                      input_rows={25,26,27,29},
                      formula_rows={28,30})
# Spacer 31
style_spacer_row(ws_inp, 31)
# Returns Summary
format_inputs_section(ws_inp, 32, 33, list(range(34, 37)),
                      formula_rows={34,35,36})
# Spacer 37
style_spacer_row(ws_inp, 37)
# NTB Registry
format_inputs_section(ws_inp, 38, 40, list(range(41, 48)),
                      input_rows=set(range(41, 48)))
ws_inp.cell(39, 2).font = mkfont(italic=True, color=SEC_TXT)
# Column widths
ws_inp.column_dimensions["A"].width = 1
ws_inp.column_dimensions["B"].width = 42
ws_inp.column_dimensions["C"].width = 16

# ══════════════════════════════════════════════════════════════════════════════
# DRIVER TREE TAB — replace hardcoded G/H values with INPUTS formulas
# ══════════════════════════════════════════════════════════════════════════════
ws_dt = wb["DRIVER TREE"]

# G/H replacements for L2 rows
dt_links = {
    8:  ("=INPUTS!D63", "=INPUTS!F63"),
    9:  ("=INPUTS!D64", "=INPUTS!F64"),
    10: ("=INPUTS!D65", "=INPUTS!F65"),
    12: ("=INPUTS!D66", "=INPUTS!F66"),
    13: ("=INPUTS!D67", "=INPUTS!F67"),
    14: ("=INPUTS!D68", "=INPUTS!F68"),
    16: ("=INPUTS!D74", "=INPUTS!F74"),
    17: ("=INPUTS!D75", "=INPUTS!F75"),
    19: (0,             "=INPUTS!F81"),
    20: (0,             "=INPUTS!F82"),
    22: ("=INPUTS!D69", "=INPUTS!F69"),
    23: ("=INPUTS!D70", "=INPUTS!F70"),
    25: ("=INPUTS!D79", "=INPUTS!F79"),
    26: ("=INPUTS!D80", "=INPUTS!F80"),
}
for r, (g_val, h_val) in dt_links.items():
    ws_dt.cell(r, 7).value = g_val
    ws_dt.cell(r, 8).value = h_val

# ── Reformat DRIVER TREE ─────────────────────────────────────────────────────
# Title row 2
style_cell(ws_dt.cell(2, 2), bold=True, size=14, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
for c in range(3, 11):
    ws_dt.cell(2, c).fill = solid(NAV_FILL)
    ws_dt.cell(2, c).border = thick_bottom()
ws_dt.row_dimensions[2].height = 22

# Subtitle row 4
ws_dt.cell(4, 2).font = mkfont(italic=True, color=SEC_TXT)
style_spacer_row(ws_dt, 3)
style_spacer_row(ws_dt, 5)

# Column header row 6
hdr_labels = {2:"Lvl", 3:"Value Driver", 4:"Type", 5:"NTB",
              6:"Evidence Tier", 7:"Base ($M)", 8:"Upside ($M)",
              9:"MOIC Delta", 10:"Notes"}
for c, lbl in hdr_labels.items():
    style_cell(ws_dt.cell(6, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL), ha="right" if c>=7 else "left")
    ws_dt.cell(6, c).value = lbl
ws_dt.row_dimensions[6].height = 18

# Evidence tier fill map
TIER_FILL = {"T1": T1_BG, "T2": T2_BG, "T3": T3_BG, "T4": T4_BG}
TIER_TXT  = {"T1": GREEN_TXT, "T2": GREEN_TXT, "T3": ORANGE_TXT, "T4": RED_TXT}

L1_rows = {7, 11, 15, 18, 21, 24}
L2_rows = {8,9,10, 12,13,14, 16,17, 19,20, 22,23, 25,26}
L28 = 28

for r in range(7, 27):
    is_l1 = r in L1_rows
    alt_idx = sum(1 for x in range(7, r) if x in L2_rows) % 2
    bg = SUB_FILL if is_l1 else (ALT_BG if alt_idx else BODY_BG)
    border = thin_bottom("9CA3AF") if is_l1 else thin_bottom()
    for c in range(2, 11):
        cell = ws_dt.cell(r, c)
        if c == 6:  # Evidence Tier — colour-coded
            tier = cell.value
            if tier and tier in TIER_FILL:
                cell.fill = solid(TIER_FILL[tier])
                cell.font = mkfont(bold=True, color=TIER_TXT[tier])
            else:
                cell.fill = solid(bg)
                cell.font = mkfont(color=BODY_TXT)
        else:
            cell.fill = solid(bg)
            is_link = (c in (7, 8) and r in dt_links)
            if is_link:
                cell.font = mkfont(color=LINK_TXT)
            elif c in (7, 8, 9):
                cell.font = mkfont(color=FORM_TXT)
            elif is_l1:
                cell.font = mkfont(bold=True, color=LABEL_TXT)
            else:
                cell.font = mkfont(color=BODY_TXT)
        if c in (7, 8, 9):
            cell.alignment = align(h="right")
            if c in (7, 8):
                cell.number_format = '#,##0'
        cell.border = border
    ws_dt.row_dimensions[r].height = 15

# Row 28 (totals)
for c in range(2, 11):
    ws_dt.cell(28, c).fill = solid(TOTAL_BG)
    ws_dt.cell(28, c).border = thick_bottom()
    if c in (7, 8, 9):
        ws_dt.cell(28, c).alignment = align(h="right")
ws_dt.row_dimensions[28].height = 15

# Column widths
ws_dt.column_dimensions["A"].width = 1
ws_dt.column_dimensions["B"].width = 5
ws_dt.column_dimensions["C"].width = 36
ws_dt.column_dimensions["D"].width = 10
ws_dt.column_dimensions["E"].width = 10
ws_dt.column_dimensions["F"].width = 13
ws_dt.column_dimensions["G"].width = 12
ws_dt.column_dimensions["H"].width = 12
ws_dt.column_dimensions["I"].width = 11
ws_dt.column_dimensions["J"].width = 40

# ══════════════════════════════════════════════════════════════════════════════
# NTB REGISTRY TAB — derive MOIC Impact from Driver Tree
# ══════════════════════════════════════════════════════════════════════════════
ws_ntb = wb["NTB REGISTRY"]

ntb_formulas = {
    8:  "=IFERROR('DRIVER TREE'!G8+'DRIVER TREE'!G9+'DRIVER TREE'!G10,\"-\")",
    9:  "=IFERROR('DRIVER TREE'!G12+'DRIVER TREE'!G13+'DRIVER TREE'!G14,\"-\")",
    10: "=IFERROR('DRIVER TREE'!H19+'DRIVER TREE'!H20,\"-\")",
    11: "=IFERROR('DRIVER TREE'!G22+'DRIVER TREE'!G23,\"-\")",
    12: "=IFERROR('DRIVER TREE'!G16+'DRIVER TREE'!G17,\"-\")",
    13: "=IFERROR(ABS('DRIVER TREE'!G26),\"-\")",
}
for r, frm in ntb_formulas.items():
    ws_ntb.cell(r, 6).value = frm
    ws_ntb.cell(r, 6).font = mkfont(color=LINK_TXT)
    ws_ntb.cell(r, 6).number_format = '#,##0'
    ws_ntb.cell(r, 6).alignment = align(h="right")

# ── Reformat NTB REGISTRY ─────────────────────────────────────────────────────
# Title row 2
style_cell(ws_ntb.cell(2, 2), bold=True, size=14, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
for c in range(3, 12):
    ws_ntb.cell(2, c).fill = solid(NAV_FILL)
    ws_ntb.cell(2, c).border = thick_bottom()
ws_ntb.row_dimensions[2].height = 22

style_spacer_row(ws_ntb, 3)
style_spacer_row(ws_ntb, 5)

# Check row 6
for c in range(2, 7):
    ws_ntb.cell(6, c).fill = solid(TOTAL_BG)
    ws_ntb.cell(6, c).border = thin_bottom()
ws_ntb.row_dimensions[6].height = 15
ws_ntb.cell(6, 2).font = mkfont(bold=True, color=LABEL_TXT)

# Header row 7
for c in range(2, 12):
    style_cell(ws_ntb.cell(7, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL))
ws_ntb.row_dimensions[7].height = 18

# Evidence status colour
EV_FILLS = {
    "Confirmed": (GREEN_BG, GREEN_TXT),
    "Partial":   (ORANGE_BG, ORANGE_TXT),
    "Gap":       (RED_BG, RED_TXT),
}
def ev_status_key(val):
    if not val:
        return None
    v = str(val).lower()
    if "gap" in v: return "Gap"
    if "partial" in v: return "Partial"
    if "confirmed" in v: return "Confirmed"
    return None

for i, r in enumerate(range(8, 14)):
    alt = (i % 2 == 1)
    bg = ALT_BG if alt else BODY_BG
    for c in range(2, 12):
        cell = ws_ntb.cell(r, c)
        if c == 4:  # Evidence Status
            key = ev_status_key(cell.value)
            if key:
                ebg, etxt = EV_FILLS[key]
                cell.fill = solid(ebg)
                cell.font = mkfont(bold=True, color=etxt)
            else:
                cell.fill = solid(bg)
                cell.font = mkfont(color=BODY_TXT)
        elif c == 6:  # MOIC Impact
            cell.fill = solid(bg)
            cell.font = mkfont(color=LINK_TXT)
            cell.alignment = align(h="right")
            cell.number_format = '#,##0'
        elif c == 7:  # MOIC (x)
            cell.fill = solid(bg)
            cell.font = mkfont(color=FORM_TXT)
            cell.alignment = align(h="right")
            cell.number_format = '0.00x'
        else:
            cell.fill = solid(bg)
            if c in (3,):
                cell.font = mkfont(color=FORM_TXT)
            else:
                cell.font = mkfont(color=BODY_TXT)
        cell.border = thin_bottom()
    ws_ntb.row_dimensions[r].height = 15

# Totals row 15
for c in range(2, 12):
    ws_ntb.cell(15, c).fill = solid(TOTAL_BG)
    ws_ntb.cell(15, c).border = thick_bottom()
ws_ntb.row_dimensions[15].height = 15
ws_ntb.cell(15, 6).alignment = align(h="right")
ws_ntb.cell(15, 6).number_format = '#,##0'
ws_ntb.cell(15, 7).alignment = align(h="right")
ws_ntb.cell(15, 7).number_format = '0.00x'

# Column widths
ws_ntb.column_dimensions["A"].width = 1
ws_ntb.column_dimensions["B"].width = 8
ws_ntb.column_dimensions["C"].width = 40
ws_ntb.column_dimensions["D"].width = 20
ws_ntb.column_dimensions["E"].width = 18
ws_ntb.column_dimensions["F"].width = 14
ws_ntb.column_dimensions["G"].width = 10
ws_ntb.column_dimensions["H"].width = 35
ws_ntb.column_dimensions["I"].width = 35
ws_ntb.column_dimensions["J"].width = 10
ws_ntb.column_dimensions["K"].width = 8

# ══════════════════════════════════════════════════════════════════════════════
# KPI TREE TAB — link Budget/Actual to FINANCIAL MODEL
# ══════════════════════════════════════════════════════════════════════════════
ws_kpi = wb["KPI TREE"]

# Budget (col G = col 7)
budget_links = {
    7:  ("='FINANCIAL MODEL'!G41", None),
    8:  ("='FINANCIAL MODEL'!G17", None),
    11: ("='FINANCIAL MODEL'!G44", None),
    12: ("='FINANCIAL MODEL'!G8",  None),
    15: ("='FINANCIAL MODEL'!G26", None),
    16: ("='FINANCIAL MODEL'!G9",  None),
    17: ("='FINANCIAL MODEL'!G10", None),
    18: ("=IFERROR('FINANCIAL MODEL'!G24/'FINANCIAL MODEL'!G17,\"-\")", None),
    20: ("='FINANCIAL MODEL'!G42", None),
    21: ("='FINANCIAL MODEL'!G43", None),
    24: ("='FINANCIAL MODEL'!G38", None),
    25: ("=IFERROR('FINANCIAL MODEL'!G38/'FINANCIAL MODEL'!G17,\"-\")", None),
    26: ("='FINANCIAL MODEL'!G14", None),
}
actual_links = {
    7:  "='FINANCIAL MODEL'!F41",
    8:  "='FINANCIAL MODEL'!F17",
    11: "='FINANCIAL MODEL'!F44",
    12: "=IFERROR('FINANCIAL MODEL'!F19/'FINANCIAL MODEL'!F17,\"-\")",
    15: "='FINANCIAL MODEL'!F26",
    20: "='FINANCIAL MODEL'!F42",
    21: "='FINANCIAL MODEL'!F43",
    24: "='FINANCIAL MODEL'!F38",
    25: "=IFERROR('FINANCIAL MODEL'!F38/'FINANCIAL MODEL'!F17,\"-\")",
}

for r, (frm, _) in budget_links.items():
    ws_kpi.cell(r, 7).value = frm
    ws_kpi.cell(r, 7).font = mkfont(color=LINK_TXT)
    ws_kpi.cell(r, 7).alignment = align(h="right")

for r, frm in actual_links.items():
    ws_kpi.cell(r, 8).value = frm
    ws_kpi.cell(r, 8).font = mkfont(color=LINK_TXT)
    ws_kpi.cell(r, 8).alignment = align(h="right")

# ── Status formula and colour-code ───────────────────────────────────────────
# Lower-is-better rows (OpEx, Capex)
lower_better = {18, 26}
for r in range(7, 27):
    cell_g = ws_kpi.cell(r, 7)
    cell_h = ws_kpi.cell(r, 8)
    if cell_g.value is None:
        continue
    k_cell = ws_kpi.cell(r, 11)  # Status col K
    j_cell = ws_kpi.cell(r, 10)  # Variance % col J
    if r in lower_better:
        status_frm = (f'=IFERROR(IF((H{r}-G{r})/G{r}<0.05,"On Track",'
                      f'IF((H{r}-G{r})/G{r}<0.15,"Watch","Behind Plan")),"-")')
    else:
        status_frm = (f'=IFERROR(IF((H{r}-G{r})/G{r}>-0.05,"On Track",'
                      f'IF((H{r}-G{r})/G{r}>-0.15,"Watch","Behind Plan")),"-")')
    # Variance formulas
    ws_kpi.cell(r, 9).value = f"=IFERROR(H{r}-G{r},\"-\")"
    ws_kpi.cell(r, 10).value = f"=IFERROR(I{r}/G{r},\"-\")"
    # Status
    k_cell.value = status_frm

# ── Reformat KPI TREE ────────────────────────────────────────────────────────
# Title row 2
style_cell(ws_kpi.cell(2, 2), bold=True, size=14, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
for c in range(3, 13):
    ws_kpi.cell(2, c).fill = solid(NAV_FILL)
    ws_kpi.cell(2, c).border = thick_bottom()
ws_kpi.row_dimensions[2].height = 22

style_spacer_row(ws_kpi, 3)
style_spacer_row(ws_kpi, 5)

# Column header row 6
for c in range(2, 13):
    style_cell(ws_kpi.cell(6, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL))
ws_kpi.row_dimensions[6].height = 18

L1_kpi = {7, 8, 12, 15, 19, 24}
NS_kpi  = {7}

for i, r in enumerate(range(7, 27)):
    is_l1 = r in L1_kpi
    is_ns  = r in NS_kpi
    bg = SUB_FILL if is_l1 else (ALT_BG if i % 2 else BODY_BG)
    border = thin_bottom("9CA3AF") if is_l1 else thin_bottom()
    for c in range(2, 13):
        cell = ws_kpi.cell(r, c)
        if c == 11:  # Status — apply conditional colour
            sv = str(cell.value) if cell.value else ""
            if "On Track" in sv or "on track" in sv.lower():
                cell.fill = solid(GREEN_BG)
                cell.font = mkfont(bold=True, color=GREEN_TXT)
            elif "Watch" in sv or "watch" in sv.lower():
                cell.fill = solid(ORANGE_BG)
                cell.font = mkfont(bold=True, color=ORANGE_TXT)
            elif "Behind" in sv or "behind" in sv.lower():
                cell.fill = solid(RED_BG)
                cell.font = mkfont(bold=True, color=RED_TXT)
            else:
                cell.fill = solid(bg)
                cell.font = mkfont(color=BODY_TXT)
        elif c in (7, 8):
            # Preserve font color set above (LINK_TXT); just update fill/border
            cell.fill = solid(bg)
        elif c in (9, 10):
            cell.fill = solid(bg)
            cell.font = mkfont(color=FORM_TXT)
            cell.alignment = align(h="right")
        elif c == 12:
            cell.font = mkfont(italic=True, color=SEC_TXT)
            cell.fill = solid(bg)
        else:
            cell.fill = solid(bg)
            if is_l1:
                cell.font = mkfont(bold=True, color=LABEL_TXT)
            else:
                cell.font = mkfont(color=BODY_TXT)
        cell.border = border
    ws_kpi.row_dimensions[r].height = 15

# Column widths
ws_kpi.column_dimensions["A"].width = 1
ws_kpi.column_dimensions["B"].width = 5
ws_kpi.column_dimensions["C"].width = 30
ws_kpi.column_dimensions["D"].width = 12
ws_kpi.column_dimensions["E"].width = 22
ws_kpi.column_dimensions["F"].width = 10
ws_kpi.column_dimensions["G"].width = 14
ws_kpi.column_dimensions["H"].width = 14
ws_kpi.column_dimensions["I"].width = 12
ws_kpi.column_dimensions["J"].width = 10
ws_kpi.column_dimensions["K"].width = 12
ws_kpi.column_dimensions["L"].width = 35

# ══════════════════════════════════════════════════════════════════════════════
# COVER TAB — reformat
# ══════════════════════════════════════════════════════════════════════════════
ws_cov = wb["COVER"]
style_cell(ws_cov.cell(2, 2), bold=True, size=14, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
for c in range(3, 5):
    ws_cov.cell(2, c).fill = solid(NAV_FILL)
    ws_cov.cell(2, c).border = thick_bottom()
ws_cov.row_dimensions[2].height = 22

ws_cov.cell(3, 2).font = mkfont(italic=True, color=SEC_TXT)
style_spacer_row(ws_cov, 4)

# Nav table
style_cell(ws_cov.cell(5, 2), bold=True, size=11, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
ws_cov.cell(5, 3).fill = solid(NAV_FILL)
ws_cov.cell(5, 4).fill = solid(NAV_FILL)
ws_cov.row_dimensions[5].height = 18

for c in range(2, 5):
    style_cell(ws_cov.cell(6, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL))
ws_cov.row_dimensions[6].height = 18

for i, r in enumerate(range(7, 14)):
    alt = (i % 2 == 1)
    bg = ALT_BG if alt else BODY_BG
    for c in range(2, 5):
        ws_cov.cell(r, c).fill = solid(bg)
        ws_cov.cell(r, c).border = thin_bottom()
        ws_cov.cell(r, c).font = mkfont(color=BODY_TXT if c>2 else LABEL_TXT,
                                      bold=(c==2))
    ws_cov.row_dimensions[r].height = 15

style_spacer_row(ws_cov, 14)

# Investment Summary
style_cell(ws_cov.cell(15, 2), bold=True, size=11, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
ws_cov.cell(15, 3).fill = solid(NAV_FILL)
ws_cov.cell(15, 4).fill = solid(NAV_FILL)
ws_cov.row_dimensions[15].height = 18

for c in range(2, 5):
    style_cell(ws_cov.cell(16, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL))
ws_cov.row_dimensions[16].height = 18

for i, r in enumerate(range(17, 27)):
    if ws_cov.cell(r, 2).value is None:
        continue
    alt = (i % 2 == 1)
    bg = ALT_BG if alt else BODY_BG
    for c in range(2, 5):
        ws_cov.cell(r, c).fill = solid(bg)
        ws_cov.cell(r, c).border = thin_bottom()
        ws_cov.cell(r, c).font = mkfont(color=BODY_TXT if c>2 else LABEL_TXT,
                                      bold=(c==2))
    ws_cov.row_dimensions[r].height = 15

style_spacer_row(ws_cov, 27)

# Colour coding table
style_cell(ws_cov.cell(28, 2), bold=True, size=11, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
ws_cov.cell(28, 3).fill = solid(NAV_FILL)
ws_cov.cell(28, 4).fill = solid(NAV_FILL)
ws_cov.row_dimensions[28].height = 18

for c in range(2, 5):
    style_cell(ws_cov.cell(29, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL))
ws_cov.row_dimensions[29].height = 18

for i, r in enumerate(range(30, 36)):
    alt = (i % 2 == 1)
    bg = ALT_BG if alt else BODY_BG
    for c in range(2, 5):
        ws_cov.cell(r, c).fill = solid(bg)
        ws_cov.cell(r, c).border = thin_bottom()
        ws_cov.cell(r, c).font = mkfont(color=BODY_TXT)
    ws_cov.row_dimensions[r].height = 15

ws_cov.column_dimensions["A"].width = 1
ws_cov.column_dimensions["B"].width = 28
ws_cov.column_dimensions["C"].width = 28
ws_cov.column_dimensions["D"].width = 35

# ══════════════════════════════════════════════════════════════════════════════
# FINANCIALS TAB — reformat
# ══════════════════════════════════════════════════════════════════════════════
ws_fin = wb["FINANCIALS"]
style_cell(ws_fin.cell(2, 2), bold=True, size=14, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
for c in range(3, 8):
    ws_fin.cell(2, c).fill = solid(NAV_FILL)
    ws_fin.cell(2, c).border = thick_bottom()
ws_fin.row_dimensions[2].height = 22

style_spacer_row(ws_fin, 3)
ws_fin.cell(4, 2).font = mkfont(italic=True, color=SEC_TXT)
style_spacer_row(ws_fin, 5)

# Col header row 6
for c in range(2, 8):
    style_cell(ws_fin.cell(6, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL), ha="right" if c>2 else "left")
ws_fin.row_dimensions[6].height = 18

section_rows_fin = {7, 26, 32}
for i, r in enumerate(range(7, 36)):
    is_sec = r in section_rows_fin
    alt = (i % 2 == 1)
    bg = SUB_FILL if is_sec else (ALT_BG if alt else BODY_BG)
    border = thin_bottom("9CA3AF") if is_sec else thin_bottom()
    for c in range(2, 8):
        cell = ws_fin.cell(r, c)
        cell.fill = solid(bg)
        cell.border = border
        if is_sec:
            cell.font = mkfont(bold=True, color=LABEL_TXT)
        elif c > 2 and c < 7:
            cell.font = mkfont(color=BODY_TXT)
            cell.alignment = align(h="right")
        elif c == 7:
            cell.font = mkfont(italic=True, color=SEC_TXT)
        else:
            cell.font = mkfont(color=LABEL_TXT)
    ws_fin.row_dimensions[r].height = 15

ws_fin.column_dimensions["A"].width = 1
ws_fin.column_dimensions["B"].width = 32
for col_l, w in [("C",12),("D",12),("E",12),("F",12),("G",35)]:
    ws_fin.column_dimensions[col_l].width = w

# ══════════════════════════════════════════════════════════════════════════════
# FINANCIAL MODEL TAB — reformat only (preserve all formulas)
# ══════════════════════════════════════════════════════════════════════════════
ws_fm = wb["FINANCIAL MODEL"]
style_cell(ws_fm.cell(2, 2), bold=True, size=14, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
for c in range(3, 13):
    ws_fm.cell(2, c).fill = solid(NAV_FILL)
    ws_fm.cell(2, c).border = thick_bottom()
ws_fm.row_dimensions[2].height = 22

style_spacer_row(ws_fm, 3)

# Header rows 4 (column headers)
for c in range(2, 13):
    style_cell(ws_fm.cell(4, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL), ha="right" if c>2 else "left")
ws_fm.row_dimensions[4].height = 18

# Assumption block 5-6
for r in (5, 6):
    for c in range(2, 13):
        ws_fm.cell(r, c).fill = solid(BODY_BG)
        ws_fm.cell(r, c).font = mkfont(italic=True, color=SEC_TXT)
    ws_fm.row_dimensions[r].height = 14

# Assumptions rows 7-14 (yellow inputs in original → INPUT_TXT)
for r in range(7, 15):
    for c in range(2, 13):
        cell = ws_fm.cell(r, c)
        old_fill = cell.fill
        is_proj = (c >= 7 and c <= 11)
        cell.fill = solid(ALT_BG if r % 2 == 0 else BODY_BG)
        cell.border = thin_bottom()
        if c == 2:
            cell.font = mkfont(color=LABEL_TXT)
        elif c == 12:
            cell.font = mkfont(italic=True, color=SEC_TXT)
        elif is_proj:
            # Projection assumption — input
            if isinstance(cell.value, (int, float)):
                cell.font = mkfont(color=INPUT_TXT)
            else:
                cell.font = mkfont(color=FORM_TXT)
            cell.alignment = align(h="right")
        else:
            cell.font = mkfont(color=BODY_TXT)
            cell.alignment = align(h="right")
    ws_fm.row_dimensions[r].height = 15

# IS section header row 16
style_cell(ws_fm.cell(16, 2), bold=True, size=9, fg=LABEL_TXT, fill=SUB_FILL,
           border=thin_bottom("9CA3AF"))
for c in range(3, 13):
    ws_fm.cell(16, c).fill = solid(SUB_FILL)
    ws_fm.cell(16, c).border = thin_bottom("9CA3AF")
ws_fm.row_dimensions[16].height = 16

# IS rows 17-32
section_rows_fm = {16, 34, 40}
for i, r in enumerate(range(17, 44)):
    is_sec = r in section_rows_fm
    is_sub = r in {18, 20}  # growth/margin sub-rows
    alt = (i % 2 == 1)
    bg = SUB_FILL if is_sec else (ALT_BG if alt else BODY_BG)
    border = thin_bottom("9CA3AF") if is_sec else thin_bottom()
    for c in range(2, 13):
        cell = ws_fm.cell(r, c)
        cell.fill = solid(bg)
        cell.border = border
        if c == 2:
            cell.font = mkfont(bold=is_sec, color=LABEL_TXT if not is_sub else SEC_TXT)
        elif c == 12:
            cell.font = mkfont(italic=True, color=SEC_TXT)
        elif c >= 7 and c <= 11:
            cell.font = mkfont(color=FORM_TXT)
            cell.alignment = align(h="right")
        elif c in (3,4,5,6):
            cell.font = mkfont(color=BODY_TXT)
            cell.alignment = align(h="right")
    ws_fm.row_dimensions[r].height = 15

# FCF section header row 34
style_cell(ws_fm.cell(34, 2), bold=True, size=9, fg=LABEL_TXT, fill=SUB_FILL,
           border=thin_bottom("9CA3AF"))
for c in range(3, 13):
    ws_fm.cell(34, c).fill = solid(SUB_FILL)
    ws_fm.cell(34, c).border = thin_bottom("9CA3AF")
ws_fm.row_dimensions[34].height = 16

# KPI section header row 40
style_cell(ws_fm.cell(40, 2), bold=True, size=9, fg=LABEL_TXT, fill=SUB_FILL,
           border=thin_bottom("9CA3AF"))
for c in range(3, 13):
    ws_fm.cell(40, c).fill = solid(SUB_FILL)
    ws_fm.cell(40, c).border = thin_bottom("9CA3AF")
ws_fm.row_dimensions[40].height = 16

ws_fm.column_dimensions["A"].width = 1
ws_fm.column_dimensions["B"].width = 30
for col_l, w in [("C",10),("D",10),("E",10),("F",10),
                 ("G",10),("H",10),("I",10),("J",10),("K",10),("L",35)]:
    ws_fm.column_dimensions[col_l].width = w

# ══════════════════════════════════════════════════════════════════════════════
# MOIC BRIDGE TAB — reformat
# ══════════════════════════════════════════════════════════════════════════════
ws_mb = wb["MOIC BRIDGE"]
style_cell(ws_mb.cell(2, 2), bold=True, size=14, fg=HDR_TXT, fill=NAV_FILL,
           border=thick_bottom())
for c in range(3, 6):
    ws_mb.cell(2, c).fill = solid(NAV_FILL)
    ws_mb.cell(2, c).border = thick_bottom()
ws_mb.row_dimensions[2].height = 22

style_spacer_row(ws_mb, 3)
ws_mb.cell(4, 2).font = mkfont(italic=True, color=SEC_TXT)
style_spacer_row(ws_mb, 5)

for c in range(2, 6):
    style_cell(ws_mb.cell(6, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL), ha="right" if c>2 else "left")
ws_mb.row_dimensions[6].height = 18

# Entry row 7 — navy
for c in range(2, 6):
    style_cell(ws_mb.cell(7, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL))
    if c > 2:
        ws_mb.cell(7, c).alignment = align(h="right")
ws_mb.row_dimensions[7].height = 15

# NTB rows 8-13
for i, r in enumerate(range(8, 14)):
    alt = (i % 2 == 1)
    bg = ALT_BG if alt else BODY_BG
    for c in range(2, 6):
        ws_mb.cell(r, c).fill = solid(bg)
        ws_mb.cell(r, c).border = thin_bottom()
        ws_mb.cell(r, c).font = mkfont(color=LINK_TXT if c in (3,4) else BODY_TXT)
        if c >= 3:
            ws_mb.cell(r, c).alignment = align(h="right")
    ws_mb.row_dimensions[r].height = 15

# Row 14 (Multiple re-rating)
for c in range(2, 6):
    ws_mb.cell(14, c).fill = solid(TOTAL_BG)
    ws_mb.cell(14, c).border = thin_bottom()
    if c >= 3:
        ws_mb.cell(14, c).alignment = align(h="right")
ws_mb.row_dimensions[14].height = 15

style_spacer_row(ws_mb, 15)

# Exit row 16 — navy
for c in range(2, 6):
    style_cell(ws_mb.cell(16, c), bold=True, size=9, fg=HDR_TXT, fill=NAV_FILL,
               border=thick_bottom(NAV_FILL))
    if c > 2:
        ws_mb.cell(16, c).alignment = align(h="right")
ws_mb.row_dimensions[16].height = 15

ws_mb.column_dimensions["A"].width = 1
ws_mb.column_dimensions["B"].width = 50
ws_mb.column_dimensions["C"].width = 18
ws_mb.column_dimensions["D"].width = 16
ws_mb.column_dimensions["E"].width = 16

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
wb.save(TMP)
print(f"Saved to {TMP}")
shutil.copy2(TMP, DST)
print(f"Copied to {DST}")
