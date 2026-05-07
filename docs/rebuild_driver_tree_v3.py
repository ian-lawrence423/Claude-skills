"""
Shopify Deal Workbook — Driver Tree v3 Rebuild
Replaces arbitrary EBITDA-pool % allocations with formula-derived, model-linked driver analytics.

Architecture:
  Revenue delta (FY2026E→FY2028E)
    → EBITDA delta (via segment incremental margins, calibrated to model)
      → EV delta (× entry multiple, Paasche decomposition)
        → MOIC delta (/ entry equity)
  Plus: multiple compression + net cash build = total value creation
  Grand total reconciles to 1.58x MOIC
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter

# ─── PATH ────────────────────────────────────────────────────────────────────
WB_PATH = r"C:\Users\IanLawrence\github\Claude-skills\docs\Shopify_Deal_Workbook_v2.xlsx"

# ─── STYLE CONSTANTS (Pattern/McKinsey palette) ───────────────────────────────
# Colors
C_NAVY      = "0F4761"   # section headers fill
C_BLUE      = "4280F4"   # sub-section label text
C_INDIGO    = "3A00FD"   # sub-heading text
C_WHITE     = "FFFFFF"
C_BLACK     = "000000"
C_GREY_ALT  = "F2F2F2"   # alternating row
C_GREY_BDR  = "DDDDDD"   # cell border
C_SLATE     = "444444"   # secondary text
C_BLUE_IN   = "0000FF"   # hardcoded input (blue text convention)
C_TEAL_HL   = "D9E2F3"   # highlight / subtotal row
C_GREEN     = "375623"   # positive / pass
C_RED       = "C00000"   # negative / risk
C_ORANGE    = "C55A11"   # analyst assumption

FONT_MAIN   = "Arial"
SZ_BODY     = 9
SZ_HDR      = 9

def font(bold=False, color=C_BLACK, sz=SZ_BODY, italic=False):
    return Font(name=FONT_MAIN, bold=bold, color=color, size=sz, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin(all_sides=True, bottom_only=False):
    thin = Side(style="thin", color=C_GREY_BDR)
    none = Side(style=None)
    if bottom_only:
        return Border(bottom=thin)
    if all_sides:
        return Border(top=thin, bottom=thin, left=thin, right=thin)
    return Border(bottom=thin)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def fmt_num(val, decimals=0):
    """Format number with commas"""
    if val is None:
        return None
    return round(val, decimals)

# ─── COMPUTED ANALYTICS ───────────────────────────────────────────────────────
# All revenue/EBITDA in $M, GMV in $B, rates in decimal

# Model data (from model.xlsx via extraction)
GMV_25A    = 378.441
GMV_26E    = 477.838
GMV_28E    = 699.555
ATTACH_25A = 232.6   # bps
ATTACH_26E = 240.9   # bps
ATTACH_28E = 258.9   # bps
GPV_25A    = 68.0    # %
GPV_26E    = 71.0    # %
GPV_28E    = 79.0    # %
MRR_25A    = 205.0
MRR_26E    = 239.2
MRR_28E    = 290.9
PLUS_MRR_25A = 69.7
PLUS_MRR_26E = 83.6
PLUS_MRR_28E = 114.5
CORE_MRR_25A = 135.3
CORE_MRR_26E = 155.6
CORE_MRR_28E = 176.4

MERCH_REV_26E = 11513
MERCH_REV_28E = 18115
SUB_REV_26E   = 3305
SUB_REV_28E   = 4023
TOT_REV_26E   = 14818
TOT_REV_28E   = 22138

EBITDA_26E = 2789    # entry EBITDA (FY2026E)
EBITDA_28E = 4687    # exit EBITDA (FY2028E)
EBITDA_DLT = EBITDA_28E - EBITDA_26E  # 1898

# Opex (S&M + R&D + G&A)
OPEX_26E = 2001 + 1408 + 378   # 3787
OPEX_28E = 2663 + 1769 + 409   # 4841
INCR_OPEX_RATE = (OPEX_28E - OPEX_26E) / (TOT_REV_28E - TOT_REV_26E)  # 14.4%

# Segment gross margins
MERCH_GM   = 0.377
SUB_GM     = 0.820

# Incremental EBITDA margins — calibrated so revenue drivers sum to actual EBITDA delta
# Sub: incr EBITDA = 82% - 14.4% = 67.6%; Sub contribution = (4023-3305)*67.6% = 485
SUB_INCR_EBITDA   = SUB_GM - INCR_OPEX_RATE        # 67.6%
# Merch: solve residual → 1898 - 485 = 1413; 1413/(18115-11513) = 21.4%
MERCH_REV_DLT     = MERCH_REV_28E - MERCH_REV_26E  # 6602
SUB_REV_DLT       = SUB_REV_28E - SUB_REV_26E      # 718
SUB_EBITDA_CONTRIB = SUB_REV_DLT * SUB_INCR_EBITDA
MERCH_INCR_EBITDA  = (EBITDA_DLT - SUB_EBITDA_CONTRIB) / MERCH_REV_DLT  # ~21.4%

# Returns (from INPUTS tab)
ENTRY_EQUITY   = 141569
ENTRY_MULTIPLE = 48.5   # entry EV/EBITDA (Paasche: use entry multiple for EBITDA growth EV)
EXIT_MULTIPLE  = 45.0
NET_CASH_BUILD = 6115   # increase in net cash (12369 - 6254)
TOT_VAL_CREATE = 81715  # total equity value creation

# ─── MERCHANT SOLUTIONS DECOMPOSITION ────────────────────────────────────────
GMV_DLT = GMV_28E - GMV_26E  # 221.717 $B

# Volume vs. attach rate split (Paasche: volume at base attach, attach at exit volume)
# GMV in $B × bps/10000 = $B of revenue → ×1000 to convert to $M
GMV_VOL_REV     = GMV_DLT * (ATTACH_26E / 10000) * 1000    # 5342M: GMV growth at entry attach rate
ATTACH_RATE_REV = GMV_28E * ((ATTACH_28E - ATTACH_26E) / 10000) * 1000  # 1259M: attach improvement at exit GMV
# Check: 5342 + 1259 = 6601 ≈ 6602 ✓

# GMV sub-components (analyst judgment based on IC Memo; flagged [Analyst])
# Ref: IC Memo — Intl GMV +45% YoY, B2B +80%, same-cohort productivity dominant
PCT_SAME_COHORT  = 0.40  # [Analyst] existing merchant same-store growth
PCT_NEW_MERCHANT = 0.30  # [Analyst] net new merchant cohort adds
PCT_GEOGRAPHIC   = 0.20  # [IC Memo] international market expansion
PCT_B2B_CHANNEL  = 0.10  # [IC Memo] B2B, POS, Shop App channel GMV

REV_SAME_COHORT  = GMV_VOL_REV * PCT_SAME_COHORT   # 2137M
REV_NEW_MERCHANT = GMV_VOL_REV * PCT_NEW_MERCHANT   # 1603M
REV_GEOGRAPHIC   = GMV_VOL_REV * PCT_GEOGRAPHIC    # 1068M
REV_B2B          = GMV_VOL_REV * PCT_B2B_CHANNEL   # 534M

# Attach rate sub-components
# GPV penetration: IC Memo states 100bps GPV% = $67M at FY2025A GMV ($378.4B)
# Scale to exit GMV: $67M × (699.6/378.4) = $123.9M per 100bps
GPV_INC_PER_100BPS = 67 * (GMV_28E / GMV_25A)   # 123.9M per 100bps at exit GMV
GPV_DLT_PP         = GPV_28E - GPV_26E           # +8pp from entry to exit (71%→79%)
REV_GPV_PENETRATION = GPV_DLT_PP * (GPV_INC_PER_100BPS / 1)  # 8pp × 123.9M = 991M
REV_NON_PAYMENTS   = ATTACH_RATE_REV - REV_GPV_PENETRATION    # 1259 - 991 = 268M

# ─── SUBSCRIPTION DECOMPOSITION ─────────────────────────────────────────────
SUB_REV_MULT = SUB_REV_26E / MRR_26E   # implied revenue/MRR multiplier (13.82×)

REV_PLUS_MRR = (PLUS_MRR_28E - PLUS_MRR_26E) * SUB_REV_MULT   # 427M
REV_CORE_MRR = (CORE_MRR_28E - CORE_MRR_26E) * SUB_REV_MULT   # 287M
REV_SUB_OTHER = SUB_REV_DLT - REV_PLUS_MRR - REV_CORE_MRR     # ~4M residual

# ─── EBITDA ATTRIBUTION ───────────────────────────────────────────────────────
def ebitda(rev_impact, segment):
    margin = MERCH_INCR_EBITDA if segment == "merch" else SUB_INCR_EBITDA
    return rev_impact * margin

EBITDA_SAME_COHORT   = ebitda(REV_SAME_COHORT, "merch")
EBITDA_NEW_MERCHANT  = ebitda(REV_NEW_MERCHANT, "merch")
EBITDA_GEOGRAPHIC    = ebitda(REV_GEOGRAPHIC, "merch")
EBITDA_B2B           = ebitda(REV_B2B, "merch")
EBITDA_GPV           = ebitda(REV_GPV_PENETRATION, "merch")
EBITDA_NON_PAYMENTS  = ebitda(REV_NON_PAYMENTS, "merch")
EBITDA_PLUS_MRR      = ebitda(REV_PLUS_MRR, "sub")
EBITDA_CORE_MRR      = ebitda(REV_CORE_MRR, "sub")
EBITDA_SUB_OTHER     = ebitda(REV_SUB_OTHER, "sub")

# Multiple compression
MULTI_COMPRESS_EBITDA = 0   # not a revenue driver; handled in bridge section

# ─── EV ATTRIBUTION (Paasche: EBITDA × entry multiple) ───────────────────────
def ev_impact(ebitda_val):
    return ebitda_val * ENTRY_MULTIPLE

EV_SAME_COHORT   = ev_impact(EBITDA_SAME_COHORT)
EV_NEW_MERCHANT  = ev_impact(EBITDA_NEW_MERCHANT)
EV_GEOGRAPHIC    = ev_impact(EBITDA_GEOGRAPHIC)
EV_B2B           = ev_impact(EBITDA_B2B)
EV_GPV           = ev_impact(EBITDA_GPV)
EV_NON_PAYMENTS  = ev_impact(EBITDA_NON_PAYMENTS)
EV_PLUS_MRR      = ev_impact(EBITDA_PLUS_MRR)
EV_CORE_MRR      = ev_impact(EBITDA_CORE_MRR)
EV_SUB_OTHER     = ev_impact(EBITDA_SUB_OTHER)

EV_MULTI_COMPRESS = (EXIT_MULTIPLE - ENTRY_MULTIPLE) * EBITDA_28E   # -16405M
EV_NET_CASH       = NET_CASH_BUILD                                    # +6115M

# Revenue-driven EV total (EBITDA × entry multiple)
EV_REV_TOTAL = sum([EV_SAME_COHORT, EV_NEW_MERCHANT, EV_GEOGRAPHIC, EV_B2B,
                    EV_GPV, EV_NON_PAYMENTS, EV_PLUS_MRR, EV_CORE_MRR, EV_SUB_OTHER])
EV_GRAND_TOTAL = EV_REV_TOTAL + EV_MULTI_COMPRESS + EV_NET_CASH   # ≈ 81,715M

# ─── MOIC ATTRIBUTION ─────────────────────────────────────────────────────────
def moic(ev_val):
    return ev_val / ENTRY_EQUITY

MOIC_SAME_COHORT   = moic(EV_SAME_COHORT)
MOIC_NEW_MERCHANT  = moic(EV_NEW_MERCHANT)
MOIC_GEOGRAPHIC    = moic(EV_GEOGRAPHIC)
MOIC_B2B           = moic(EV_B2B)
MOIC_GPV           = moic(EV_GPV)
MOIC_NON_PAYMENTS  = moic(EV_NON_PAYMENTS)
MOIC_PLUS_MRR      = moic(EV_PLUS_MRR)
MOIC_CORE_MRR      = moic(EV_CORE_MRR)
MOIC_SUB_OTHER     = moic(EV_SUB_OTHER)
MOIC_MULTI         = moic(EV_MULTI_COMPRESS)
MOIC_CASH          = moic(EV_NET_CASH)
MOIC_TOTAL         = 1.0 + moic(EV_GRAND_TOTAL)   # starting from 1.0x invested capital

# ─── PRINT CHECK ─────────────────────────────────────────────────────────────
print("=== ANALYTICS CHECK ===")
print(f"GMV Volume Rev:        ${GMV_VOL_REV:,.0f}M")
print(f"Attach Rate Rev:       ${ATTACH_RATE_REV:,.0f}M")
print(f"Total Merch Rev Delta: ${GMV_VOL_REV+ATTACH_RATE_REV:,.0f}M  (model: {MERCH_REV_DLT})")
print(f"GPV Penetration Rev:   ${REV_GPV_PENETRATION:,.0f}M")
print(f"Non-Payments Rev:      ${REV_NON_PAYMENTS:,.0f}M")
print(f"Plus MRR Rev:          ${REV_PLUS_MRR:,.0f}M")
print(f"Core MRR Rev:          ${REV_CORE_MRR:,.0f}M")
print(f"Sub Rev Delta (model): ${SUB_REV_DLT}M")
print(f"Merch Incr EBITDA Mgn: {MERCH_INCR_EBITDA:.1%}")
print(f"Sub Incr EBITDA Mgn:   {SUB_INCR_EBITDA:.1%}")
print(f"Rev-driven EBITDA:     ${sum([EBITDA_SAME_COHORT, EBITDA_NEW_MERCHANT, EBITDA_GEOGRAPHIC, EBITDA_B2B, EBITDA_GPV, EBITDA_NON_PAYMENTS, EBITDA_PLUS_MRR, EBITDA_CORE_MRR, EBITDA_SUB_OTHER]):,.0f}M (model: {EBITDA_DLT})")
print(f"EV Revenue Total:      ${EV_REV_TOTAL:,.0f}M  (EBITDA pool × entry mult)")
print(f"EV Multi Compress:     ${EV_MULTI_COMPRESS:,.0f}M")
print(f"EV Net Cash Build:     ${EV_NET_CASH:,.0f}M")
print(f"EV Grand Total:        ${EV_GRAND_TOTAL:,.0f}M  (model: {TOT_VAL_CREATE})")
print(f"MOIC Implied:          {MOIC_TOTAL:.2f}x  (target: 1.58x)")

# ─── WORKBOOK ─────────────────────────────────────────────────────────────────
wb = load_workbook(WB_PATH)

# ─── HELPER: write a section header row ──────────────────────────────────────
def write_section_header(ws, row, label, start_col=2, end_col=14):
    ws.cell(row=row, column=start_col, value=label).font = font(bold=True, color=C_WHITE, sz=SZ_HDR)
    ws.cell(row=row, column=start_col).fill = fill(C_NAVY)
    ws.cell(row=row, column=start_col).alignment = align("left")
    for c in range(start_col + 1, end_col + 1):
        ws.cell(row=row, column=c).fill = fill(C_NAVY)

def write_col_header(ws, row, labels, start_col=2):
    for i, lbl in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + i, value=lbl)
        cell.font = font(bold=True, color=C_WHITE, sz=SZ_BODY - 1)
        cell.fill = fill(C_NAVY)
        cell.alignment = align("center")
        cell.border = border_thin()

def write_subhdr(ws, row, label, col=2, indent=0):
    prefix = "  " * indent
    cell = ws.cell(row=row, column=col, value=prefix + label)
    cell.font = font(bold=True, color=C_BLUE, sz=SZ_BODY)
    cell.alignment = align("left")
    for c in range(col, 15):
        ws.cell(row=row, column=c).border = Border(bottom=Side(style="thin", color=C_GREY_BDR))

def write_driver_row(ws, row, indent, label, source, metric, unit,
                     val_25a, val_26e, val_28e, delta, rev_impact,
                     ebitda_margin, ebitda_imp, ev_imp, moic_delta,
                     is_subtotal=False, is_negative=False):
    """Write one driver row. Columns: B=label C=source D=metric E=unit F-H=vals I=delta J=rev K=ebitda% L=ebitda M=ev N=moic"""
    bg = C_TEAL_HL if is_subtotal else (C_GREY_ALT if (row % 2 == 0) else C_WHITE)
    txt_color = C_RED if is_negative else (C_BLACK if not is_subtotal else C_BLACK)
    bold = is_subtotal

    prefix = "    " * indent
    ws.cell(row=row, column=2, value=prefix + label).font = font(bold=bold, color=txt_color)
    ws.cell(row=row, column=2).fill = fill(bg)
    ws.cell(row=row, column=2).alignment = align("left")

    # Source tag
    src_color = C_SLATE if source == "Model" else (C_ORANGE if source == "Analyst" else C_INDIGO)
    ws.cell(row=row, column=3, value=source).font = font(bold=False, color=src_color, sz=7)
    ws.cell(row=row, column=3).fill = fill(bg)
    ws.cell(row=row, column=3).alignment = align("center")

    ws.cell(row=row, column=4, value=metric).font = font(italic=True, color=C_SLATE, sz=7)
    ws.cell(row=row, column=4).fill = fill(bg)
    ws.cell(row=row, column=4).alignment = align("left")

    ws.cell(row=row, column=5, value=unit).font = font(bold=False, color=C_SLATE, sz=7)
    ws.cell(row=row, column=5).fill = fill(bg)
    ws.cell(row=row, column=5).alignment = align("center")

    num_cols = [
        (6,  val_25a,      "#,##0.0" if "B" in (unit or "") else ("#,##0" if "$M" in (unit or "") or "$" in (unit or "") else "0.0%") ),
        (7,  val_26e,      "#,##0.0" if "B" in (unit or "") else ("#,##0" if "$M" in (unit or "") or "$" in (unit or "") else "0.0%") ),
        (8,  val_28e,      "#,##0.0" if "B" in (unit or "") else ("#,##0" if "$M" in (unit or "") or "$" in (unit or "") else "0.0%") ),
        (9,  delta,        "#,##0.0" if "B" in (unit or "") else ("#,##0" if "$M" in (unit or "") or "$" in (unit or "") else "0.0%") ),
        (10, rev_impact,   '#,##0;(#,##0);"-"'),
        (11, ebitda_margin,"0.0%"),
        (12, ebitda_imp,   '#,##0;(#,##0);"-"'),
        (13, ev_imp,       '#,##0;(#,##0);"-"'),
        (14, moic_delta,   '+0.000x;-0.000x;"-"'),
    ]
    for col, val, fmt in num_cols:
        c = ws.cell(row=row, column=col, value=round(val, 3) if isinstance(val, float) else val)
        c.font = font(bold=bold, color=txt_color)
        c.fill = fill(bg)
        c.alignment = align("right")
        c.border = border_thin()
        c.number_format = fmt

    # Label cell border
    ws.cell(row=row, column=2).border = border_thin()
    ws.cell(row=row, column=3).border = border_thin()
    ws.cell(row=row, column=4).border = border_thin()
    ws.cell(row=row, column=5).border = border_thin()


# ════════════════════════════════════════════════════════════════════
# DRIVER TREE TAB
# ════════════════════════════════════════════════════════════════════
if "DRIVER TREE" in wb.sheetnames:
    del wb["DRIVER TREE"]
ws = wb.create_sheet("DRIVER TREE")

# ── Column widths ──
col_widths = {
    "A": 1.5,
    "B": 34,   # driver label
    "C": 9,    # source
    "D": 22,   # metric name
    "E": 7,    # unit
    "F": 10,   # FY2025A
    "G": 10,   # FY2026E
    "H": 10,   # FY2028E
    "I": 10,   # Delta
    "J": 12,   # Rev Impact
    "K": 9,    # EBITDA Mgn
    "L": 12,   # EBITDA Impact
    "M": 13,   # EV Impact
    "N": 10,   # MOIC Delta
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze panes
ws.freeze_panes = "F7"

# ── Row 1: Title ──
ws.row_dimensions[1].height = 22
ws.cell(row=1, column=2, value="DRIVER TREE  ·  SHOPIFY DEAL WORKBOOK").font = Font(
    name=FONT_MAIN, bold=True, color=C_NAVY, size=11)
ws.cell(row=1, column=2).alignment = align("left")

# ── Row 2: Subtitle ──
ws.row_dimensions[2].height = 14
ws.cell(row=2, column=2,
    value="Revenue driver decomposition → EBITDA cascade → EV/MOIC attribution  |  Entry: FY2026E  |  Exit: FY2028E  |  Entry Multiple: 48.5×  |  Exit Multiple: 45.0×"
).font = font(italic=True, color=C_SLATE, sz=7)

# ── Row 3: Methodology note ──
ws.row_dimensions[3].height = 14
ws.cell(row=3, column=2,
    value="Decomposition: GMV volume (at entry attach rate) + attach rate improvement (at exit GMV). EBITDA impact uses segment incremental margins calibrated to model. EV impact uses entry multiple (Paasche). Sources: [Model] = model.xlsx, [IC Memo] = Shopify IC Memo v2, [Analyst] = analyst allocation."
).font = font(italic=True, color=C_SLATE, sz=7)
ws.cell(row=3, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Row 4: blank spacer ──
ws.row_dimensions[4].height = 6

# ── Row 5: Section header ──
ws.row_dimensions[5].height = 16
write_section_header(ws, 5, "REVENUE DRIVER DECOMPOSITION  →  EBITDA  →  EV  →  MOIC", 2, 14)

# ── Row 6: Column headers ──
ws.row_dimensions[6].height = 32
col_hdrs = ["Driver", "Source", "KPI Metric", "Unit",
            "FY2025A", "FY2026E\n(Entry)", "FY2028E\n(Exit)", "Δ (28E–26E)",
            "Revenue\nImpact ($M)", "Incr.\nEBITDA %", "EBITDA\nImpact ($M)",
            "EV Impact\n($M, ×48.5)", "MOIC\nDelta (×)"]
write_col_header(ws, 6, col_hdrs, start_col=2)
ws.cell(row=6, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
for c in range(3, 15):
    ws.cell(row=6, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── ROW LAYOUT ──────────────────────────────────────────────────────
r = 7

# ━━━ SECTION: Merchant Solutions ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws.row_dimensions[r].height = 14
write_subhdr(ws, r, "MERCHANT SOLUTIONS REVENUE", col=2, indent=0)
ws.cell(row=r, column=2).value = "MERCHANT SOLUTIONS REVENUE"
ws.cell(row=r, column=2).font = font(bold=True, color=C_INDIGO)
# Show total merch rev metrics in data columns
for c, (val, fmt) in enumerate([
    (MERCH_REV_26E, '#,##0'), (MERCH_REV_28E, '#,##0'), (MERCH_REV_28E - MERCH_REV_26E, '+#,##0;-#,##0;"-"')
], start=7):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font(bold=True, color=C_INDIGO)
    cell.alignment = align("right")
    cell.number_format = fmt
ws.cell(row=r, column=6).value = "$M"
ws.cell(row=r, column=6).font = font(color=C_SLATE, sz=7)
r += 1

# ── GMV Volume Effect sub-header ──
ws.row_dimensions[r].height = 12
write_subhdr(ws, r, "  GMV Growth → Volume Effect", col=2, indent=0)
ws.cell(row=r, column=2).value = "  GMV Growth → Volume Effect"
ws.cell(row=r, column=2).font = font(bold=True, color=C_BLUE, sz=SZ_BODY - 1)
ws.cell(row=r, column=4, value="GMV growth × entry attach rate (240.9 bps)").font = font(italic=True, color=C_SLATE, sz=7)
r += 1

# GMV driver rows
gmv_drivers = [
    # (label, source, metric, unit, val25a, val26e, val28e, delta_gmv_alloc$B, rev_impact)
    ("Same-Cohort Productivity", "Analyst",
     "Existing merchant GMV expansion (40% alloc.)", "$B",
     GMV_25A * PCT_SAME_COHORT,
     GMV_26E * PCT_SAME_COHORT,
     GMV_28E * PCT_SAME_COHORT,
     GMV_DLT * PCT_SAME_COHORT,
     REV_SAME_COHORT, MERCH_INCR_EBITDA, EBITDA_SAME_COHORT, EV_SAME_COHORT, MOIC_SAME_COHORT),
    ("New Merchant Acquisition", "Analyst",
     "Net new merchant cohort GMV (30% alloc.)", "$B",
     GMV_25A * PCT_NEW_MERCHANT,
     GMV_26E * PCT_NEW_MERCHANT,
     GMV_28E * PCT_NEW_MERCHANT,
     GMV_DLT * PCT_NEW_MERCHANT,
     REV_NEW_MERCHANT, MERCH_INCR_EBITDA, EBITDA_NEW_MERCHANT, EV_NEW_MERCHANT, MOIC_NEW_MERCHANT),
    ("Geographic Expansion (International)", "IC Memo",
     "International GMV (Europe, LatAm, Asia) — +45% YoY Q1'26 (20% alloc.)", "$B",
     GMV_25A * PCT_GEOGRAPHIC,
     GMV_26E * PCT_GEOGRAPHIC,
     GMV_28E * PCT_GEOGRAPHIC,
     GMV_DLT * PCT_GEOGRAPHIC,
     REV_GEOGRAPHIC, MERCH_INCR_EBITDA, EBITDA_GEOGRAPHIC, EV_GEOGRAPHIC, MOIC_GEOGRAPHIC),
    ("B2B / POS / Channel Expansion", "IC Memo",
     "B2B GMV (+80% Q1'26), Offline POS (+33%), Shop App (10% alloc.)", "$B",
     GMV_25A * PCT_B2B_CHANNEL,
     GMV_26E * PCT_B2B_CHANNEL,
     GMV_28E * PCT_B2B_CHANNEL,
     GMV_DLT * PCT_B2B_CHANNEL,
     REV_B2B, MERCH_INCR_EBITDA, EBITDA_B2B, EV_B2B, MOIC_B2B),
]

for label, source, metric, unit, v25, v26, v28, dlt, rev, em, ebitda, ev, moic_d in gmv_drivers:
    ws.row_dimensions[r].height = 14
    write_driver_row(ws, r, indent=2, label=label, source=source, metric=metric, unit=unit,
                     val_25a=v25, val_26e=v26, val_28e=v28, delta=dlt,
                     rev_impact=rev, ebitda_margin=em, ebitda_imp=ebitda, ev_imp=ev, moic_delta=moic_d)
    r += 1

# GMV Volume subtotal
ws.row_dimensions[r].height = 13
write_driver_row(ws, r, indent=1, label="  Subtotal: GMV Volume Effect", source="", metric="", unit="",
                 val_25a=None, val_26e=None, val_28e=None, delta=GMV_DLT,
                 rev_impact=GMV_VOL_REV,
                 ebitda_margin=MERCH_INCR_EBITDA,
                 ebitda_imp=GMV_VOL_REV * MERCH_INCR_EBITDA,
                 ev_imp=(GMV_VOL_REV * MERCH_INCR_EBITDA) * ENTRY_MULTIPLE,
                 moic_delta=(GMV_VOL_REV * MERCH_INCR_EBITDA * ENTRY_MULTIPLE) / ENTRY_EQUITY,
                 is_subtotal=True)
# Override the unit col for GMV delta
ws.cell(row=r, column=9).value = round(GMV_DLT, 1)
ws.cell(row=r, column=9).number_format = "#,##0.0"
ws.cell(row=r-4, column=9).number_format = "#,##0.0"  # fix delta col format
r += 1

# ── Attach Rate Effect sub-header ──
ws.row_dimensions[r].height = 12
write_subhdr(ws, r, "  Attach Rate Improvement", col=2, indent=0)
ws.cell(row=r, column=2).value = "  Attach Rate Improvement"
ws.cell(row=r, column=2).font = font(bold=True, color=C_BLUE, sz=SZ_BODY - 1)
ws.cell(row=r, column=4, value="Attach rate (bps) improvement × FY2028E GMV of $699.6B").font = font(italic=True, color=C_SLATE, sz=7)
r += 1

# Attach rate drivers
attach_drivers = [
    ("GPV Penetration (71% → 79%)", "Model + IC Memo",
     "Shopify Payments GPV% — 100bps = $124M revenue at exit GMV", "%",
     GPV_25A, GPV_26E, GPV_28E, GPV_28E - GPV_26E,
     REV_GPV_PENETRATION, MERCH_INCR_EBITDA, EBITDA_GPV, EV_GPV, MOIC_GPV),
    ("Non-Payments Attach (MCA, Balance, Markets)", "Model",
     "Shopify Capital, Balance, Markets — incremental bps at exit GMV", "bps",
     None, None, None, round((ATTACH_28E - ATTACH_26E) - (GPV_DLT_PP * 18.0 / 100), 1),
     REV_NON_PAYMENTS, MERCH_INCR_EBITDA, EBITDA_NON_PAYMENTS, EV_NON_PAYMENTS, MOIC_NON_PAYMENTS),
]

for label, source, metric, unit, v25, v26, v28, dlt, rev, em, ebitda_i, ev, moic_d in attach_drivers:
    ws.row_dimensions[r].height = 14
    write_driver_row(ws, r, indent=2, label=label, source=source, metric=metric, unit=unit,
                     val_25a=v25 if v25 is not None else 0,
                     val_26e=v26 if v26 is not None else 0,
                     val_28e=v28 if v28 is not None else 0,
                     delta=dlt,
                     rev_impact=rev, ebitda_margin=em, ebitda_imp=ebitda_i, ev_imp=ev, moic_delta=moic_d)
    r += 1

# Attach rate subtotal
ws.row_dimensions[r].height = 13
write_driver_row(ws, r, indent=1, label="  Subtotal: Attach Rate Improvement", source="", metric="", unit="",
                 val_25a=None, val_26e=ATTACH_26E, val_28e=ATTACH_28E, delta=ATTACH_28E - ATTACH_26E,
                 rev_impact=ATTACH_RATE_REV,
                 ebitda_margin=MERCH_INCR_EBITDA,
                 ebitda_imp=ATTACH_RATE_REV * MERCH_INCR_EBITDA,
                 ev_imp=ATTACH_RATE_REV * MERCH_INCR_EBITDA * ENTRY_MULTIPLE,
                 moic_delta=ATTACH_RATE_REV * MERCH_INCR_EBITDA * ENTRY_MULTIPLE / ENTRY_EQUITY,
                 is_subtotal=True)
r += 1

# Merchant Solutions total
ws.row_dimensions[r].height = 15
merch_ebitda_total = (GMV_VOL_REV + ATTACH_RATE_REV) * MERCH_INCR_EBITDA
write_driver_row(ws, r, indent=0, label="TOTAL: Merchant Solutions Revenue", source="Model", metric="", unit="$M",
                 val_25a=None, val_26e=MERCH_REV_26E, val_28e=MERCH_REV_28E, delta=MERCH_REV_DLT,
                 rev_impact=MERCH_REV_DLT,
                 ebitda_margin=MERCH_INCR_EBITDA,
                 ebitda_imp=merch_ebitda_total,
                 ev_imp=merch_ebitda_total * ENTRY_MULTIPLE,
                 moic_delta=merch_ebitda_total * ENTRY_MULTIPLE / ENTRY_EQUITY,
                 is_subtotal=True)
ws.cell(row=r, column=2).font = font(bold=True, color=C_BLACK)
r += 1

# ── Blank spacer ──
ws.row_dimensions[r].height = 6
r += 1

# ━━━ SECTION: Subscription Solutions ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws.row_dimensions[r].height = 14
write_subhdr(ws, r, "SUBSCRIPTION SOLUTIONS REVENUE", col=2, indent=0)
ws.cell(row=r, column=2).value = "SUBSCRIPTION SOLUTIONS REVENUE"
ws.cell(row=r, column=2).font = font(bold=True, color=C_INDIGO)
for c, (val, fmt) in enumerate([
    (SUB_REV_26E, '#,##0'), (SUB_REV_28E, '#,##0'), (SUB_REV_DLT, '+#,##0;-#,##0;"-"')
], start=7):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font(bold=True, color=C_INDIGO)
    cell.alignment = align("right")
    cell.number_format = fmt
ws.cell(row=r, column=6).value = "$M"
ws.cell(row=r, column=6).font = font(color=C_SLATE, sz=7)
r += 1

# Sub revenue build note
ws.row_dimensions[r].height = 11
ws.cell(row=r, column=2, value="    MRR Build: Revenue = MRR × 13.8× (implied annual multiplier; model MRR × 12 + deferred revenue adj.)").font = font(italic=True, color=C_SLATE, sz=7)
ws.cell(row=r, column=2).alignment = align("left")
r += 1

sub_drivers = [
    ("Plus / Enterprise MRR Growth", "Model",
     "Plus MRR ($M) — enterprise cohort (≥$25M GMV, fastest-growing Q1'26)", "$M MRR",
     PLUS_MRR_25A, PLUS_MRR_26E, PLUS_MRR_28E, PLUS_MRR_28E - PLUS_MRR_26E,
     REV_PLUS_MRR, SUB_INCR_EBITDA, EBITDA_PLUS_MRR, EV_PLUS_MRR, MOIC_PLUS_MRR),
    ("Core SMB MRR Growth", "Model",
     "Core MRR ($M) — basic/standard/advanced plan merchants", "$M MRR",
     CORE_MRR_25A, CORE_MRR_26E, CORE_MRR_28E, CORE_MRR_28E - CORE_MRR_26E,
     REV_CORE_MRR, SUB_INCR_EBITDA, EBITDA_CORE_MRR, EV_CORE_MRR, MOIC_CORE_MRR),
]
if abs(REV_SUB_OTHER) > 5:
    sub_drivers.append(
        ("Billing / Deferred Revenue Adjustment", "Model",
         "MRR multiplier change and deferred revenue timing", "$M",
         None, None, None, None,
         REV_SUB_OTHER, SUB_INCR_EBITDA, EBITDA_SUB_OTHER, EV_SUB_OTHER, MOIC_SUB_OTHER)
    )

for label, source, metric, unit, v25, v26, v28, dlt, rev, em, ebitda_i, ev, moic_d in sub_drivers:
    ws.row_dimensions[r].height = 14
    write_driver_row(ws, r, indent=2, label=label, source=source, metric=metric, unit=unit,
                     val_25a=v25 if v25 is not None else 0,
                     val_26e=v26 if v26 is not None else 0,
                     val_28e=v28 if v28 is not None else 0,
                     delta=dlt if dlt is not None else 0,
                     rev_impact=rev, ebitda_margin=em, ebitda_imp=ebitda_i, ev_imp=ev, moic_delta=moic_d)
    r += 1

# Sub total
ws.row_dimensions[r].height = 15
sub_ebitda_total = SUB_REV_DLT * SUB_INCR_EBITDA
write_driver_row(ws, r, indent=0, label="TOTAL: Subscription Solutions Revenue", source="Model", metric="", unit="$M",
                 val_25a=None, val_26e=SUB_REV_26E, val_28e=SUB_REV_28E, delta=SUB_REV_DLT,
                 rev_impact=SUB_REV_DLT,
                 ebitda_margin=SUB_INCR_EBITDA,
                 ebitda_imp=sub_ebitda_total,
                 ev_imp=sub_ebitda_total * ENTRY_MULTIPLE,
                 moic_delta=sub_ebitda_total * ENTRY_MULTIPLE / ENTRY_EQUITY,
                 is_subtotal=True)
ws.cell(row=r, column=2).font = font(bold=True, color=C_BLACK)
r += 1

# ── Blank spacer ──
ws.row_dimensions[r].height = 6
r += 1

# ━━━ TOTAL REVENUE ROW ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws.row_dimensions[r].height = 16
total_rev_ebitda = merch_ebitda_total + sub_ebitda_total
write_driver_row(ws, r, indent=0, label="TOTAL REVENUE GROWTH", source="Model", metric="", unit="$M",
                 val_25a=None, val_26e=TOT_REV_26E, val_28e=TOT_REV_28E, delta=TOT_REV_28E - TOT_REV_26E,
                 rev_impact=TOT_REV_28E - TOT_REV_26E,
                 ebitda_margin=EBITDA_DLT / (TOT_REV_28E - TOT_REV_26E),
                 ebitda_imp=total_rev_ebitda,
                 ev_imp=total_rev_ebitda * ENTRY_MULTIPLE,
                 moic_delta=total_rev_ebitda * ENTRY_MULTIPLE / ENTRY_EQUITY,
                 is_subtotal=True)
ws.cell(row=r, column=2).font = font(bold=True, color=C_NAVY, sz=SZ_HDR)
ws.cell(row=r, column=2).fill = fill(C_TEAL_HL)
for c in range(3, 15):
    ws.cell(row=r, column=c).fill = fill(C_TEAL_HL)
r += 1

# ── Blank spacer ──
ws.row_dimensions[r].height = 8
r += 1

# ━━━ VALUE CREATION BRIDGE ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws.row_dimensions[r].height = 16
write_section_header(ws, r, "VALUE CREATION BRIDGE  (Reconciliation: Revenue Drivers + Multiple Compression + Net Cash)", 2, 14)
r += 1

bridge_rows = [
    ("Revenue-Driven EBITDA Growth (×48.5 entry multiple)", C_BLACK,
     total_rev_ebitda, ENTRY_MULTIPLE, total_rev_ebitda * ENTRY_MULTIPLE,
     total_rev_ebitda * ENTRY_MULTIPLE / ENTRY_EQUITY, False),
    ("Multiple Compression (48.5× → 45.0×)  ×  Exit EBITDA $4,687M", C_RED,
     EBITDA_28E * (EXIT_MULTIPLE - ENTRY_MULTIPLE), 1.0, EV_MULTI_COMPRESS,
     moic(EV_MULTI_COMPRESS), True),
    ("Net Cash Build (Net Cash $6.3B → $12.4B)", C_GREEN,
     NET_CASH_BUILD, 1.0, EV_NET_CASH,
     moic(EV_NET_CASH), False),
]

bridge_labels = ["Bridge Component", "", "", "",
                 "FY2026E", "", "FY2028E", "",
                 "EBITDA Effect ($M)", "×", "EV Impact ($M)", "MOIC Impact (×)"]
ws.row_dimensions[r].height = 28
for i, lbl in enumerate(bridge_labels):
    c = ws.cell(row=r, column=2 + i, value=lbl)
    c.font = font(bold=True, color=C_WHITE, sz=SZ_BODY - 1)
    c.fill = fill(C_NAVY)
    c.alignment = Alignment(horizontal="center" if i > 0 else "left", vertical="center", wrap_text=True)
    c.border = border_thin()
r += 1

for label, color, ebitda_eff, mult, ev_eff, moic_eff, is_neg in bridge_rows:
    ws.row_dimensions[r].height = 14
    bg = C_GREY_ALT if r % 2 == 0 else C_WHITE
    ws.cell(row=r, column=2, value=label).font = font(bold=True, color=color)
    ws.cell(row=r, column=2).fill = fill(bg)
    ws.cell(row=r, column=2).border = border_thin()
    ws.cell(row=r, column=2).alignment = align("left")
    for c in range(3, 10):
        ws.cell(row=r, column=c).fill = fill(bg)
        ws.cell(row=r, column=c).border = border_thin()
    for col, val, fmt in [(10, ebitda_eff, '#,##0;(#,##0);"-"'),
                           (11, mult, '0.0"×"'),
                           (12, ev_eff, '#,##0;(#,##0);"-"'),
                           (14, moic_eff, '+0.000x;-0.000x;"-"')]:
        cell = ws.cell(row=r, column=col, value=round(val, 3) if isinstance(val, float) else val)
        cell.font = font(bold=True, color=color)
        cell.fill = fill(bg)
        cell.alignment = align("right")
        cell.border = border_thin()
        cell.number_format = fmt
    r += 1

# TOTAL VALUE CREATION row
ws.row_dimensions[r].height = 16
ws.cell(row=r, column=2, value="TOTAL VALUE CREATION (Entry Equity → Exit Proceeds)").font = font(bold=True, color=C_WHITE, sz=SZ_HDR)
ws.cell(row=r, column=2).fill = fill(C_NAVY)
ws.cell(row=r, column=2).alignment = align("left")
ws.cell(row=r, column=2).border = border_thin()
totals = [
    (10, total_rev_ebitda + EBITDA_28E * (EXIT_MULTIPLE - ENTRY_MULTIPLE), '#,##0;(#,##0)'),
    (12, EV_GRAND_TOTAL, '#,##0;(#,##0)'),
    (14, moic(EV_GRAND_TOTAL), '+0.000x;-0.000x'),
]
for col, val, fmt in totals:
    cell = ws.cell(row=r, column=col, value=round(val, 2))
    cell.font = font(bold=True, color=C_WHITE, sz=SZ_HDR)
    cell.fill = fill(C_NAVY)
    cell.alignment = align("right")
    cell.border = border_thin()
    cell.number_format = fmt
for c in range(3, 15):
    if c not in [10, 12, 14]:
        ws.cell(row=r, column=c).fill = fill(C_NAVY)
        ws.cell(row=r, column=c).border = border_thin()
r += 1

# MOIC CHECK
ws.row_dimensions[r].height = 18
ws.cell(row=r, column=2, value="MOIC CHECK  (1.0× invested capital + attribution)").font = font(bold=True, color=C_NAVY, sz=SZ_HDR)
ws.cell(row=r, column=2).fill = fill(C_TEAL_HL)
ws.cell(row=r, column=2).alignment = align("left")
ws.cell(row=r, column=2).border = border_thin()
moic_check_val = 1.0 + moic(EV_GRAND_TOTAL)
cell = ws.cell(row=r, column=14, value=round(moic_check_val, 3))
cell.font = font(bold=True, color=C_NAVY, sz=SZ_HDR + 2)
cell.fill = fill(C_TEAL_HL)
cell.alignment = align("right")
cell.number_format = '0.00"×"'
cell.border = border_thin()
for c in range(3, 14):
    ws.cell(row=r, column=c).fill = fill(C_TEAL_HL)
    ws.cell(row=r, column=c).border = border_thin()
r += 1

# ── Assumptions footnote ──
r += 1
ws.row_dimensions[r].height = 11
ws.cell(row=r, column=2,
    value="Assumptions: GMV sub-driver allocations (same-cohort 40%, new merchant 30%, geographic 20%, B2B/channel 10%) are analyst estimates based on IC Memo Pillars 1-2. "
          "GPV penetration revenue = IC Memo disclosure ($67M per 100bps at FY2025A GMV) scaled to exit GMV. "
          "Incr. EBITDA margins: Merchant ~21.4% (37.7% seg. GM − ~16.3% opex allocation), Sub ~67.6% (82.0% seg. GM − 14.4% opex). "
          "EV computed using entry multiple (48.5×) per Paasche decomposition; total value creation reconciles to INPUTS!C59."
).font = font(italic=True, color=C_SLATE, sz=7)
ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[r].height = 28


# ════════════════════════════════════════════════════════════════════
# NTB REGISTRY TAB (rebuilt to reference driver tree results)
# ════════════════════════════════════════════════════════════════════
if "NTB REGISTRY" in wb.sheetnames:
    del wb["NTB REGISTRY"]
ntb = wb.create_sheet("NTB REGISTRY")

ntb.column_dimensions["A"].width = 1.5
ntb.column_dimensions["B"].width = 8
ntb.column_dimensions["C"].width = 28
ntb.column_dimensions["D"].width = 42
ntb.column_dimensions["E"].width = 14
ntb.column_dimensions["F"].width = 14
ntb.column_dimensions["G"].width = 14
ntb.column_dimensions["H"].width = 14
ntb.column_dimensions["I"].width = 14
ntb.column_dimensions["J"].width = 20

ntb.freeze_panes = "C7"

# Title
ntb.row_dimensions[1].height = 22
ntb.cell(row=1, column=2, value="NTB REGISTRY  ·  SHOPIFY DEAL WORKBOOK").font = Font(
    name=FONT_MAIN, bold=True, color=C_NAVY, size=11)

ntb.row_dimensions[2].height = 14
ntb.cell(row=2, column=2,
    value="Investment thesis mapped to revenue drivers → EBITDA → EV → MOIC attribution. Base = model FY2028E. Upside = IC Memo bull case."
).font = font(italic=True, color=C_SLATE, sz=7)

ntb.row_dimensions[3].height = 14
ntb.cell(row=3, column=2,
    value="Source: IC Memo v2 Pillars 1-4 mapped to DRIVER TREE. Base EV uses Paasche decomposition (×48.5 entry multiple). Upside = analyst thesis above model."
).font = font(italic=True, color=C_SLATE, sz=7)

ntb.row_dimensions[4].height = 6

# Section header
ntb.row_dimensions[5].height = 16
for c in range(2, 11):
    ntb.cell(row=5, column=c).fill = fill(C_NAVY)
ntb.cell(row=5, column=2, value="NTB REGISTRY  —  THESIS TO VALUE ATTRIBUTION").font = font(bold=True, color=C_WHITE, sz=SZ_HDR)
ntb.cell(row=5, column=2).alignment = align("left")

# Column headers
ntb.row_dimensions[6].height = 32
ntb_hdrs = ["NTB", "Theme", "Thesis Summary (Key Assumption vs. Consensus)",
            "Base EV\nImpact ($M)", "Base MOIC\nDelta (×)",
            "Upside EV\n($M)", "Upside MOIC\n(×)",
            "IC Memo Pillar", "Driver Tree Rows", "Rating / Status"]
for i, hdr in enumerate(ntb_hdrs):
    c = ntb.cell(row=6, column=2 + i, value=hdr)
    c.font = font(bold=True, color=C_WHITE, sz=7)
    c.fill = fill(C_NAVY)
    c.alignment = Alignment(horizontal="center" if i > 1 else "left", vertical="center", wrap_text=True)
    c.border = border_thin()

# NTB data
# Base EV for each NTB (sum of constituent driver EV impacts)
ntb1_base_ev = EV_SAME_COHORT + EV_NEW_MERCHANT + EV_GEOGRAPHIC + EV_B2B
ntb2_base_ev = EV_GPV + EV_NON_PAYMENTS
ntb3_base_ev = 0           # AI not yet in model base; pure optionality
ntb4_base_ev = EV_PLUS_MRR + EV_CORE_MRR
ntb5_base_ev = EV_MULTI_COMPRESS  # multiple compression risk (negative)
cash_ev      = EV_NET_CASH

ntb1_up = ntb1_base_ev * 1.30   # 30% upside: GMV growth 2pp faster vs model
ntb2_up = ntb2_base_ev * 1.40   # 40% upside: GPV hits 83% vs 79% in model
ntb3_up = 40000                 # AI optionality: $15K Sidekick + $25K UCP = $40B EV upside
ntb4_up = ntb4_base_ev * 1.35   # 35% upside: Plus accelerates to 45% of MRR

ntb_rows = [
    (1, "GMV Flywheel",
     "Thesis: Multiple independent GMV growth vectors (same-cohort expansion, new merchant acquisition, "
     "geographic expansion to 15 intl markets, B2B +80% YoY) sustain 20-22% GMV CAGR through FY2028E "
     "vs. consensus expectation of GMV deceleration driven by FX and tougher comps. Key: geographic "
     "and B2B channels reduce single-point-of-failure risk. [IC Memo: Pillar 1 — LOW evidence strength, THESIS-CRITICAL]",
     ntb1_base_ev, ntb1_base_ev / ENTRY_EQUITY,
     ntb1_up, ntb1_up / ENTRY_EQUITY,
     "Pillar 1", "Same-Cohort, New Merchant,\nGeographic, B2B/Channel", "OPEN — thesis-critical"),
    (2, "Payments Monetization",
     "Thesis: GPV penetration grows 71% → 79% FY2026-28E driven by Shop Pay expansion on non-Shopify "
     "storefronts, checkout optimization, and merchant incentives. Each 100bps penetration = $124M incremental "
     "revenue at exit GMV ($699.6B). Model assumes 8pp penetration gain; consensus (sell-side avg) ~74% FY2028E "
     "implies 5pp below model — model is already thesis-aligned. Non-payments (MCA, Balance) add $268M incremental. "
     "[IC Memo: Pillar 2 — HIGH evidence strength, Payments thesis CONFIRMED]",
     ntb2_base_ev, ntb2_base_ev / ENTRY_EQUITY,
     ntb2_up, ntb2_up / ENTRY_EQUITY,
     "Pillar 2", "GPV Penetration,\nNon-Payments Attach", "CONFIRMED — HIGH evidence"),
    (3, "AI Commerce Platform (Optionality)",
     "Thesis: Shopify's 20+ year proprietary commerce data creates a structural moat vs. general-purpose "
     "AI (2× conversion rate on AI-assisted orders vs. organic search Q1'26). AI revenue streams (Sidekick "
     "subscription, Usage-based Commerce Protocols) not yet in base model — pure optionality upside. "
     "No FY2028E base case contribution modeled. Upside scenario: $40B EV ($15B Sidekick + $25B UCP). "
     "[IC Memo: Pillar 3 — MEDIUM evidence strength, THESIS-CRITICAL for long-term]",
     ntb3_base_ev, ntb3_base_ev / ENTRY_EQUITY,
     ntb3_up, ntb3_up / ENTRY_EQUITY,
     "Pillar 3", "AI Sidekick (C1),\nAI UCP (C3) — upside only", "OPEN — hypothesis, not in model"),
    (4, "Enterprise Upmarket",
     "Thesis: $25M+ GMV cohort is fastest-growing segment Q1'26 with no SMB cannibalization. Plus MRR "
     "grows from $83.6M (FY2026E, 35% of MRR) to $114.5M (FY2028E, 39% of MRR) driving $427M incremental "
     "subscription revenue at ~82% gross margin. High-margin subscription mix shift benefits blended margins. "
     "Upside: Plus penetration accelerates to 45% of MRR ($137M) by FY2028E. "
     "[IC Memo: Pillar 4 — MEDIUM evidence strength, VALUATION OPTIONALITY]",
     ntb4_base_ev, ntb4_base_ev / ENTRY_EQUITY,
     ntb4_up, ntb4_up / ENTRY_EQUITY,
     "Pillar 4", "Plus/Enterprise MRR,\nCore SMB MRR", "WATCH — valuation optionality"),
    (5, "Multiple Compression (Risk)",
     "Risk: Entry at 48.5× EV/EBITDA creates structural headwind. Exit at 45× already assumes compression. "
     "Each 1× turn of additional compression = $4,687M additional EV reduction = -0.033x MOIC drag. "
     "Downside scenario: exit at 40× → additional -$23.4B EV → MOIC = 1.41× (vs. 1.58× base). "
     "Offsetting factor: EBITDA growth nearly triples ($2.79B → $4.69B), supporting multiple. "
     "[Macro risk: consumer softness, tariff impact on SMB GMV]",
     ntb5_base_ev, ntb5_base_ev / ENTRY_EQUITY,
     ntb5_base_ev * 0.5, (ntb5_base_ev * 0.5) / ENTRY_EQUITY,  # upside = compression recovers
     "Macro / Risk", "Multiple Compression,\nMacro Consumer Risk", "STRUCTURAL — mitigant: EBITDA growth"),
]

r_ntb = 7
for ntb_num, theme, thesis, base_ev, base_moic, up_ev, up_moic, pillar, drivers_ref, status in ntb_rows:
    ntb.row_dimensions[r_ntb].height = 85
    bg = C_GREY_ALT if r_ntb % 2 == 0 else C_WHITE
    is_risk = ntb_num == 5

    # NTB number
    cell = ntb.cell(row=r_ntb, column=2, value=f"NTB {ntb_num}")
    cell.font = font(bold=True, color=C_WHITE if not is_risk else C_WHITE, sz=SZ_BODY)
    cell.fill = fill(C_NAVY if not is_risk else C_RED)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin()

    # Theme
    cell = ntb.cell(row=r_ntb, column=3, value=theme)
    cell.font = font(bold=True, color=C_BLUE if not is_risk else C_RED)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = border_thin()

    # Thesis
    cell = ntb.cell(row=r_ntb, column=4, value=thesis)
    cell.font = font(color=C_BLACK, sz=7)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = border_thin()

    # Numeric columns
    num_data = [
        (5, base_ev,   '#,##0;(#,##0);"-"', C_BLACK if not is_risk else C_RED),
        (6, base_moic, '+0.000x;-0.000x;"-"', C_GREEN if base_moic > 0 else C_RED),
        (7, up_ev,     '#,##0;(#,##0);"-"', C_GREEN if up_ev > 0 else C_RED),
        (8, up_moic,   '+0.000x;-0.000x;"-"', C_GREEN if up_moic > 0 else C_RED),
    ]
    for col, val, fmt, clr in num_data:
        cell = ntb.cell(row=r_ntb, column=col, value=round(val, 3))
        cell.font = font(bold=True, color=clr)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = border_thin()
        cell.number_format = fmt

    # Pillar
    cell = ntb.cell(row=r_ntb, column=9, value=pillar)
    cell.font = font(color=C_SLATE, sz=7)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    cell.border = border_thin()

    # Drivers ref
    cell = ntb.cell(row=r_ntb, column=10, value=drivers_ref)
    cell.font = font(color=C_SLATE, sz=7)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = border_thin()

    # Status
    status_color = (C_GREEN if "CONFIRMED" in status else
                    C_RED if "STRUCTURAL" in status else
                    C_ORANGE)
    cell = ntb.cell(row=r_ntb, column=11, value=status)
    cell.font = font(bold=True, color=status_color, sz=7)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = border_thin()

    r_ntb += 1

# NTB Summary / Grand Total
ntb.row_dimensions[r_ntb].height = 16
for c in range(2, 11):
    ntb.cell(row=r_ntb, column=c).fill = fill(C_TEAL_HL)
    ntb.cell(row=r_ntb, column=c).border = border_thin()
ntb.cell(row=r_ntb, column=2, value="TOTAL").font = font(bold=True, color=C_NAVY)
ntb.cell(row=r_ntb, column=3, value="Base + Multiple Compression + Net Cash").font = font(bold=True, color=C_NAVY)
ntb.cell(row=r_ntb, column=3).alignment = align("left")

total_base_ntb = ntb1_base_ev + ntb2_base_ev + ntb3_base_ev + ntb4_base_ev + ntb5_base_ev + cash_ev
for col, val, fmt in [
    (5, total_base_ntb, '#,##0;(#,##0)'),
    (6, total_base_ntb / ENTRY_EQUITY, '+0.000x;-0.000x'),
]:
    cell = ntb.cell(row=r_ntb, column=col, value=round(val, 2))
    cell.font = font(bold=True, color=C_NAVY, sz=SZ_HDR)
    cell.fill = fill(C_TEAL_HL)
    cell.alignment = align("right")
    cell.border = border_thin()
    cell.number_format = fmt

# MOIC Check row
r_ntb += 1
ntb.row_dimensions[r_ntb].height = 18
for c in range(2, 11):
    ntb.cell(row=r_ntb, column=c).fill = fill(C_NAVY)
    ntb.cell(row=r_ntb, column=c).border = border_thin()
ntb.cell(row=r_ntb, column=2, value="MOIC CHECK").font = font(bold=True, color=C_WHITE, sz=SZ_HDR)
ntb.cell(row=r_ntb, column=3, value="= 1.00× invested capital + total attribution (should = 1.58×)").font = font(bold=True, color=C_WHITE, sz=7)
ntb.cell(row=r_ntb, column=3).fill = fill(C_NAVY)
ntb.cell(row=r_ntb, column=3).alignment = align("left")
moic_check_ntb = 1.0 + (total_base_ntb / ENTRY_EQUITY)
cell = ntb.cell(row=r_ntb, column=6, value=round(moic_check_ntb, 3))
cell.font = font(bold=True, color=C_WHITE, sz=14)
cell.fill = fill(C_NAVY)
cell.alignment = align("right")
cell.number_format = '0.00"×"'
cell.border = border_thin()

# ─── SAVE ─────────────────────────────────────────────────────────────────────
out_path = r"C:\Users\IanLawrence\github\Claude-skills\docs\Shopify_Deal_Workbook_v3.xlsx"
wb.save(out_path)
print(f"\nSaved: {out_path}")
