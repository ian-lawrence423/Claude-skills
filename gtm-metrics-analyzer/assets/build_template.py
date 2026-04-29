from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / 'gtm_metrics_template.xlsx'
OUT.parent.mkdir(parents=True, exist_ok=True)

# ── Pattern brand constants ───────────────────────────────────────────────
NAVY    = '0F4761'
BLUE    = '4280F4'
LT_BLUE = 'EBF2FA'
ALT_ROW = 'F5F8FF'
BORDER  = 'DDDDDD'
WHITE   = 'FFFFFF'
BLACK   = '000000'
GRAY    = '888888'
INPUT_V = 'EAF5EA'    # light mint — signals "fill this column"

_b  = Side(style='thin', color=BORDER)
_bh = Side(style='medium', color=NAVY)
BDR       = Border(left=_b, right=_b, top=_b, bottom=_b)
WRAP_TOP  = Alignment(horizontal='left', vertical='top', wrap_text=True)
LEFT_MID  = Alignment(horizontal='left', vertical='center')
CTR       = Alignment(horizontal='center', vertical='center')
CTR_WRAP  = Alignment(horizontal='center', vertical='center', wrap_text=True)

def fill(c):
    return PatternFill('solid', fgColor=c)

def font(name='Wix Madefor Display', size=9, bold=False, color=BLACK, italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

# ── Data ──────────────────────────────────────────────────────────────────

input_rows = [
    ('Company & Reporting Context','business_model','Business Model','e.g., B2B SaaS, usage-based SaaS, hybrid, PLG, SLG','Required',None),
    ('Company & Reporting Context','primary_revenue_metric','Primary Revenue Metric','ARR, MRR, CARR, bookings, billings, GAAP revenue','Required',None),
    ('Company & Reporting Context','reporting_cadence','Reporting Cadence','Monthly, quarterly, annual, or LTM','Required',None),
    ('Company & Reporting Context','analysis_period','Analysis Period','Period being analyzed','Required',None),
    ('Company & Reporting Context','segment_scheme','Segmentation Scheme','SMB, MM, Enterprise, geo, product, channel, etc.','Optional',None),
    ('Revenue & ARR','beginning_arr','Beginning ARR','ARR at start of period','Required for ARR funnel',0),
    ('Revenue & ARR','new_logo_arr','New Logo ARR','ARR from new customers in period','Required for ARR funnel',0),
    ('Revenue & ARR','expansion_arr','Expansion ARR','ARR expansion from existing customers','Required for ARR funnel/retention',0),
    ('Revenue & ARR','upsell_arr','Upsell ARR','Expansion from same product / contract uplift','Optional',0),
    ('Revenue & ARR','cross_sell_arr','Cross-sell ARR','Expansion from new product / contract','Optional',0),
    ('Revenue & ARR','downsell_arr','Downsell ARR','ARR lost from customer downgrades','Required for ARR funnel/retention',0),
    ('Revenue & ARR','logo_churn_arr','Logo Churn ARR','ARR lost from customer churn','Required for ARR funnel/retention',0),
    ('Revenue & ARR','ending_arr','Ending ARR','ARR at end of period if already known','Optional / validation',0),
    ('Revenue & ARR','bookings','Bookings','Total contract value signed in period','Optional',0),
    ('Revenue & ARR','billings','Billings','Amount invoiced in period','Optional',0),
    ('Revenue & ARR','revenue','Revenue','Recognized revenue in period','Optional',0),
    ('Revenue & ARR','deferred_revenue','Deferred Revenue','Billings less recognized revenue','Optional',0),
    ('Customer & Retention','bop_customers','Beginning Customer Count','Customers at start of period','Required for logo retention',0),
    ('Customer & Retention','eop_customers','Ending Customer Count','Customers at end of period','Required for logo retention',0),
    ('Customer & Retention','gross_new_logo_customers','Gross New Logo Customers','New logos acquired in period','Required for CAC',0),
    ('Customer & Retention','churned_customers','Churned Customers','Customers lost in period','Required for logo churn',0),
    ('Customer & Retention','logo_churn_rate','Logo Churn Rate','If already calculated, can be provided directly','Optional / shortcut',0),
    ('Customer & Retention','avg_arr_per_customer','Average ARR per Customer','Average ARR per retained customer','Required for LTV',0),
    ('Customer & Retention','dau','Daily Active Users','Customer-level or aggregate DAU','Optional',0),
    ('Customer & Retention','mau','Monthly Active Users','Customer-level or aggregate MAU','Optional',0),
    ('Customer & Retention','total_users','Total Users','Total licensed / relevant users','Optional',0),
    ('Customer & Retention','feature_active_users','Feature Active Users','Users actively using a feature','Optional',0),
    ('Customer & Retention','expected_implementation_days','Expected Implementation Days','Promised or target implementation duration','Optional',0),
    ('Customer & Retention','actual_implementation_days','Actual Implementation Days','Observed implementation duration','Optional',0),
    ('Customer & Retention','promoter_pct','Promoter %','Percent of promoters for NPS','Optional',0),
    ('Customer & Retention','detractor_pct','Detractor %','Percent of detractors for NPS','Optional',0),
    ('Customer & Retention','satisfied_responses','Satisfied Responses','Positive CSAT responses','Optional',0),
    ('Customer & Retention','total_responses','Total Responses','Total survey responses','Optional',0),
    ('Funnel & Pipeline','pipeline_dollars','Pipeline $','Unweighted pipeline value for period','Required for pipeline coverage',0),
    ('Funnel & Pipeline','weighted_pipeline_dollars','Weighted Pipeline $','Forecast-weighted pipeline value','Optional / preferred',0),
    ('Funnel & Pipeline','sales_target','Sales Target','Bookings / ARR target for period','Required for coverage',0),
    ('Funnel & Pipeline','forecast_dollars','Forecast $','Forecast for period','Required for forecast vs target',0),
    ('Funnel & Pipeline','closed_won_opportunities','Closed Won Opportunities','Count of won opportunities','Optional',0),
    ('Funnel & Pipeline','closed_lost_opportunities','Closed Lost Opportunities','Count of lost opportunities','Optional',0),
    ('Funnel & Pipeline','opportunities_created','Opportunities Created in Period','Used for close rate','Optional',0),
    ('Funnel & Pipeline','closed_won_same_period','Closed Won in Same Period','Used for close rate','Optional',0),
    ('Funnel & Pipeline','sql_count','SQL Count','Total sales qualified leads','Optional',0),
    ('Funnel & Pipeline','closed_won_sql','Closed Won SQLs','SQLs that closed won','Optional',0),
    ('Funnel & Pipeline','stage_forecast_probability','Average Stage Forecast Probability','If using a simplified weighted forecast','Optional',0),
    ('Efficiency & Economics','sm_opex','S&M OpEx (Current Period)','Sales & marketing operating expense in current period','Required for several metrics',0),
    ('Efficiency & Economics','prior_qtr_sm_opex','Prior Quarter S&M OpEx','Preferred for magic number and time-adjusted CAC','Required for lagged metrics',0),
    ('Efficiency & Economics','gross_profit','Gross Profit','Total gross profit in period','Required for gross margin',0),
    ('Efficiency & Economics','total_revenue','Total Revenue','Total revenue in period','Required for gross margin',0),
    ('Efficiency & Economics','subscription_gross_profit','Subscription Gross Profit','Gross profit from subscription revenue line','Optional',0),
    ('Efficiency & Economics','subscription_revenue','Subscription Revenue','Revenue from subscription line only','Optional',0),
    ('Efficiency & Economics','services_gross_profit','Services Gross Profit','Gross profit from professional services','Optional',0),
    ('Efficiency & Economics','services_revenue','Services Revenue','Revenue from professional services','Optional',0),
    ('Efficiency & Economics','gross_margin_pct','Gross Margin %','Can be input directly if already calculated','Optional / shortcut',0),
    ('Efficiency & Economics','channel_spend','Channel Spend','Spend for a specific channel','Optional',0),
    ('Efficiency & Economics','channel_new_logos','Channel New Logos','Customers acquired via channel','Optional',0),
    ('Team & Productivity','avg_sm_ftes','Average S&M FTEs','Average S&M employees in period','Required for FTE metrics',0),
    ('Team & Productivity','qcr_count','Quota-Carrying Reps','Average QCRs in period','Required for rep productivity/capacity',0),
    ('Team & Productivity','csm_count','CSM Count','Customer success managers','Optional',0),
    ('Team & Productivity','sales_manager_count','Sales Manager Count','AE managers','Optional',0),
    ('Team & Productivity','sdr_count','SDR / BDR Count','Sales development / business development reps','Optional',0),
    ('Team & Productivity','qcr_departed','QCRs Departed','Quota-carrying reps departed in period','Optional',0),
    ('Team & Productivity','quota_allocated','Quota Allocated','Total allocated quota in period','Required for attainment',0),
    ('Team & Productivity','quota_attained','Quota Attained','Total quota achieved in period','Required for attainment',0),
    ('Team & Productivity','qcrs_attaining_100','QCRs Attaining 100%+','Count of reps hitting quota','Optional',0),
    ('Team & Productivity','qcr_retention_rate','QCR Retention Rate','Percent of QCRs retained in period','Required for capacity if used',0),
    ('Team & Productivity','reps_first_full_ramp_period','Reps in First Fully Ramped Period','Denominator for ramp rate','Optional',0),
    ('Team & Productivity','reps_hit_quota_first_full_ramp','Reps Hitting Quota in First Fully Ramped Period','Numerator for ramp rate','Optional',0),
    ('Planning / Forecast','plan_net_new_bookings','Plan Net New Bookings','Budgeted net new bookings','Optional',0),
    ('Planning / Forecast','plan_gross_new_bookings','Plan Gross New Bookings','Budgeted gross new bookings','Optional',0),
    ('Planning / Forecast','plan_ending_arr','Plan Ending ARR','Budgeted ending ARR','Optional',0),
    ('Planning / Forecast','actual_yoy_arr_growth_pct','Actual YoY ARR Growth %','If actual growth already calculated','Optional',0),
    ('Planning / Forecast','plan_yoy_arr_growth_pct','Plan YoY ARR Growth %','Budgeted YoY ARR growth','Optional',0),
    ('Planning / Forecast','actual_fcf','Actual Free Cash Flow','Optional',0),
    ('Planning / Forecast','plan_fcf','Plan Free Cash Flow','Optional',0),
    ('Planning / Forecast','revenue_guidance','Revenue Guidance','Management guidance','Optional',0),
    ('Planning / Forecast','consensus_estimate','Consensus Estimate','Public-company context only','Optional',0),
]

metrics = [
    ('Growth Drivers','gross_new_arr','Gross New ARR','Derived','New Logo ARR + Expansion ARR','new_logo_arr, expansion_arr', '=IF(COUNTIF(Input_Fields!$B:$B,"new_logo_arr")*COUNTIF(Input_Fields!$B:$B,"expansion_arr")=0,"",XLOOKUP("new_logo_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")+XLOOKUP("expansion_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,""))'),
    ('Growth Drivers','churned_arr','Churned ARR','Derived','Downsell ARR + Logo Churn ARR','downsell_arr, logo_churn_arr', '=IF(COUNTIF(Input_Fields!$B:$B,"downsell_arr")*COUNTIF(Input_Fields!$B:$B,"logo_churn_arr")=0,"",XLOOKUP("downsell_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")+XLOOKUP("logo_churn_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,""))'),
    ('Growth Drivers','net_new_arr','Net New ARR','Derived','Gross New ARR - Churned ARR','new_logo_arr, expansion_arr, downsell_arr, logo_churn_arr', '=IFERROR(G2-G3,"")'),
    ('Growth Drivers','ending_arr_calc','Ending ARR (Calculated)','Derived','Beginning ARR + Net New ARR','beginning_arr, new_logo_arr, expansion_arr, downsell_arr, logo_churn_arr', '=IFERROR(XLOOKUP("beginning_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")+G4,"")'),
    ('Growth Drivers','arr_growth_yoy','ARR Growth YoY %','Derived','((EOP ARR / EOP ARR 12 months ago) - 1) * 100','Requires prior-year EOP ARR upload', ''),
    ('Growth Drivers','new_logo_mix','% Gross New ARR from New Logo','Derived','New Logo ARR / Gross New ARR','new_logo_arr, expansion_arr', '=IFERROR(XLOOKUP("new_logo_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/G2,"")'),
    ('Growth Drivers','expansion_mix','% Gross New ARR from Expansion','Derived','Expansion ARR / Gross New ARR','new_logo_arr, expansion_arr', '=IFERROR(XLOOKUP("expansion_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/G2,"")'),
    ('Sales Funnel','pipeline_coverage','Pipeline Coverage','Derived','Pipeline $ / Sales Target','pipeline_dollars, sales_target', '=IFERROR(XLOOKUP("pipeline_dollars",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("sales_target",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Sales Funnel','weighted_pipeline_coverage','Weighted Pipeline Coverage','Derived','Weighted Pipeline $ / Sales Target','weighted_pipeline_dollars, sales_target', '=IFERROR(XLOOKUP("weighted_pipeline_dollars",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("sales_target",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Sales Funnel','win_rate','Win Rate','Derived','Closed Won / (Closed Won + Closed Lost)','closed_won_opportunities, closed_lost_opportunities', '=IFERROR(XLOOKUP("closed_won_opportunities",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/(XLOOKUP("closed_won_opportunities",Input_Fields!$B:$B,Input_Fields!$F:$F,"")+XLOOKUP("closed_lost_opportunities",Input_Fields!$B:$B,Input_Fields!$F:$F,"")),"")'),
    ('Sales Funnel','close_rate','Close Rate','Derived','Closed Won in same period / Opportunities created in period','closed_won_same_period, opportunities_created', '=IFERROR(XLOOKUP("closed_won_same_period",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("opportunities_created",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Sales Funnel','sql_to_closed_won','SQL to Closed Won','Derived','Closed Won SQLs / SQL Count','closed_won_sql, sql_count', '=IFERROR(XLOOKUP("closed_won_sql",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("sql_count",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Sales Funnel','forecast_pct_target','Forecast as % of Sales Target','Derived','Forecast $ / Sales Target','forecast_dollars, sales_target', '=IFERROR(XLOOKUP("forecast_dollars",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("sales_target",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Retention','quarterly_ndr','Quarterly Annualized NDR','Derived','1 + ((Quarter Expansion ARR - Quarter Churned ARR) * 4) / avg(BOQ ARR + EOQ ARR)','beginning_arr, ending_arr or calculated ending arr, expansion_arr, downsell_arr, logo_churn_arr', '=IFERROR(1+(((XLOOKUP("expansion_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")-(XLOOKUP("downsell_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")+XLOOKUP("logo_churn_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")))*4)/AVERAGE(XLOOKUP("beginning_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,""),IF(XLOOKUP("ending_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")="",G5,XLOOKUP("ending_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")))),"")'),
    ('Retention','quarterly_gdr','Quarterly Annualized GDR','Derived','1 - (Quarter Gross Churned ARR * 4) / avg(BOQ ARR + EOQ ARR)','beginning_arr, ending_arr or calculated ending arr, downsell_arr, logo_churn_arr', '=IFERROR(1-((((XLOOKUP("downsell_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")+XLOOKUP("logo_churn_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,""))*4)/AVERAGE(XLOOKUP("beginning_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,""),IF(XLOOKUP("ending_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")="",G5,XLOOKUP("ending_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")))),"")'),
    ('Retention','logo_retention','Logo Retention','Derived','1 - Churned Customers / avg(BOP Customers + EOP Customers)','bop_customers, eop_customers, churned_customers', '=IFERROR(1-(XLOOKUP("churned_customers",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/AVERAGE(XLOOKUP("bop_customers",Input_Fields!$B:$B,Input_Fields!$F:$F,""),XLOOKUP("eop_customers",Input_Fields!$B:$B,Input_Fields!$F:$F,""))),"")'),
    ('Retention','logo_churn','Logo Churn','Derived','Churned Customers / avg(BOP Customers + EOP Customers)','bop_customers, eop_customers, churned_customers', '=IFERROR(XLOOKUP("churned_customers",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/AVERAGE(XLOOKUP("bop_customers",Input_Fields!$B:$B,Input_Fields!$F:$F,""),XLOOKUP("eop_customers",Input_Fields!$B:$B,Input_Fields!$F:$F,"")),"")'),
    ('Retention','dau_rate','DAU Rate','Derived','DAU / Total Users','dau, total_users', '=IFERROR(XLOOKUP("dau",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("total_users",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Retention','mau_rate','MAU Rate','Derived','MAU / Total Users','mau, total_users', '=IFERROR(XLOOKUP("mau",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("total_users",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Retention','adoption_rate','Adoption Rate','Derived','Feature Active Users / Total Users','feature_active_users, total_users', '=IFERROR(XLOOKUP("feature_active_users",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("total_users",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Retention','stickiness_rate','Stickiness Rate','Derived','DAU / MAU','dau, mau', '=IFERROR(XLOOKUP("dau",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("mau",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Retention','time_to_implement_vs_goal','Time to Implement vs Goal','Derived','Actual Implementation Days / Expected Implementation Days','actual_implementation_days, expected_implementation_days', '=IFERROR(XLOOKUP("actual_implementation_days",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("expected_implementation_days",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Retention','nps','NPS','Derived','Promoter % - Detractor %','promoter_pct, detractor_pct', '=IFERROR(XLOOKUP("promoter_pct",Input_Fields!$B:$B,Input_Fields!$F:$F,"")-XLOOKUP("detractor_pct",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Retention','csat','CSAT','Derived','Satisfied Responses / Total Responses','satisfied_responses, total_responses', '=IFERROR(XLOOKUP("satisfied_responses",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("total_responses",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Efficiency & Economics','gross_margin','Gross Margin','Derived','Gross Profit / Total Revenue','gross_profit, total_revenue', '=IFERROR(XLOOKUP("gross_profit",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("total_revenue",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Efficiency & Economics','subscription_gm','Subscription Gross Margin','Derived','Subscription Gross Profit / Subscription Revenue','subscription_gross_profit, subscription_revenue', '=IFERROR(XLOOKUP("subscription_gross_profit",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("subscription_revenue",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Efficiency & Economics','services_gm','Services Gross Margin','Derived','Services Gross Profit / Services Revenue','services_gross_profit, services_revenue', '=IFERROR(XLOOKUP("services_gross_profit",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("services_revenue",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Efficiency & Economics','net_magic_number','Net Magic Number','Derived','Current Quarter Net New ARR / Prior Quarter S&M OpEx','new_logo_arr, expansion_arr, downsell_arr, logo_churn_arr, prior_qtr_sm_opex', '=IFERROR(G4/XLOOKUP("prior_qtr_sm_opex",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Efficiency & Economics','gross_magic_number','Gross Magic Number','Derived','Current Quarter Gross New ARR / Prior Quarter S&M OpEx','new_logo_arr, expansion_arr, prior_qtr_sm_opex', '=IFERROR(G2/XLOOKUP("prior_qtr_sm_opex",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Efficiency & Economics','simple_cac','Simple CAC','Derived','S&M OpEx / Gross New Logo Customers','sm_opex, gross_new_logo_customers', '=IFERROR(XLOOKUP("sm_opex",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("gross_new_logo_customers",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Efficiency & Economics','time_adjusted_cac','Time-Adjusted CAC','Derived','Prior Quarter S&M OpEx / Gross New Logo Customers This Quarter','prior_qtr_sm_opex, gross_new_logo_customers', '=IFERROR(XLOOKUP("prior_qtr_sm_opex",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("gross_new_logo_customers",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Efficiency & Economics','simple_ltv','Simple LTV','Derived','(ARR per Customer * Gross Margin) / Logo Churn Rate','avg_arr_per_customer, gross_margin or gross_margin_pct, logo_churn or provided logo_churn_rate', '=IFERROR((XLOOKUP("avg_arr_per_customer",Input_Fields!$B:$B,Input_Fields!$F:$F,"")*IF(XLOOKUP("gross_margin_pct",Input_Fields!$B:$B,Input_Fields!$F:$F,"")="",G26,XLOOKUP("gross_margin_pct",Input_Fields!$B:$B,Input_Fields!$F:$F,"")))/IF(XLOOKUP("logo_churn_rate",Input_Fields!$B:$B,Input_Fields!$F:$F,"")="",G18,XLOOKUP("logo_churn_rate",Input_Fields!$B:$B,Input_Fields!$F:$F,"")),"")'),
    ('Efficiency & Economics','ltv_cac','LTV / CAC','Derived','Customer LTV / Customer Acquisition Cost','avg_arr_per_customer, margin, churn, sm_opex, gross_new_logo_customers', '=IFERROR(G33/G31,"")'),
    ('Efficiency & Economics','payback_period','Payback Period','Derived','CAC / (ARR per Customer * Gross Margin)','avg_arr_per_customer, margin, sm_opex, gross_new_logo_customers', '=IFERROR(G31/(XLOOKUP("avg_arr_per_customer",Input_Fields!$B:$B,Input_Fields!$F:$F,"")*IF(XLOOKUP("gross_margin_pct",Input_Fields!$B:$B,Input_Fields!$F:$F,"")="",G26,XLOOKUP("gross_margin_pct",Input_Fields!$B:$B,Input_Fields!$F:$F,""))),"")'),
    ('Team & Productivity','sm_opex_per_fte','S&M OpEx per S&M FTE','Derived','(Quarter S&M OpEx * 4) / Average S&M FTEs','sm_opex, avg_sm_ftes', '=IFERROR((XLOOKUP("sm_opex",Input_Fields!$B:$B,Input_Fields!$F:$F,"")*4)/XLOOKUP("avg_sm_ftes",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Team & Productivity','ae_per_csm','AE per CSM','Derived','Quota-Carrying AEs / CSMs','qcr_count, csm_count', '=IFERROR(XLOOKUP("qcr_count",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("csm_count",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Team & Productivity','ae_per_manager','AE per Sales Manager','Derived','Quota-Carrying AEs / Sales Managers','qcr_count, sales_manager_count', '=IFERROR(XLOOKUP("qcr_count",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("sales_manager_count",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Team & Productivity','ae_per_sdr','AE per SDR','Derived','Quota-Carrying AEs / SDRs','qcr_count, sdr_count', '=IFERROR(XLOOKUP("qcr_count",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("sdr_count",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Team & Productivity','employee_attrition','Employee Attrition','Derived','QCRs Departed / Average QCRs in Period','qcr_departed, qcr_count', '=IFERROR(XLOOKUP("qcr_departed",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("qcr_count",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Team & Productivity','nnarr_per_sm_fte','NNARR per S&M FTE','Derived','(Net New ARR * 4) / Average S&M FTEs','new_logo_arr, expansion_arr, downsell_arr, logo_churn_arr, avg_sm_ftes', '=IFERROR((G4*4)/XLOOKUP("avg_sm_ftes",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Team & Productivity','sales_capacity','Sales Capacity','Derived','Allocated Quota * Quota Attainment % * QCR Retention Rate','quota_allocated, quota_attained, qcr_retention_rate', '=IFERROR(XLOOKUP("quota_allocated",Input_Fields!$B:$B,Input_Fields!$F:$F,"")*(XLOOKUP("quota_attained",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("quota_allocated",Input_Fields!$B:$B,Input_Fields!$F:$F,""))*XLOOKUP("qcr_retention_rate",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Team & Productivity','capacity_per_qcr','Capacity per QCR','Derived','Sales Capacity / QCR Count','quota_allocated, quota_attained, qcr_retention_rate, qcr_count', '=IFERROR(G42/XLOOKUP("qcr_count",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Team & Productivity','net_productivity','Net Productivity','Derived','Net New ARR / Average QCRs','new_logo_arr, expansion_arr, downsell_arr, logo_churn_arr, qcr_count', '=IFERROR(G4/XLOOKUP("qcr_count",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Team & Productivity','team_quota_attainment','Team Quota Attainment','Derived','QCRs attaining 100%+ / total QCRs','qcrs_attaining_100, qcr_count', '=IFERROR(XLOOKUP("qcrs_attaining_100",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("qcr_count",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Team & Productivity','rep_attainment','Rep Attainment','Derived','Quota Attained / Quota Allocated','quota_attained, quota_allocated', '=IFERROR(XLOOKUP("quota_attained",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("quota_allocated",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Team & Productivity','ramp_rate','Ramp Rate','Derived','Reps hitting quota in first fully ramped period / reps in first fully ramped period','reps_hit_quota_first_full_ramp, reps_first_full_ramp_period', '=IFERROR(XLOOKUP("reps_hit_quota_first_full_ramp",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("reps_first_full_ramp_period",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Fiscal Maturity','net_new_bookings_attainment','Net New Bookings Attainment','Derived','Actual Net New Bookings / Plan Net New Bookings','net new ARR approximation or actual net bookings, plan_net_new_bookings', '=IFERROR(G4/XLOOKUP("plan_net_new_bookings",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Fiscal Maturity','gross_new_bookings_attainment','Gross New Bookings Attainment','Derived','Actual Gross New Bookings / Plan Gross New Bookings','gross_new_arr, plan_gross_new_bookings', '=IFERROR(G2/XLOOKUP("plan_gross_new_bookings",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Fiscal Maturity','ending_arr_attainment','Ending ARR Attainment','Derived','Actual Ending ARR / Plan Ending ARR','ending arr or calculated ending arr, plan_ending_arr', '=IFERROR(IF(XLOOKUP("ending_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,"")="",G5,XLOOKUP("ending_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,""))/XLOOKUP("plan_ending_arr",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Fiscal Maturity','yoy_arr_growth_attainment','YoY ARR Growth Attainment','Derived','Actual YoY ARR Growth / Plan YoY ARR Growth','actual_yoy_arr_growth_pct, plan_yoy_arr_growth_pct', '=IFERROR(XLOOKUP("actual_yoy_arr_growth_pct",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("plan_yoy_arr_growth_pct",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Fiscal Maturity','fcf_attainment','Free Cash Flow Attainment','Derived','Actual FCF / Plan FCF','actual_fcf, plan_fcf', '=IFERROR(XLOOKUP("actual_fcf",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("plan_fcf",Input_Fields!$B:$B,Input_Fields!$F:$F,""),"")'),
    ('Fiscal Maturity','beat_vs_guidance','Beat Against Revenue Guidance','Derived','(Actual Revenue / Revenue Guidance) - 1','revenue, revenue_guidance', '=IFERROR((XLOOKUP("revenue",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("revenue_guidance",Input_Fields!$B:$B,Input_Fields!$F:$F,""))-1,"")'),
    ('Fiscal Maturity','beat_vs_consensus','Beat Against Revenue Consensus','Derived','(Actual Revenue / Consensus Estimate) - 1','revenue, consensus_estimate', '=IFERROR((XLOOKUP("revenue",Input_Fields!$B:$B,Input_Fields!$F:$F,"")/XLOOKUP("consensus_estimate",Input_Fields!$B:$B,Input_Fields!$F:$F,""))-1,"")'),
]

# ── Number format helper ──────────────────────────────────────────────────
_PCT_KEYS  = {'rate','retention','attainment','margin','coverage','pct','mix','churn','stickiness','adoption','csat','beat','_gdr','_ndr','_mix'}
_DOLLAR_KEYS = {'arr','bookings','cac','ltv','capacity','opex_per','nnarr_per','productivity'}
_RATIO_KEYS  = {'magic_number','ltv_cac','ae_per','payback'}

def num_fmt(key):
    k = key.lower()
    if any(p in k for p in _PCT_KEYS):   return '0.0%'
    if any(p in k for p in _DOLLAR_KEYS): return '$#,##0'
    if any(p in k for p in _RATIO_KEYS):  return '0.00'
    return '#,##0.0'

# ── Workbook ──────────────────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════
#  README
# ═══════════════════════════════════════════════════════
ws = wb.active
ws.title = 'README'
ws.sheet_properties.tabColor = LT_BLUE[1:] if LT_BLUE.startswith('#') else LT_BLUE

ws.column_dimensions['A'].width = 90
ws.column_dimensions['B'].width = 20

def readme_row(r, height=18):
    ws.row_dimensions[r].height = height

# Title block
ws.merge_cells('A1:B1')
c = ws['A1']
c.value = 'GTM Metrics Analyzer'
c.font = Font(name='Wix Madefor Display SemiBold', size=16, bold=True, color=NAVY)
c.alignment = LEFT_MID
readme_row(1, 28)

ws.merge_cells('A2:B2')
c = ws['A2']
c.value = 'Input Template + Metric Calculator  ·  B2B SaaS / GTM-Driven Businesses'
c.font = Font(name='Wix Madefor Display', size=9, color=GRAY)
c.alignment = LEFT_MID
readme_row(2, 18)

readme_row(3, 8)

# Section: Purpose
ws['A4'].value = 'Purpose'
ws['A4'].font = Font(name='Wix Madefor Display SemiBold', size=10, bold=True, color=BLUE)
ws['A4'].alignment = LEFT_MID
readme_row(4, 20)

ws.merge_cells('A5:B5')
ws['A5'].value = (
    'Use this workbook to load raw company data into Input_Fields, then review '
    'derived GTM metric calculations in Metric_Calcs, and surface a clean diagnostic '
    'summary in Diagnostic_Output. Supports all six GTM metric families: Growth Drivers, '
    'Sales Funnel, Retention, Efficiency & Economics, Team & Productivity, and Fiscal Maturity.'
)
ws['A5'].font = Font(name='Wix Madefor Display', size=9, color=BLACK)
ws['A5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.row_dimensions[5].height = 40

readme_row(6, 8)

# Section: How to Use
ws['A7'].value = 'How to Use'
ws['A7'].font = Font(name='Wix Madefor Display SemiBold', size=10, bold=True, color=BLUE)
ws['A7'].alignment = LEFT_MID
readme_row(7, 20)

steps = [
    '1.  Input_Fields    →  Populate the User Value column (column F) with values from uploaded source files. Do not overwrite raw exports — add them as separate tabs if needed.',
    '2.  Metric_Calcs    →  Review formulas, required inputs, and calculated values. Column G auto-calculates from Input_Fields.',
    '3.  Diagnostic_Output  →  Clean summary of the 20 most decision-useful GTM metrics, pulled automatically from Metric_Calcs.',
    '4.  Do not modify column B (Input Key) or column B of Metric_Calcs (Metric Key) — these are the XLOOKUP lookup arrays.',
]
for i, step in enumerate(steps, start=8):
    ws.merge_cells(f'A{i}:B{i}')
    ws[f'A{i}'].value = step
    ws[f'A{i}'].font = Font(name='Wix Madefor Display', size=9, color=BLACK)
    ws[f'A{i}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[i].height = 30

readme_row(12, 8)

# Section: Requirements
ws['A13'].value = 'Requirements'
ws['A13'].font = Font(name='Wix Madefor Display SemiBold', size=10, bold=True, color=BLUE)
ws['A13'].alignment = LEFT_MID
readme_row(13, 20)

ws.merge_cells('A14:B14')
ws['A14'].value = (
    '⚠️  Requires Excel 365 or Excel 2021+. Formulas use XLOOKUP and FILTER, which are '
    'not available in older Excel versions or Google Sheets without modification.'
)
ws['A14'].font = Font(name='Wix Madefor Display', size=9, color='7F3F00', italic=True)
ws['A14'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.row_dimensions[14].height = 28

readme_row(15, 8)

# Section: Metric Families
ws['A16'].value = 'Metric Families'
ws['A16'].font = Font(name='Wix Madefor Display SemiBold', size=10, bold=True, color=BLUE)
ws['A16'].alignment = LEFT_MID
readme_row(16, 20)

families = [
    ('Growth Drivers',       '7 metrics  ·  Gross/Net New ARR, ARR Growth, Mix'),
    ('Sales Funnel',         '6 metrics  ·  Pipeline Coverage, Win Rate, Close Rate, Forecast'),
    ('Retention',            '10 metrics  ·  NDR, GDR, Logo Retention, Adoption, NPS, CSAT'),
    ('Efficiency & Economics', '9 metrics  ·  Gross Margin, Magic Number, CAC, LTV, Payback'),
    ('Team & Productivity',  '10 metrics  ·  Sales Capacity, Attainment, Ramp Rate, FTE Productivity'),
    ('Fiscal Maturity',      '6 metrics  ·  Plan Attainment, YoY Growth, FCF, Beat vs Guidance'),
]
for i, (fam, desc) in enumerate(families, start=17):
    ws[f'A{i}'].value = fam
    ws[f'A{i}'].font = Font(name='Wix Madefor Display SemiBold', size=9, bold=True, color=NAVY)
    ws[f'A{i}'].alignment = LEFT_MID
    ws[f'B{i}'].value = desc
    ws[f'B{i}'].font = Font(name='Wix Madefor Display', size=9, color=GRAY)
    ws[f'B{i}'].alignment = LEFT_MID
    ws.column_dimensions['B'].width = 52
    ws.row_dimensions[i].height = 18

# ═══════════════════════════════════════════════════════
#  INPUT_FIELDS
# ═══════════════════════════════════════════════════════
inp = wb.create_sheet('Input_Fields')
inp.sheet_properties.tabColor = NAVY
inp.freeze_panes = 'A2'

inp_headers = ['Section', 'Input Key', 'Input Label', 'Definition / What to Upload', 'Input Requirement', 'User Value', 'Notes']
inp_col_widths = {'A': 24, 'B': 28, 'C': 28, 'D': 44, 'E': 22, 'F': 18, 'G': 24}

# Header row
for ci, hdr in enumerate(inp_headers, start=1):
    c = inp.cell(row=1, column=ci, value=hdr)
    c.fill = fill(NAVY)
    c.font = Font(name='Wix Madefor Display SemiBold', size=9, bold=True, color=WHITE)
    c.border = BDR
    c.alignment = CTR_WRAP
inp.row_dimensions[1].height = 22

# Data rows
prev_section = None
alt_counter = 0
for ri, row_data in enumerate(input_rows, start=2):
    section = row_data[0]
    is_section_start = (section != prev_section)
    if is_section_start:
        alt_counter = 0
    else:
        alt_counter += 1

    is_alt = (alt_counter % 2 == 1)
    base_fill = fill(ALT_ROW) if is_alt else fill(WHITE)
    sect_fill = fill(LT_BLUE)

    for ci, val in enumerate(row_data, start=1):
        c = inp.cell(row=ri, column=ci, value=val)
        c.border = BDR
        c.alignment = WRAP_TOP

        if is_section_start:
            c.fill = sect_fill
            c.font = Font(name='Wix Madefor Display SemiBold', size=9, bold=True, color=NAVY)
        elif ci == 6:   # User Value column — light green tint
            c.fill = fill(INPUT_V)
            c.font = Font(name='Wix Madefor Display', size=9, color=BLACK)
        else:
            c.fill = base_fill
            c.font = Font(name='Wix Madefor Display', size=9, color=BLACK)

    inp.row_dimensions[ri].height = 18
    prev_section = section

# User Value header extra treatment
inp.cell(1, 6).fill = fill('1A6B2C')   # darker green for "fill here" column header
inp.cell(1, 6).font = Font(name='Wix Madefor Display SemiBold', size=9, bold=True, color=WHITE)

for col, w in inp_col_widths.items():
    inp.column_dimensions[col].width = w

inp.auto_filter.ref = f'A1:G{inp.max_row}'

# ═══════════════════════════════════════════════════════
#  METRIC_CALCS
# ═══════════════════════════════════════════════════════
calc = wb.create_sheet('Metric_Calcs')
calc.sheet_properties.tabColor = BLUE
calc.freeze_panes = 'A2'

calc_headers = ['Metric Family', 'Metric Key', 'Metric Name', 'Type', 'Formula / Logic', 'Required Inputs', 'Calculated Value', 'Comments']
calc_col_widths = {'A': 22, 'B': 24, 'C': 30, 'D': 10, 'E': 38, 'F': 38, 'G': 18, 'H': 24}

# Header row
for ci, hdr in enumerate(calc_headers, start=1):
    c = calc.cell(row=1, column=ci, value=hdr)
    c.fill = fill(NAVY)
    c.font = Font(name='Wix Madefor Display SemiBold', size=9, bold=True, color=WHITE)
    c.border = BDR
    c.alignment = CTR_WRAP
calc.row_dimensions[1].height = 22

# Data rows
prev_family = None
alt_counter = 0
for ri, row_data in enumerate(metrics, start=2):
    family = row_data[0]
    is_family_start = (family != prev_family)
    if is_family_start:
        alt_counter = 0
    else:
        alt_counter += 1

    is_alt = (alt_counter % 2 == 1)
    base_fill = fill(ALT_ROW) if is_alt else fill(WHITE)

    row_vals = list(row_data) + ['']
    for ci, val in enumerate(row_vals, start=1):
        c = calc.cell(row=ri, column=ci, value=val)
        c.border = BDR
        c.alignment = WRAP_TOP

        if is_family_start:
            c.fill = fill(LT_BLUE)
            c.font = Font(name='Wix Madefor Display SemiBold', size=9, bold=True, color=NAVY)
        else:
            c.fill = base_fill
            c.font = Font(name='Wix Madefor Display', size=9, color=BLACK)

    # Number format on Calculated Value column (G = col 7)
    key = row_data[1]
    calc.cell(ri, 7).number_format = num_fmt(key)
    calc.row_dimensions[ri].height = 18
    prev_family = family

for col, w in calc_col_widths.items():
    calc.column_dimensions[col].width = w

calc.auto_filter.ref = f'A1:H{calc.max_row}'

# ═══════════════════════════════════════════════════════
#  DIAGNOSTIC_OUTPUT
# ═══════════════════════════════════════════════════════
out = wb.create_sheet('Diagnostic_Output')
out.sheet_properties.tabColor = NAVY

out.column_dimensions['A'].width = 32
out.column_dimensions['B'].width = 18
out.column_dimensions['C'].width = 24
out.column_dimensions['D'].width = 4
out.column_dimensions['E'].width = 38

# Title block
out.merge_cells('A1:C1')
out['A1'].value = 'GTM Diagnostic Output'
out['A1'].font = Font(name='Wix Madefor Display SemiBold', size=14, bold=True, color=NAVY)
out['A1'].alignment = LEFT_MID
out.row_dimensions[1].height = 28

out.merge_cells('A2:C2')
out['A2'].value = 'Auto-calculated from Input_Fields. Populate the User Value column in Input_Fields to populate this summary.'
out['A2'].font = Font(name='Wix Madefor Display', size=9, color=GRAY, italic=True)
out['A2'].alignment = LEFT_MID
out.row_dimensions[2].height = 18
out.row_dimensions[3].height = 10

# Section labels + table headers (row 4 = section label, row 5 = headers)
out['A4'].value = 'Key Metrics'
out['A4'].font = Font(name='Wix Madefor Display SemiBold', size=10, bold=True, color=BLUE)
out['A4'].alignment = LEFT_MID
out['E4'].value = 'Data Quality'
out['E4'].font = Font(name='Wix Madefor Display SemiBold', size=10, bold=True, color=BLUE)
out['E4'].alignment = LEFT_MID
out.row_dimensions[4].height = 20

for ci, hdr in enumerate(['Metric', 'Value', 'Family'], start=1):
    c = out.cell(row=5, column=ci, value=hdr)
    c.fill = fill(NAVY)
    c.font = Font(name='Wix Madefor Display SemiBold', size=9, bold=True, color=WHITE)
    c.border = BDR
    c.alignment = CTR
out.row_dimensions[5].height = 22

out.cell(5, 5).value = 'Missing Required Inputs'
out.cell(5, 5).fill = fill(NAVY)
out.cell(5, 5).font = Font(name='Wix Madefor Display SemiBold', size=9, bold=True, color=WHITE)
out.cell(5, 5).border = BDR
out.cell(5, 5).alignment = CTR

# Summary metric rows
summary_metrics = [
    ('Gross New ARR',              'gross_new_arr',              'Growth Drivers'),
    ('Net New ARR',                'net_new_arr',                'Growth Drivers'),
    ('Ending ARR (Calculated)',    'ending_arr_calc',            'Growth Drivers'),
    ('Pipeline Coverage',         'pipeline_coverage',          'Sales Funnel'),
    ('Weighted Pipeline Coverage','weighted_pipeline_coverage',  'Sales Funnel'),
    ('Win Rate',                   'win_rate',                   'Sales Funnel'),
    ('Quarterly Annualized NDR',   'quarterly_ndr',              'Retention'),
    ('Quarterly Annualized GDR',   'quarterly_gdr',              'Retention'),
    ('Logo Retention',             'logo_retention',             'Retention'),
    ('Gross Margin',               'gross_margin',               'Efficiency & Economics'),
    ('Net Magic Number',           'net_magic_number',           'Efficiency & Economics'),
    ('Simple CAC',                 'simple_cac',                 'Efficiency & Economics'),
    ('Simple LTV',                 'simple_ltv',                 'Efficiency & Economics'),
    ('LTV / CAC',                  'ltv_cac',                    'Efficiency & Economics'),
    ('S&M OpEx per FTE',           'sm_opex_per_fte',            'Team & Productivity'),
    ('NNARR per S&M FTE',          'nnarr_per_sm_fte',           'Team & Productivity'),
    ('Sales Capacity',             'sales_capacity',             'Team & Productivity'),
    ('Rep Attainment',             'rep_attainment',             'Team & Productivity'),
    ('Ramp Rate',                  'ramp_rate',                  'Team & Productivity'),
    ('Ending ARR Attainment',      'ending_arr_attainment',      'Fiscal Maturity'),
]

prev_family = None
for i, (label, key, family) in enumerate(summary_metrics):
    r = i + 6
    is_alt = (i % 2 == 1)
    base = fill(ALT_ROW) if is_alt else fill(WHITE)

    # Metric label
    c_a = out.cell(row=r, column=1, value=label)
    c_a.fill = base
    c_a.font = Font(name='Wix Madefor Display', size=9, color=BLACK)
    c_a.border = BDR
    c_a.alignment = LEFT_MID

    # Value (XLOOKUP from Metric_Calcs)
    c_b = out.cell(row=r, column=2)
    c_b.value = f'=XLOOKUP("{key}",Metric_Calcs!$B:$B,Metric_Calcs!$G:$G,"")'
    c_b.fill = fill(LT_BLUE)   # tinted to distinguish calculated from inputs
    c_b.font = Font(name='Wix Madefor Display SemiBold', size=9, bold=True, color=NAVY)
    c_b.border = BDR
    c_b.alignment = CTR
    c_b.number_format = num_fmt(key)

    # Family tag
    c_c = out.cell(row=r, column=3, value=family)
    c_c.fill = base
    c_c.font = Font(name='Wix Madefor Display', size=9, color=GRAY)
    c_c.border = BDR
    c_c.alignment = LEFT_MID

    out.row_dimensions[r].height = 18

# Missing inputs panel — spans all data rows, right column
missing_start = 6
missing_end = 6 + len(summary_metrics) - 1
out.merge_cells(f'E{missing_start}:E{missing_end}')
c_e = out.cell(missing_start, 5)
c_e.value = (
    '=TEXTJOIN(CHAR(10),TRUE,'
    'FILTER(Input_Fields!$C$2:$C$999,'
    '(Input_Fields!$E$2:$E$999<>"Optional")'
    '*(Input_Fields!$E$2:$E$999<>"Optional / shortcut")'
    '*(Input_Fields!$E$2:$E$999<>"Optional / validation")'
    '*(Input_Fields!$F$2:$F$999=0)'
    ',"— All required inputs populated"))'
)
c_e.fill = fill(WHITE)
c_e.font = Font(name='Wix Madefor Display', size=9, color=BLACK)
c_e.border = BDR
c_e.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

wb.save(OUT)
print(f'Wrote {OUT}')
